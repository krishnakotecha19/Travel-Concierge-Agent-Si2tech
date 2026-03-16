"""
Travel Concierge Agent — main.py
Merged: flight.py + hotel.py + email_agent.py + app_ui.py
Round-trip support added (search_return_flights, build_flights_tfs, UI changes)
FIXES:
  1. Outbound Book button no longer crashes — arr_time newline stripped before URL build
  2. Excel export splits outbound vs return flights into dedicated columns
  3. DB insert correctly maps return_origin / return_destination / return_flight_no
"""

import os
import re
import sys
import io
import json
import uuid
import time
import base64
import urllib.parse
import webbrowser
import concurrent.futures
from math import radians, sin, cos, sqrt, atan2
from datetime import datetime, date, timedelta
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.image import MIMEImage


import requests
import msal
import streamlit as st
import streamlit.components.v1 as components
from PIL import Image
from groq import Groq
from dotenv import load_dotenv
from serpapi import GoogleSearch
from geopy.geocoders import Nominatim
import openrouteservice
import airportsdata

try:
    import pdfScrapper
except ImportError:
    pdfScrapper = None

try:
    from addtodatabase_v2 import add_flight_booking, add_hotel_booking, build_ui_inputs, fetch_flight_transactions, fetch_hotel_transactions
except ImportError:
    pass

load_dotenv(override=True)

_IATA_CACHE_FILE = os.path.join(os.path.dirname(__file__), "flight_iata_cache.json")

def _load_iata_cache():
    if os.path.exists(_IATA_CACHE_FILE):
        try:
            with open(_IATA_CACHE_FILE, "r") as f:
                return json.load(f)
        except:
            pass
    return {}

def _save_iata_cache(cache):
    try:
        with open(_IATA_CACHE_FILE, "w") as f:
            json.dump(cache, f, indent=2)
    except:
        pass

_flight_iata_cache = _load_iata_cache()

# ── Search history persistence ────────────────────────────────────────────────
_SEARCH_HISTORY_FILE = os.path.join(os.path.dirname(__file__), "search_history.json")

def _load_search_history() -> dict:
    if os.path.exists(_SEARCH_HISTORY_FILE):
        try:
            with open(_SEARCH_HISTORY_FILE, "r") as f:
                return json.load(f)
        except Exception:
            pass
    return {"origins": [], "destinations": []}

def _save_search_history(history: dict):
    try:
        with open(_SEARCH_HISTORY_FILE, "w") as f:
            json.dump(history, f, indent=2)
    except Exception:
        pass

def _add_to_search_history(field: str, value: str):
    """Add a city name to the search history (origins or destinations list)."""
    if not value or not value.strip():
        return
    value = value.strip()
    history = _load_search_history()
    lst = history.get(field, [])
    # Remove duplicates (case-insensitive) and re-add at top
    lst = [v for v in lst if v.lower() != value.lower()]
    lst.insert(0, value)
    lst = lst[:20]  # keep last 20
    history[field] = lst
    _save_search_history(history)

# ── Fuzzy IATA resolution (local airportsdata first, then LLM) ───────────────

def _fuzzy_match_airport(query: str) -> str:
    """
    Try to match a city/airport name against airportsdata using fuzzy substring
    matching. Returns IATA code or empty string.
    Handles misspellings by checking if query is close enough to a known city.
    """
    if not query:
        return ""
    q = query.strip().lower()

    # 1. Exact IATA code (user typed "BOM", "DEL", etc.)
    q_upper = q.upper()
    if len(q_upper) == 3 and q_upper in airport_db:
        return q_upper

    # 2. Exact city name match
    for code, data in airport_db.items():
        city = (data.get("city") or "").lower()
        name = (data.get("name") or "").lower()
        if q == city or q == name:
            return code

    # 3. Fuzzy: find best substring / edit-distance match
    best_code, best_score = "", 0
    for code, data in airport_db.items():
        city = (data.get("city") or "").lower()
        if not city:
            continue
        # Substring match (query inside city or city inside query)
        if q in city or city in q:
            score = len(city)  # longer match = better
            if score > best_score:
                best_score, best_code = score, code
                continue
        # Simple character-overlap ratio for typo tolerance
        if len(q) >= 3 and len(city) >= 3:
            common = sum(1 for c in q if c in city)
            ratio = common / max(len(q), len(city))
            # Also check if first 2 chars match (strong signal)
            prefix_match = q[:2] == city[:2]
            if ratio >= 0.7 and prefix_match and len(city) > best_score:
                best_score, best_code = len(city), code

    if best_code:
        return best_code
    return ""


def get_flight_iata_from_city(city_name: str) -> str:
    if not city_name:
        return ""
    cache_key = city_name.strip().upper()
    if cache_key in _flight_iata_cache:
        return _flight_iata_cache[cache_key]

    # Try local fuzzy match first (no LLM call needed)
    fuzzy_result = _fuzzy_match_airport(city_name)
    if fuzzy_result:
        _flight_iata_cache[cache_key] = fuzzy_result
        _save_iata_cache(_flight_iata_cache)
        print(f"  ✅ Fuzzy IATA match: '{city_name}' → {fuzzy_result}")
        return fuzzy_result

    # Fallback to LLM
    api_key = os.getenv("GROQ_API_KEY")
    if not api_key:
        return ""
    try:
        client = Groq(api_key=api_key)
    except Exception:
        return ""
    prompt = f"You are a travel logic system.\nGiven a city name, return ONLY its primary 3-letter IATA airport code.\nInput:\n- City: {city_name}\nRespond with EXACTLY 3 uppercase letters, no explanation."
    try:
        completion = client.chat.completions.create(
            model="llama-3.3-70b-versatile",
            messages=[{"role": "user", "content": prompt}],
            temperature=0,
            max_tokens=10
        )
        iata = completion.choices[0].message.content.strip()
        if len(iata) >= 3:
            iata = iata[:3].upper()
            _flight_iata_cache[cache_key] = iata
            _save_iata_cache(_flight_iata_cache)
            return iata
    except Exception as e:
        print(f"Groq Flight IATA error: {e}")
    return ""


# ==============================================================================
#  FLIGHT ENGINE
# ==============================================================================

geolocator = Nominatim(user_agent="ai_flight_agent_v1")
airport_db = airportsdata.load('IATA')

def get_clean_key():
    key = os.getenv("ORS_API_KEY")
    if key:
        return key.strip().strip('"').strip("'")
    return None

_flight_key_index = 0

def get_serp_results_with_fallback(params):
    global _flight_key_index
    all_keys = [
        os.getenv("SERP_API_1"), os.getenv("SERP_API_2"), os.getenv("SERP_API_3"),
        os.getenv("SERP_API_4"), os.getenv("SERP_API_5"), os.getenv("SERP_API_6"),
        os.getenv("SERP_API_7"), os.getenv("SERP_API_8"), os.getenv("SERP_API_9")
    ]
    keys = [k for k in all_keys if k]
    if not keys:
        print("❌ No SerpApi keys configured.")
        return {}
    n = len(keys)
    start = _flight_key_index % n
    _flight_key_index = (start + 1) % n
    for i in range(n):
        key = keys[(start + i) % n]
        key_num = all_keys.index(key) + 1 if key in all_keys else "?"
        try:
            params["api_key"] = key
            results = GoogleSearch(params).get_dict()
            if "error" in results:
                print(f"⚠️ Key {key_num} Quota Error: {results['error']} — trying next…")
                continue
            print(f"  ✅ Key {key_num} used")
            return results
        except Exception as e:
            print(f"⚠️ Key {key_num} Connection Error: {e} — trying next…")
            continue
    print("❌ Critical: All SerpApi keys exhausted.")
    return {}

cleaned_key = get_clean_key()
ors_client = openrouteservice.Client(key=cleaned_key) if cleaned_key else None


def _get_country(iata_code: str) -> str:
    info = airport_db.get(iata_code.upper().strip())
    if info:
        return info.get("country", "").upper()
    return ""

def is_international(origin: str, destination: str) -> bool:
    orig_country = _get_country(origin)
    dest_country = _get_country(destination)
    if not orig_country or not dest_country:
        return False
    return orig_country != dest_country


# ── Protobuf primitives ────────────────────────────────────────────────────────

def _varint(v: int) -> bytes:
    v = int(v) & 0xFFFFFFFFFFFFFFFF
    out = []
    while v > 0x7F:
        out.append((v & 0x7F) | 0x80)
        v >>= 7
    out.append(v & 0x7F)
    return bytes(out)

def _vfield(f: int, v: int) -> bytes:
    return _varint((f << 3) | 0) + _varint(v)

def _lfield(f: int, data: bytes) -> bytes:
    return _varint((f << 3) | 2) + _varint(len(data)) + data

def _airport(iata: str) -> bytes:
    return _vfield(1, 1) + _lfield(2, iata.strip().upper().encode())


# ── TFS Builder — supports both one-way AND round-trip ────────────────────────

def build_flights_tfs(
    origin: str,
    destination: str,
    date_str: str,
    travel_count: int,
    airline_code: str = None,
    dep_start_hour: int = None,
    dep_end_hour: int = None,
    arr_cutoff_hour: int = None,
    max_price_per_person: int = None,
    max_stops: int = None,
    trip_type: int = 2,           # 2=one-way  1=round-trip
    return_date_str: str = None,  # only used when trip_type=1
) -> str:
    adults  = max(1, int(travel_count))
    max_u64 = (1 << 64) - 1

    def _leg(orig, dest, ds, dep_start=None, dep_end=None, arr_cut=None):
        leg = _lfield(2, ds.encode())
        if max_stops is not None:
            leg += _vfield(3, int(max_stops))
        if airline_code and len(airline_code) == 2:
            leg += _lfield(6, airline_code.strip().upper().encode())
        if dep_start is not None:
            leg += _vfield(8, int(dep_start))
        if dep_end is not None:
            leg += _vfield(9, int(dep_end))
        leg += _vfield(10, 0)
        if arr_cut is not None:
            leg += _vfield(11, int(arr_cut))
        leg += _lfield(13, _airport(orig))
        leg += _lfield(14, _airport(dest))
        return leg

    time_filter  = _vfield(1, max_u64)
    outbound_leg = _leg(origin, destination, date_str,
                        dep_start=dep_start_hour, dep_end=dep_end_hour, arr_cut=arr_cutoff_hour)

    proto = (
        _vfield(1, 28)
        + _vfield(2, trip_type)
        + _lfield(3, outbound_leg)
    )

    # Round-trip: add return leg
    if trip_type == 1 and return_date_str:
        return_leg = _leg(destination, origin, return_date_str)
        proto += _lfield(3, return_leg)

    for _ in range(adults):
        proto += _vfield(8, 1)
    proto += (
        _vfield(9, 1)
        + _vfield(14, 1)
        + _lfield(16, time_filter)
        + _vfield(19, 2)
    )
    if max_price_per_person is not None:
        proto += _vfield(26, int(max_price_per_person))

    return base64.urlsafe_b64encode(proto).decode().rstrip("=")


# Keep old name as alias so any existing callers don't break
def build_flights_tfs_oneway(*args, **kwargs):
    return build_flights_tfs(*args, **kwargs)


def _clean_time_str(time_str: str) -> str:
    """Strip newlines and markdown annotations from arrival/departure time strings."""
    if not time_str:
        return ""
    return time_str.split('\n')[0].strip()


def make_google_flights_link(
    origin: str,
    destination: str,
    travel_date: str,
    airline_name: str,
    travel_count: int,
    dep_time_str: str = None,
    arr_time_str: str = None,
    price_per_person: int = None,
    _override_airline_code: str = None,
    round_trip: bool = False,
    return_date: str = None,
    **kwargs
) -> str:
    """
    Build a Google Flights search URL.
    NEVER raises — returns a safe fallback URL on any error so Book buttons always work.
    Outbound flights must always pass round_trip=False / return_date=None.
    """
    try:
        dep_time_clean = _clean_time_str(dep_time_str)
        arr_time_clean = _clean_time_str(arr_time_str)
        airline_code = _override_airline_code if _override_airline_code else get_airline_code(airline_name or "")
        dep_hour   = _parse_hour(dep_time_clean)
        arr_hour   = _parse_hour(arr_time_clean)
        dep_start  = max(0,  dep_hour - 1) if dep_hour is not None else None
        dep_end    = min(23, dep_hour + 1) if dep_hour is not None else None
        arr_cutoff = min(23, arr_hour + 1) if arr_hour is not None else None
        max_price  = int(price_per_person * 1.10) if price_per_person else None
        max_stops  = kwargs.get("max_stops")
        trip_type  = 1 if round_trip else 2
        tfs = build_flights_tfs(
            origin=origin, destination=destination, date_str=travel_date,
            travel_count=travel_count, airline_code=airline_code,
            dep_start_hour=dep_start, dep_end_hour=dep_end, arr_cutoff_hour=arr_cutoff,
            max_price_per_person=max_price, max_stops=max_stops,
            trip_type=trip_type, return_date_str=return_date,
        )
        params = {"tfs": tfs, "hl": "en", "curr": "INR"}
        return f"https://www.google.com/travel/flights/search?{urllib.parse.urlencode(params)}"
    except Exception as _gfl_err:
        print(f"\u26a0\ufe0f make_google_flights_link error ({origin}\u2192{destination} {travel_date}): {_gfl_err}")
        # Fallback: minimal valid TFS with only origin/dest/date (no filters)
        try:
            tfs_min = build_flights_tfs(
                origin=str(origin or "").strip().upper()[:3] or "BOM",
                destination=str(destination or "").strip().upper()[:3] or "DEL",
                date_str=str(travel_date or "")[:10],
                travel_count=max(1, int(travel_count or 1)),
                trip_type=2,
            )
            return f"https://www.google.com/travel/flights/search?{urllib.parse.urlencode({'tfs': tfs_min, 'hl': 'en', 'curr': 'INR'})}"
        except Exception:
            return "https://www.google.com/travel/flights"


def _parse_hour(time_str: str):
    if not time_str:
        return None
    time_str = _clean_time_str(time_str).upper()
    for fmt in ("%I:%M %p", "%H:%M"):
        try:
            return datetime.strptime(time_str, fmt).hour
        except ValueError:
            pass
    return None


# ── Logistics Engine ──────────────────────────────────────────────────────────

def _haversine_km(lon1, lat1, lon2, lat2):
    R = 6371
    dlon = radians(lon2 - lon1); dlat = radians(lat2 - lat1)
    a = sin(dlat/2)**2 + cos(radians(lat1)) * cos(radians(lat2)) * sin(dlon/2)**2
    return R * 2 * atan2(sqrt(a), sqrt(1-a))


def calculate_dynamic_cutoff(dest_airport_code, meeting_address, meeting_dt):
    default_deadline = meeting_dt - timedelta(hours=2)   # 2h default buffer when ORS unavailable
    if not ors_client:
        return default_deadline, None
    try:
        iata         = dest_airport_code.upper().strip()
        airport_data = airport_db.get(iata)
        if not airport_data:
            return default_deadline, None
        airport_coords = (airport_data['lon'], airport_data['lat'])
        city_name      = airport_data.get('city', '')
        country_code   = airport_data.get('country', '')
        addr_parts = [meeting_address]
        if city_name and city_name.lower() not in meeting_address.lower():
            addr_parts.append(city_name)
        if country_code:
            addr_parts.append(country_code)
        meeting_loc = geolocator.geocode(", ".join(addr_parts))
        if not meeting_loc:
            return default_deadline, None
        meeting_coords = (meeting_loc.longitude, meeting_loc.latitude)
        straight_km = _haversine_km(airport_coords[0], airport_coords[1],
                                    meeting_coords[0], meeting_coords[1])
        if straight_km > 300:
            return meeting_dt - timedelta(hours=2), {"drive_mins": 60, "distance_km": round(straight_km, 1), "total_buffer": 120}
        try:
            route      = ors_client.directions(coordinates=(airport_coords, meeting_coords),
                                               profile='driving-car', format='geojson', radiuses=[10000, 10000])
            summary    = route['features'][0]['properties']['summary']
            drive_secs = summary['duration']
            drive_mins = max(15, drive_secs / 60)   # minimum 15 min drive (ORS sometimes returns 0)
            distance_km  = summary['distance'] / 1000
            # buffer = drive time + 1h fixed buffer for check-in / transit
            total_buffer_min = drive_mins + 60
            return meeting_dt - timedelta(minutes=total_buffer_min), {
                "drive_mins":   round(drive_mins),
                "distance_km":  round(distance_km, 1),
                "total_buffer": round(total_buffer_min),
            }
        except Exception as e:
            print(f"⚠️ ORS routing failed: {e}")
            # Fallback: estimate drive from straight-line distance (~40 km/h average city speed)
            est_drive_mins = max(15, round((straight_km / 40) * 60))
            total_buffer_min = est_drive_mins + 60
            return meeting_dt - timedelta(minutes=total_buffer_min), {
                "drive_mins":   est_drive_mins,
                "distance_km":  round(straight_km, 1),
                "total_buffer": round(total_buffer_min),
            }
    except Exception as e:
        print(f"⚠️ General Logistics Error: {e}")
        return default_deadline, None


# ── Search Engine ─────────────────────────────────────────────────────────────

def _search_flights_single_date(origin, destination, travel_date):
    params = {
        "engine": "google_flights", "departure_id": origin, "arrival_id": destination,
        "outbound_date": travel_date, "currency": "INR", "travel_class": "1",
        "type": "2", "hl": "en"
    }
    try:
        results = get_serp_results_with_fallback(params)
        return results.get('best_flights', []) + results.get('other_flights', [])
    except Exception as e:
        print(f"Error connecting to SerpAPI: {e}")
        return []


def search_flights_raw(origin, destination, travel_date, meeting_date_obj=None):
    intl = is_international(origin, destination)
    if not intl:
        print(f"🏠 Domestic route ({origin}→{destination}): searching {travel_date} only")
        return _search_flights_single_date(origin, destination, travel_date)
    base_date = meeting_date_obj or datetime.strptime(travel_date, "%Y-%m-%d").date()
    # International: only search D-1 (day before meeting).
    # Long-haul flights departing on meeting day always arrive after the meeting.
    d1_str = (base_date - timedelta(days=1)).strftime("%Y-%m-%d")
    dates_to_search = [d1_str]
    print(f"🌍 International route ({origin}→{destination}): searching D-1 only ({d1_str})")
    all_flights = []
    with concurrent.futures.ThreadPoolExecutor(max_workers=2) as executor:
        futures = {executor.submit(_search_flights_single_date, origin, destination, d): d for d in dates_to_search}
        for future in concurrent.futures.as_completed(futures):
            date_searched = futures[future]
            try:
                result = future.result()
                print(f"  → {date_searched}: {len(result)} flight group(s)")
                all_flights.extend(result)
            except Exception as e:
                print(f"  ⚠️ Error searching {date_searched}: {e}")
    seen = set(); deduped = []
    for fg in all_flights:
        legs = fg.get("flights", [])
        if not legs:
            continue
        first = legs[0]
        key = (first.get("flight_number", ""), first.get("departure_airport", {}).get("time", ""))
        if key not in seen:
            seen.add(key); deduped.append(fg)
    print(f"  📊 Merged: {len(all_flights)} raw → {len(deduped)} unique flight groups")
    return deduped


# ── Return-flight search ──────────────────────────────────────────────────────

def search_return_flights(
    origin: str,
    destination: str,
    meeting_end_dt,
    travel_count: int = 1,
) -> list:
    end_date_str  = meeting_end_dt.strftime("%Y-%m-%d")
    next_date_str = (meeting_end_dt + timedelta(days=1)).strftime("%Y-%m-%d")
    dates_to_search = [end_date_str, next_date_str]
    print(f"🔄 Return flights ({destination}→{origin}): searching {dates_to_search}")

    all_raw = []
    with concurrent.futures.ThreadPoolExecutor(max_workers=2) as ex:
        futures = {ex.submit(_search_flights_single_date, destination, origin, d): d for d in dates_to_search}
        for future in concurrent.futures.as_completed(futures):
            d = futures[future]
            try:
                res = future.result()
                print(f"  → {d}: {len(res)} group(s)")
                all_raw.extend(res)
            except Exception as e:
                print(f"  ⚠️ Return search error {d}: {e}")

    seen = set(); deduped = []
    for fg in all_raw:
        legs = fg.get("flights", [])
        if not legs:
            continue
        key = (legs[0].get("flight_number", ""), legs[0].get("departure_airport", {}).get("time", ""))
        if key not in seen:
            seen.add(key); deduped.append(fg)

    results = []
    for fg in deduped:
        legs = fg.get("flights", [])
        if not legs:
            continue
        first_leg = legs[0]; last_leg = legs[-1]
        dep_raw = first_leg.get("departure_airport", {}).get("time", "")
        arr_raw = last_leg.get("arrival_airport",   {}).get("time", "")
        if not dep_raw or not arr_raw:
            continue
        try:
            dep_dt = datetime.strptime(dep_raw, "%Y-%m-%d %H:%M")
            arr_dt = datetime.strptime(arr_raw, "%Y-%m-%d %H:%M")
        except Exception:
            continue

        if dep_dt < meeting_end_dt + timedelta(hours=1):
            continue

        price = fg.get("price", float("inf"))
        if not isinstance(price, (int, float)) or price <= 0 or price == float("inf"):
            continue

        dep_date = dep_dt.date()
        arr_date = arr_dt.date()
        end_date = meeting_end_dt.date()

        same_day_return = (dep_date == end_date and arr_date == end_date)
        needs_hotel     = (dep_date > end_date) or (arr_date > dep_date)

        duration_mins = fg.get("total_duration") or int((arr_dt - dep_dt).total_seconds() / 60)
        h, m = divmod(duration_mins, 60)

        airline_names = []
        for leg in legs:
            a = leg.get("airline", "")
            if a and a not in airline_names:
                airline_names.append(a)
        airline_name         = ", ".join(airline_names) if airline_names else ""
        airline_name_primary = first_leg.get("airline", "")
        num_stops            = len(legs) - 1
        flight_no_base       = first_leg.get("flight_number", "")
        flight_no            = flight_no_base
        if num_stops > 0:
            last_no = last_leg.get("flight_number", "")
            if last_no and last_no != flight_no_base:
                flight_no = f"{flight_no_base}, {last_no}"

        thumb = fg.get("airline_logo") or first_leg.get("airline_logo")
        if not (isinstance(thumb, str) and thumb.startswith("http")):
            thumb = None

        airline_code = get_airline_code(airline_name_primary)
        extracted    = flight_no_base[:2].upper() if len(flight_no_base) >= 2 else ""
        gl_code      = airline_code or extracted

        dep_ui  = dep_dt.strftime("%I:%M %p")
        arr_ui  = arr_dt.strftime("%I:%M %p")
        dep_str = dep_dt.strftime("%Y-%m-%d")

        if needs_hotel and arr_date > dep_date:
            arr_label = f"{arr_ui}  \n*(Next day — {arr_dt.strftime('%b %d')})*"
        else:
            arr_label = arr_ui

        total_price     = price * travel_count
        hotel_check_in  = end_date
        hotel_check_out = dep_date

        # Build Google Flights URL — pass travel_count for correct total pricing
        # Bulletproof: make_google_flights_link has its own full try/except fallback
        _price_pp = max(1, int(price)) if price and price > 0 else None
        gl_url = make_google_flights_link(
            destination, origin, dep_str,
            airline_name_primary, travel_count,
            dep_time_str=dep_ui,
            arr_time_str=arr_ui,   # already clean (no newlines)
            price_per_person=_price_pp,
            _override_airline_code=gl_code,
            round_trip=False,      # return leg is always a one-way search
            return_date=None,
        )

        results.append({
            "airline":         airline_name,
            "airline_primary": airline_name_primary,
            "flight_no_base":  flight_no_base,
            "flight_no":       flight_no,
            "dep_time":        dep_ui,
            "arr_time":        arr_label,
            "arr_time_raw":    arr_ui,
            "duration":        f"{h}h {m}m" if h > 0 else f"{m}m",
            "price":           total_price,
            "thumbnail":       thumb,
            "link":            gl_url,
            "stops":           num_stops,
            "departure_date":  dep_str,
            "arrival_date":    arr_dt.strftime("%Y-%m-%d"),
            "same_day_return": same_day_return,
            "needs_hotel":     needs_hotel,
            "hotel_check_in":  hotel_check_in,
            "hotel_check_out": hotel_check_out,
            "dep_dt":          dep_dt,
            "arr_dt":          arr_dt,
            # Mark as return so Excel/DB can distinguish
            "_is_return":      True,
            "return_origin":   destination,
            "return_dest":     origin,
        })

    same_day_results = [r for r in results if r["same_day_return"]]
    if same_day_results:
        results = same_day_results

    results.sort(key=lambda x: (not x["same_day_return"], x["price"]))
    print(f"  ✅ {len(results)} return flight(s) valid after {meeting_end_dt.strftime('%H:%M')}")
    return results


# ── Flight Processor ──────────────────────────────────────────────────────────

def process_flight_results(
    flights, cutoff_dt, origin, destination, meeting_datetime,
    travel_count=1, dedupe=True, meeting_location: str = None,
    allow_prev_night: bool = False,
):
    """
    allow_prev_night=True: if no same-day flights pass the cutoff, return flights that
    arrive on the *previous* night (so employee can check in a night early).
    Returned flights in that case get category="prev_night" and needs_hotel=True.
    """
    processed_flights       = []
    best_flight_per_airline = {}
    prev_night_flights      = []      # fallback bucket for domestic
    meeting_date_only       = meeting_datetime.date()
    _is_intl                = is_international(origin, destination)
    _stop_limits            = [1, 2, 3] if _is_intl else [1, 2]

    for stop_limit in _stop_limits:
        for flight_group in flights:
            legs = flight_group.get('flights', [])
            if not legs or len(legs) != stop_limit:
                continue
            first_leg = legs[0]; last_leg = legs[-1]
            try:
                dep_time_raw = first_leg.get('departure_airport', {}).get('time', '')
                arr_time_raw = last_leg.get('arrival_airport',   {}).get('time', '')
                if not dep_time_raw or not arr_time_raw:
                    continue
                dep_dt = datetime.strptime(dep_time_raw, "%Y-%m-%d %H:%M")
                arr_dt = datetime.strptime(arr_time_raw, "%Y-%m-%d %H:%M")
                arr_date_only = arr_dt.date()
                days_diff     = (meeting_date_only - arr_date_only).days
                flight_category = "same_day"
                _is_prev_night  = False
                if _is_intl:
                    if days_diff < 0:
                        continue
                    elif days_diff == 0:
                        # Arrives on meeting day — must land before meeting starts
                        if arr_dt >= meeting_datetime:
                            continue
                    elif days_diff == 1:
                        flight_category = "early"
                    elif days_diff > 1:
                        continue
                else:
                    # Domestic: strict same-day filter first
                    if days_diff == 0:
                        # Must arrive before meeting start time
                        if arr_dt >= meeting_datetime:
                            continue
                        # Must respect the calculated cutoff buffer
                        if cutoff_dt is not None and arr_dt >= cutoff_dt:
                            continue
                        # Passes — valid same-day flight
                    elif days_diff == 1 and allow_prev_night:
                        # Arrived yesterday evening — qualifies as prev-night fallback
                        _is_prev_night  = True
                        flight_category = "prev_night"
                    else:
                        continue

                airline_names = []
                for leg in legs:
                    aname = leg.get('airline', '')
                    if aname and aname not in airline_names:
                        airline_names.append(aname)
                airline_name         = ", ".join(airline_names) if airline_names else ""
                airline_name_primary = first_leg.get('airline', '')
                price_per_person     = flight_group.get('price', float('inf'))
                if not isinstance(price_per_person, (int, float)) or price_per_person <= 0 or price_per_person == float('inf'):
                    continue
                total_price  = price_per_person * travel_count
                dep_ui_fmt   = dep_dt.strftime("%I:%M %p")
                base_arr_fmt = arr_dt.strftime("%I:%M %p")
                if _is_prev_night:
                    arr_display_date = arr_dt.strftime("%b %d")
                    arr_ui_fmt = f"{base_arr_fmt}  \n*(Prev night — {arr_display_date})*"
                elif _is_intl:
                    arrival_display_date = arr_dt.strftime("%b %d").replace(" 0", " ")
                    arr_ui_fmt = f"{base_arr_fmt}  \n*(Arrives {arrival_display_date})*"
                else:
                    arr_ui_fmt = base_arr_fmt
                departure_date_str = dep_dt.strftime("%Y-%m-%d")
                arrival_date_str   = arr_dt.strftime("%Y-%m-%d")
                duration_mins = flight_group.get('total_duration')
                if duration_mins is None:
                    duration_mins = int((arr_dt - dep_dt).total_seconds() / 60)
                hours, mins  = divmod(duration_mins, 60)
                duration_str = f"{hours}h {mins}m" if hours > 0 else f"{mins}m"
                flight_no_base = first_leg.get('flight_number', '')
                airline_code   = get_airline_code(airline_name_primary)
                mmt_specific_link = make_mmt_link(
                    origin, destination, meeting_datetime,
                    travel_count, arrival_cutoff_dt=cutoff_dt, airline_code=airline_code
                )
                extracted_iata = flight_no_base[:2].upper() if flight_no_base and len(flight_no_base) >= 2 else ""
                gl_airline_code = airline_code or extracted_iata
                specific_link = make_google_flights_link(
                    origin, destination, departure_date_str,
                    airline_name_primary, travel_count,
                    dep_time_str=dep_ui_fmt,
                    arr_time_str=base_arr_fmt,
                    price_per_person=price_per_person,
                    _override_airline_code=gl_airline_code,
                )
                thumb = flight_group.get('airline_logo')
                if not (isinstance(thumb, str) and thumb.startswith("http")):
                    thumb = first_leg.get('airline_logo')
                if not (isinstance(thumb, str) and thumb.startswith("http")):
                    thumb = next((e for e in first_leg.get('extensions', [])
                                  if isinstance(e, str) and e.startswith("http")), None)
                flight_no = flight_no_base
                # prev_night always needs hotel (check-in = arrival date)
                needs_hotel = _is_prev_night
                num_stops = stop_limit - 1
                if num_stops > 0:
                    last_flight_no = last_leg.get('flight_number', '')
                    if last_flight_no and flight_no != last_flight_no:
                        flight_no = f"{flight_no}, {last_flight_no}"
                    if len(legs) > 1:
                        first_arr_raw  = first_leg.get('arrival_airport', {}).get('time', '')
                        second_dep_raw = legs[1].get('departure_airport', {}).get('time', '')
                        if first_arr_raw and second_dep_raw:
                            first_arr_dt  = datetime.strptime(first_arr_raw, "%Y-%m-%d %H:%M")
                            second_dep_dt = datetime.strptime(second_dep_raw, "%Y-%m-%d %H:%M")
                            if (second_dep_dt - first_arr_dt).total_seconds() / 60 > 120:
                                needs_hotel = True
                if flight_category == "early":
                    needs_hotel = True
                # For prev_night: hotel check-in = arrival date, check-out = meeting date
                hotel_checkin_date  = arr_dt.date() if _is_prev_night else None
                hotel_checkout_date = meeting_date_only if _is_prev_night else None
                flight_data = {
                    "airline": airline_name, "flight_no_base": flight_no_base,
                    "flight_no": flight_no, "dep_time": dep_ui_fmt, "arr_time": arr_ui_fmt,
                    "arr_time_raw": base_arr_fmt,
                    "duration": duration_str, "price": total_price, "thumbnail": thumb,
                    "link": specific_link, "mmt_link": mmt_specific_link,
                    "stops": num_stops, "needs_hotel": needs_hotel,
                    "departure_date": departure_date_str, "arrival_date": arrival_date_str,
                    "category": flight_category,
                    "hotel_checkin_date":  hotel_checkin_date,
                    "hotel_checkout_date": hotel_checkout_date,
                    "_is_return": False,
                }
                if _is_prev_night:
                    # Collect separately — only used when no same-day flights exist
                    prev_night_flights.append(flight_data)
                    continue
                if dedupe:
                    if airline_name not in best_flight_per_airline:
                        best_flight_per_airline[airline_name] = flight_data
                    else:
                        existing = best_flight_per_airline[airline_name]
                        if flight_data["category"] == "same_day" and existing["category"] == "early":
                            best_flight_per_airline[airline_name] = flight_data
                        elif flight_data["category"] == "early" and existing["category"] == "same_day":
                            pass
                        elif flight_data["stops"] < existing["stops"]:
                            best_flight_per_airline[airline_name] = flight_data
                        elif flight_data["stops"] == existing["stops"] and total_price < existing["price"]:
                            best_flight_per_airline[airline_name] = flight_data
                else:
                    processed_flights.append(flight_data)
            except Exception as e:
                print(f"Error processing flight: {e}"); continue

        if (dedupe and best_flight_per_airline) or (not dedupe and processed_flights):
            break

    results = list(best_flight_per_airline.values()) if dedupe else processed_flights
    if not dedupe:
        seen_keys = set(); deduped_results = []
        for f in results:
            key = (f.get("flight_no_base",""), f.get("dep_time",""), f.get("arrival_date",""))
            if key not in seen_keys:
                seen_keys.add(key); deduped_results.append(f)
        results = deduped_results

    same_day  = sorted([f for f in results if f.get("category") == "same_day"], key=lambda x: x['price'])
    early     = sorted([f for f in results if f.get("category") == "early"],    key=lambda x: x['price'])
    prev_night_sorted = sorted(prev_night_flights, key=lambda x: x['price'])
    if _is_intl:
        return [f.copy() for f in same_day] + [f.copy() for f in early]
    else:
        if same_day:
            return [f.copy() for f in same_day]
        # No same-day flights pass cutoff — return prev-night fallback if available
        return [f.copy() for f in prev_night_sorted]


# ── Helpers ───────────────────────────────────────────────────────────────────

def get_airline_code(airline_name: str) -> str:
    if not airline_name:
        return ""
    mapping = {
        "indigo": "6E", "air india": "AI", "air india express": "IX",
        "vistara": "UK", "spicejet": "SG", "akasa air": "QP",
        "alliance air": "9I", "star air": "S5",
        "emirates": "EK", "qatar airways": "QR", "etihad airways": "EY",
        "etihad": "EY", "oman air": "WY", "gulf air": "GF", "saudia": "SV",
        "saudi arabian airlines": "SV", "flydubai": "FZ", "air arabia": "G9",
        "lufthansa": "LH", "british airways": "BA", "air france": "AF",
        "klm": "KL", "klm royal dutch airlines": "KL", "swiss": "LX",
        "swiss international air lines": "LX", "turkish airlines": "TK",
        "virgin atlantic": "VS", "iberia": "IB", "finnair": "AY",
        "austrian": "OS", "austrian airlines": "OS", "brussels airlines": "SN",
        "scandinavian airlines": "SK", "sas": "SK", "lot polish airlines": "LO",
        "tap air portugal": "TP", "aeroflot": "SU",
        "singapore airlines": "SQ", "cathay pacific": "CX", "thai airways": "TG",
        "japan airlines": "JL", "jal": "JL", "ana": "NH",
        "all nippon airways": "NH", "korean air": "KE", "asiana airlines": "OZ",
        "malaysia airlines": "MH", "garuda indonesia": "GA",
        "philippine airlines": "PR", "vietnam airlines": "VN",
        "china airlines": "CI", "china eastern": "MU", "china southern": "CZ",
        "air china": "CA", "eva air": "BR", "bangkok airways": "PG",
        "cebu pacific": "5J", "airasia": "AK", "airasia india": "I5",
        "scoot": "TR", "dragonair": "KA",
        "united airlines": "UA", "united": "UA", "american airlines": "AA",
        "delta air lines": "DL", "delta": "DL", "air canada": "AC",
        "westjet": "WS", "southwest airlines": "WN", "jetblue": "B6",
        "spirit airlines": "NK", "frontier airlines": "F9",
        "qantas": "QF", "air new zealand": "NZ", "south african airways": "SA",
        "kenya airways": "KQ", "ethiopian airlines": "ET",
    }
    return mapping.get(airline_name.lower().strip(), "")


import json as _json_mod
import os as _os_mod

MMT_CACHE_FILE = "mmt_city_cache.json"

def _load_mmt_cache():
    if os.path.exists(MMT_CACHE_FILE):
        try:
            with open(MMT_CACHE_FILE, "r") as f:
                return json.load(f)
        except Exception:
            return {}
    return {}

def _save_mmt_cache(cache_dict):
    try:
        with open(MMT_CACHE_FILE, "w") as f:
            json.dump(cache_dict, f, indent=4)
    except Exception as e:
        print(f"Error saving MMT cache: {e}")

_mmt_city_cache = _load_mmt_cache()

# ── Hardcoded locality → MMT city code lookup (never hallucinates) ────────────
# Format: "keyword" → "MMT_LOCUS_ID"
# Always locusType=city, type=city — no fake area tags ever
_MMT_LOCATION_TABLE = {
    # ── MUMBAI & surroundings ─────────────────────────────────────────────────
    "mumbai": "RGMUM", "bombay": "RGMUM",
    "andheri": "RGMUM", "andheri east": "RGMUM", "andheri west": "RGMUM",
    "bandra": "RGMUM", "bandra east": "RGMUM", "bandra west": "RGMUM",
    "borivali": "RGMUM", "borivali east": "RGMUM", "borivali west": "RGMUM",
    "malad": "RGMUM", "malad east": "RGMUM", "malad west": "RGMUM",
    "kandivali": "RGMUM", "kandivali east": "RGMUM", "kandivali west": "RGMUM",
    "goregaon": "RGMUM", "goregaon east": "RGMUM", "goregaon west": "RGMUM",
    "jogeshwari": "RGMUM", "jogeshwari east": "RGMUM", "jogeshwari west": "RGMUM",
    "santacruz": "RGMUM", "santacruz east": "RGMUM", "santacruz west": "RGMUM",
    "vile parle": "RGMUM", "vile parle east": "RGMUM", "vile parle west": "RGMUM",
    "juhu": "RGMUM", "versova": "RGMUM",
    "kurla": "RGMUM", "kurla east": "RGMUM", "kurla west": "RGMUM",
    "ghatkopar": "RGMUM", "ghatkopar east": "RGMUM", "ghatkopar west": "RGMUM",
    "vikhroli": "RGMUM", "kanjurmarg": "RGMUM", "bhandup": "RGMUM",
    "mulund": "RGMUM", "mulund east": "RGMUM", "mulund west": "RGMUM",
    "thane": "RGMUM", "thane east": "RGMUM", "thane west": "RGMUM",
    "kalyan": "RGMUM", "dombivli": "RGMUM", "ulhasnagar": "RGMUM",
    "mira road": "RGMUM", "bhayander": "RGMUM", "virar": "RGMUM", "vasai": "RGMUM",
    "dadar": "RGMUM", "prabhadevi": "RGMUM", "worli": "RGMUM", "lower parel": "RGMUM",
    "elphinstone": "RGMUM", "parel": "RGMUM", "sewri": "RGMUM",
    "dharavi": "RGMUM", "sion": "RGMUM", "chunabhatti": "RGMUM",
    "chembur": "RGMUM", "govandi": "RGMUM", "mankhurd": "RGMUM",
    "trombay": "RGMUM", "tilak nagar": "RGMUM",
    "colaba": "RGMUM", "nariman point": "RGMUM", "churchgate": "RGMUM",
    "fort": "RGMUM", "byculla": "RGMUM", "grant road": "RGMUM",
    "mumbai central": "RGMUM", "mahalaxmi": "RGMUM",
    "powai": "RGMUM", "hiranandani": "RGMUM", "chandivali": "RGMUM",
    "navi mumbai": "RGMUM", "nerul": "RGMUM", "vashi": "RGMUM",
    "kharghar": "RGMUM", "airoli": "RGMUM", "ghansoli": "RGMUM",
    "belapur": "RGMUM", "cbd belapur": "RGMUM", "sanpada": "RGMUM",
    "turbhe": "RGMUM", "kopar khairane": "RGMUM",
    "panvel": "RGMUM", "new panvel": "RGMUM", "old panvel": "RGMUM",
    "taloja": "RGMUM", "khopoli": "RGMUM",
    "bkc": "RGMUM", "bandra kurla complex": "RGMUM",
    "lower parel": "RGMUM", "peninsula": "RGMUM",
    # Raigad district — nearest MMT city is Mumbai
    "alibaug": "RGMUM", "alibag": "RGMUM", "alibaugh": "RGMUM",
    "pen": "RGMUM", "karjat": "RGMUM", "khalapur": "RGMUM",
    "roha": "RGMUM", "murud": "RGMUM", "mandwa": "RGMUM",
    "uran": "RGMUM", "dronagiri": "RGMUM",

    # ── DELHI / NCR ───────────────────────────────────────────────────────────
    "delhi": "RGNCR", "new delhi": "RGNCR", "ncr": "RGNCR",
    "connaught place": "RGNCR", "cp": "RGNCR", "karol bagh": "RGNCR",
    "paharganj": "RGNCR", "chandni chowk": "RGNCR", "old delhi": "RGNCR",
    "lajpat nagar": "RGNCR", "south extension": "RGNCR", "defence colony": "RGNCR",
    "greater kailash": "RGNCR", "gk1": "RGNCR", "gk2": "RGNCR",
    "hauz khas": "RGNCR", "green park": "RGNCR", "safdarjung": "RGNCR",
    "vasant kunj": "RGNCR", "vasant vihar": "RGNCR", "mehrauli": "RGNCR",
    "saket": "RGNCR", "malviya nagar": "RGNCR",
    "pitampura": "RGNCR", "rohini": "RGNCR", "shalimar bagh": "RGNCR",
    "janakpuri": "RGNCR", "dwarka": "RGNCR", "uttam nagar": "RGNCR",
    "vikaspuri": "RGNCR", "palam": "RGNCR",
    "shahdara": "RGNCR", "dilshad garden": "RGNCR", "preet vihar": "RGNCR",
    "laxmi nagar": "RGNCR", "east delhi": "RGNCR",
    "gurgaon": "RGNCR", "gurugram": "RGNCR", "sohna road": "RGNCR",
    "golf course road": "RGNCR", "mg road": "RGNCR", "cyber city": "RGNCR",
    "udyog vihar": "RGNCR", "sector 29": "RGNCR", "dlf cyber hub": "RGNCR",
    "noida": "RGNCR", "greater noida": "RGNCR", "noida expressway": "RGNCR",
    "sector 18": "RGNCR", "sector 62": "RGNCR",
    "faridabad": "RGNCR", "ballabhgarh": "RGNCR",
    "ghaziabad": "RGNCR", "indirapuram": "RGNCR", "vaishali": "RGNCR",

    # ── BENGALURU ─────────────────────────────────────────────────────────────
    "bangalore": "RGBLR", "bengaluru": "RGBLR", "blr": "RGBLR",
    "koramangala": "RGBLR", "indiranagar": "RGBLR", "whitefield": "RGBLR",
    "electronic city": "RGBLR", "btm layout": "RGBLR", "hsr layout": "RGBLR",
    "jayanagar": "RGBLR", "jp nagar": "RGBLR", "banashankari": "RGBLR",
    "malleshwaram": "RGBLR", "rajajinagar": "RGBLR", "basavanagudi": "RGBLR",
    "mg road": "RGBLR", "brigade road": "RGBLR", "commercial street": "RGBLR",
    "marathahalli": "RGBLR", "sarjapur": "RGBLR", "bellandur": "RGBLR",
    "yelahanka": "RGBLR", "hebbal": "RGBLR", "devanahalli": "RGBLR",
    "kengeri": "RGBLR", "tumkur road": "RGBLR", "yeshwanthpur": "RGBLR",
    "bannerghatta": "RGBLR", "bommanahalli": "RGBLR",

    # ── HYDERABAD ─────────────────────────────────────────────────────────────
    "hyderabad": "RGHYD", "secunderabad": "RGHYD", "cyberabad": "RGHYD",
    "banjara hills": "RGHYD", "jubilee hills": "RGHYD", "madhapur": "RGHYD",
    "hitec city": "RGHYD", "hitech city": "RGHYD", "gachibowli": "RGHYD",
    "kondapur": "RGHYD", "manikonda": "RGHYD", "kukatpally": "RGHYD",
    "dilsukhnagar": "RGHYD", "lb nagar": "RGHYD", "uppal": "RGHYD",
    "ameerpet": "RGHYD", "sr nagar": "RGHYD", "begumpet": "RGHYD",
    "paradise": "RGHYD", "nampally": "RGHYD", "abids": "RGHYD",
    "shamshabad": "RGHYD", "rajendranagar": "RGHYD",

    # ── PUNE ─────────────────────────────────────────────────────────────────
    "pune": "RGPNQ", "pune city": "RGPNQ",
    "shivajinagar": "RGPNQ", "fc road": "RGPNQ", "jm road": "RGPNQ",
    "kothrud": "RGPNQ", "karve nagar": "RGPNQ", "erandwane": "RGPNQ",
    "aundh": "RGPNQ", "baner": "RGPNQ", "balewadi": "RGPNQ",
    "hinjewadi": "RGPNQ", "wakad": "RGPNQ", "pimpri": "RGPNQ",
    "chinchwad": "RGPNQ", "pimpri chinchwad": "RGPNQ", "pcmc": "RGPNQ",
    "viman nagar": "RGPNQ", "kalyani nagar": "RGPNQ", "koregaon park": "RGPNQ",
    "kharadi": "RGPNQ", "magarpatta": "RGPNQ", "hadapsar": "RGPNQ",
    "kondhwa": "RGPNQ", "katraj": "RGPNQ", "sinhagad road": "RGPNQ",
    "deccan": "RGPNQ", "camp": "RGPNQ", "cantonment": "RGPNQ",
    "pune cantonment": "RGPNQ", "yerawada": "RGPNQ", "nagar road": "RGPNQ",
    "lonavala": "RGPNQ", "khandala": "RGPNQ", "talegaon": "RGPNQ",

    # ── CHENNAI ──────────────────────────────────────────────────────────────
    "chennai": "RGMAA", "madras": "RGMAA",
    "t nagar": "RGMAA", "nungambakkam": "RGMAA", "egmore": "RGMAA",
    "adyar": "RGMAA", "velachery": "RGMAA", "sholinganallur": "RGMAA",
    "omr": "RGMAA", "old mahabalipuram road": "RGMAA",
    "porur": "RGMAA", "guindy": "RGMAA", "mount road": "RGMAA",
    "anna nagar": "RGMAA", "kilpauk": "RGMAA", "perambur": "RGMAA",
    "ambattur": "RGMAA", "avadi": "RGMAA",
    "tambaram": "RGMAA", "chromepet": "RGMAA", "pallavaram": "RGMAA",
    "perungudi": "RGMAA", "thoraipakkam": "RGMAA", "siruseri": "RGMAA",

    # ── KOLKATA ───────────────────────────────────────────────────────────────
    "kolkata": "RGCCU", "calcutta": "RGCCU",
    "park street": "RGCCU", "esplanade": "RGCCU", "bbd bagh": "RGCCU",
    "salt lake": "RGCCU", "sector v": "RGCCU", "rajarhat": "RGCCU",
    "new town": "RGCCU", "action area": "RGCCU",
    "howrah": "RGCCU", "tollygunge": "RGCCU", "jadavpur": "RGCCU",
    "ballygunge": "RGCCU", "alipore": "RGCCU",
    "gariahat": "RGCCU", "lake town": "RGCCU", "dum dum": "RGCCU",

    # ── AHMEDABAD ─────────────────────────────────────────────────────────────
    "ahmedabad": "RGAMD", "amdavad": "RGAMD",
    "sg highway": "RGAMD", "satellite": "RGAMD", "bopal": "RGAMD",
    "prahlad nagar": "RGAMD", "bodakdev": "RGAMD", "vastrapur": "RGAMD",
    "navrangpura": "RGAMD", "cg road": "RGAMD",
    "maninagar": "RGAMD", "naroda": "RGAMD", "gota": "RGAMD",
    "gandhinagar": "RGAMD", "gift city": "RGAMD",

    # ── VADODARA ──────────────────────────────────────────────────────────────
    "vadodara": "RGBDQ", "baroda": "RGBDQ",
    "alkapuri": "RGBDQ", "fatehgunj": "RGBDQ", "race course": "RGBDQ",
    "gotri": "RGBDQ", "harni": "RGBDQ", "manjalpur": "RGBDQ",

    # ── GOA ──────────────────────────────────────────────────────────────────
    "goa": "RGGOI", "north goa": "RGGOI", "south goa": "RGGOI",
    "panaji": "RGGOI", "panjim": "RGGOI", "margao": "RGGOI",
    "calangute": "RGGOI", "baga": "RGGOI", "candolim": "RGGOI",
    "mapusa": "RGGOI", "ponda": "RGGOI", "vasco": "RGGOI",
    "colva": "RGGOI", "benaulim": "RGGOI", "anjuna": "RGGOI",
    "vagator": "RGGOI", "morjim": "RGGOI", "arambol": "RGGOI",

    # ── JAIPUR ───────────────────────────────────────────────────────────────
    "jaipur": "RGJPR", "pink city": "RGJPR",
    "malviya nagar jaipur": "RGJPR", "vaishali nagar": "RGJPR",
    "mansarovar": "RGJPR", "civil lines": "RGJPR", "c scheme": "RGJPR",
    "tonk road": "RGJPR", "ajmer road": "RGJPR",

    # ── SURAT ────────────────────────────────────────────────────────────────
    "surat": "RGSTV",
    "adajan": "RGSTV", "vesu": "RGSTV", "katargam": "RGSTV",
    "athwa": "RGSTV", "citylight": "RGSTV", "piplod": "RGSTV",

    # ── KOCHI ────────────────────────────────────────────────────────────────
    "kochi": "RGCOK", "cochin": "RGCOK", "ernakulam": "RGCOK",
    "fort kochi": "RGCOK", "mg road kochi": "RGCOK", "marine drive": "RGCOK",
    "edapally": "RGCOK", "kakkanad": "RGCOK", "infopark": "RGCOK",

    # ── LUCKNOW ──────────────────────────────────────────────────────────────
    "lucknow": "RGLKO",
    "hazratganj": "RGLKO", "gomti nagar": "RGLKO", "alambagh": "RGLKO",
    "mahanagar": "RGLKO", "aliganj": "RGLKO",

    # ── CHANDIGARH ───────────────────────────────────────────────────────────
    "chandigarh": "RGIXC", "mohali": "RGIXC", "panchkula": "RGIXC",
    "sector 17": "RGIXC", "sector 35": "RGIXC",

    # ── NAGPUR ───────────────────────────────────────────────────────────────
    "nagpur": "RGNAG",
    "sitabuldi": "RGNAG", "dharampeth": "RGNAG", "wardha road": "RGNAG",

    # ── COIMBATORE ───────────────────────────────────────────────────────────
    "coimbatore": "RGCJB", "kovai": "RGCJB",

    # ── VISAKHAPATNAM ────────────────────────────────────────────────────────
    "visakhapatnam": "RGVTZ", "vizag": "RGVTZ",
    "beach road vizag": "RGVTZ", "mvp colony": "RGVTZ",

    # ── BHUBANESWAR ──────────────────────────────────────────────────────────
    "bhubaneswar": "RGBBI", "cuttack": "RGBBI",

    # ── INTERNATIONAL ─────────────────────────────────────────────────────────
    "london": "CTLONDO", "canary wharf": "CTLONDO", "the city": "CTLONDO",
    "mayfair": "CTLONDO", "westminster": "CTLONDO", "shoreditch": "CTLONDO",
    "heathrow": "CTLONDO", "gatwick": "CTLONDO", "stansted": "CTLONDO",
    "dubai": "CTDUBAI", "deira": "CTDUBAI", "bur dubai": "CTDUBAI",
    "jbr": "CTDUBAI", "marina": "CTDUBAI", "jlt": "CTDUBAI",
    "business bay": "CTDUBAI", "difc": "CTDUBAI", "downtown dubai": "CTDUBAI",
    "singapore": "CTSINGAP", "orchard": "CTSINGAP", "raffles place": "CTSINGAP",
    "marina bay": "CTSINGAP", "sentosa": "CTSINGAP",
    "new york": "CTNEWYORK", "manhattan": "CTNEWYORK", "midtown": "CTNEWYORK",
    "times square": "CTNEWYORK", "brooklyn": "CTNEWYORK",
    "paris": "CTPARIS", "champs elysees": "CTPARIS", "le marais": "CTPARIS",
    "frankfurt": "CTFRANKF",
    "amsterdam": "CTAMSTER",
    "zurich": "CTZURICH",
    "tokyo": "CTTOKYO", "shinjuku": "CTTOKYO", "shibuya": "CTTOKYO",
    "osaka": "CTOSAKA",
    "hong kong": "CTHONGKO",
    "bangkok": "CTBANGKO", "sukhumvit": "CTBANGKO", "silom": "CTBANGKO",
    "kuala lumpur": "CTKUALAL", "klcc": "CTKUALAL", "bukit bintang": "CTKUALAL",
    "sydney": "CTSYDNEY", "melbourne": "CTMELB",
    "toronto": "CTTORONT", "new york city": "CTNEWYORK",
    "chicago": "CTCHICAG", "san francisco": "CTSANFRA",
    "los angeles": "CTLOSANG", "la": "CTLOSANG",
    "doha": "CTDOHA",
    "abu dhabi": "CTABUDHA",
    "muscat": "CTMUSCAT",
}

def _resolve_mmt_locus(meeting_location: str, dest_iata: str) -> dict:
    """
    Resolve MMT locusId from meeting_location using hardcoded table.
    Falls back to IATA-based lookup, then LLM, then Mumbai.
    Detects area-level locations and returns locusType=region, type=area.
    """
    loc_lower = meeting_location.strip().lower() if meeting_location else ""
    matched_key = ""   # the key that matched in the table

    # 1. Direct lookup — exact match first
    if loc_lower in _MMT_LOCATION_TABLE:
        city_code = _MMT_LOCATION_TABLE[loc_lower]
        matched_key = loc_lower
    else:
        # 2. Substring match — find longest matching key inside the location string
        best_key = ""
        city_code = ""
        for key, code in _MMT_LOCATION_TABLE.items():
            if key in loc_lower and len(key) > len(best_key):
                best_key = key
                city_code = code
        matched_key = best_key

    # 3. Fallback: derive from dest_iata airport country/city
    if not city_code:
        iata_upper = (dest_iata or "").strip().upper()
        _iata_to_mmt = {
            "BOM": "RGMUM", "BDQ": "RGBDQ", "DEL": "RGNCR", "BLR": "RGBLR",
            "HYD": "RGHYD", "MAA": "RGMAA", "CCU": "RGCCU", "PNQ": "RGPNQ",
            "AMD": "RGAMD", "GOI": "RGGOI", "JAI": "RGJPR", "STV": "RGSTV",
            "COK": "RGCOK", "LKO": "RGLKO", "VTZ": "RGVTZ", "NAG": "RGNAG",
            "CJB": "RGCJB", "IXC": "RGIXC", "BBI": "RGBBI",
            "LHR": "CTLONDO", "LGW": "CTLONDO", "STN": "CTLONDO",
            "DXB": "CTDUBAI", "AUH": "CTABUDHA", "DOH": "CTDOHA",
            "SIN": "CTSINGAP", "JFK": "CTNEWYORK", "EWR": "CTNEWYORK",
            "CDG": "CTPARIS", "ORY": "CTPARIS",
            "FRA": "CTFRANKF", "AMS": "CTAMSTER", "ZRH": "CTZURICH",
            "NRT": "CTTOKYO", "HND": "CTTOKYO", "KIX": "CTOSAKA",
            "HKG": "CTHONGKO", "BKK": "CTBANGKO", "KUL": "CTKUALAL",
            "SYD": "CTSYDNEY", "MEL": "CTMELB",
            "YYZ": "CTTORONT", "ORD": "CTCHICAG", "SFO": "CTSANFRA",
            "LAX": "CTLOSANG",
        }
        city_code = _iata_to_mmt.get(iata_upper, "")

    # 4. LLM fallback — ask Groq for the nearest known MMT city code
    if not city_code:
        _all_known_codes = list(_LOCUS_TO_CITY_NAME.items())  # [(code, name), ...]
        _known_str = ", ".join(f"{name}={code}" for code, name in _all_known_codes)
        try:
            _groq_key = os.getenv("GROQ_API_KEY")
            if _groq_key and Groq:
                _gc = Groq(api_key=_groq_key)
                _prompt = (
                    f"Given this location: \"{meeting_location}\" (airport: {dest_iata}), "
                    f"which of these MakeMyTrip city codes is the NEAREST major city?\n"
                    f"Known codes: {_known_str}\n"
                    f"Reply with ONLY the code (e.g. RGMUM). Nothing else."
                )
                _resp = _gc.chat.completions.create(
                    model="llama-3.3-70b-versatile",
                    messages=[{"role": "user", "content": _prompt}],
                    max_tokens=20, temperature=0,
                )
                _llm_code = _resp.choices[0].message.content.strip().upper()
                if _llm_code in _LOCUS_TO_CITY_NAME:
                    city_code = _llm_code
                    print(f"  ✅ LLM resolved MMT city: {meeting_location} → {city_code} ({_LOCUS_TO_CITY_NAME[city_code]})")
        except Exception as _mmt_llm_err:
            print(f"  ⚠️ MMT LLM city resolution failed: {_mmt_llm_err}")

    # 5. Ultimate fallback — Mumbai for India, or best-guess for international
    if not city_code:
        city_code = "RGMUM"
        print(f"  ⚠️ MMT fallback to Mumbai for: {meeting_location} ({dest_iata})")

    # Determine country
    country = "GB" if city_code.startswith("CTLONDO") else \
              "AE" if city_code in ("CTDUBAI", "CTABUDHA") else \
              "SG" if city_code == "CTSINGAP" else \
              "US" if city_code in ("CTNEWYORK", "CTCHICAG", "CTSANFRA", "CTLOSANG") else \
              "FR" if city_code == "CTPARIS" else \
              "DE" if city_code == "CTFRANKF" else \
              "NL" if city_code == "CTAMSTER" else \
              "CH" if city_code == "CTZURICH" else \
              "JP" if city_code in ("CTTOKYO", "CTOSAKA") else \
              "HK" if city_code == "CTHONGKO" else \
              "TH" if city_code == "CTBANGKO" else \
              "MY" if city_code == "CTKUALAL" else \
              "AU" if city_code in ("CTSYDNEY", "CTMELB") else \
              "CA" if city_code == "CTTORONT" else \
              "QA" if city_code == "CTDOHA" else \
              "OM" if city_code == "CTMUSCAT" else "IN"

    # Detect area vs city: if the matched key is NOT the canonical city name,
    # it's a neighbourhood/area within that city (e.g. "andheri east" → RGMUM)
    canonical_city_name = _locus_to_city_name(city_code)           # e.g. "Mumbai"
    canonical_lower     = canonical_city_name.lower()
    is_area = bool(
        matched_key
        and canonical_lower
        and matched_key != canonical_lower
        and matched_key not in (canonical_lower, canonical_lower.replace(" ", ""))
    )

    # ── Area-specific resolution ──────────────────────────────────────────────
    # MMT area URLs need: locusId=CTxxx, type=area, mmAreaTag=AreaName|CODE,
    # searchText=AreaName, semanticResults=true
    area_city_code = ""    # CT-format code for area searches
    mm_area_tag    = ""    # e.g. "Andheri East|ARANDH"

    if is_area:
        # Map RG codes → CT codes (MMT uses CT prefix for area-level searches)
        _RG_TO_CT = {
            "RGMUM": "CTBOM", "RGNCR": "CTDEL", "RGBLR": "CTBLR",
            "RGHYD": "CTHYD", "RGMAA": "CTMAA", "RGCCU": "CTCCU",
            "RGPNQ": "CTPNQ", "RGAMD": "CTAMD", "RGGOI": "CTGOI",
            "RGJPR": "CTJPR", "RGBDQ": "CTBDQ", "RGSTV": "CTSTV",
            "RGCOK": "CTCOK", "RGLKO": "CTLKO", "RGIXC": "CTIXC",
            "RGNAG": "CTNAG", "RGCJB": "CTCJB", "RGVTZ": "CTVTZ",
            "RGBBI": "CTBBI",
        }
        area_city_code = _RG_TO_CT.get(city_code, city_code)

        # Use LLM to get the MMT area tag code
        area_display_name = matched_key.title()  # e.g. "Andheri East"
        try:
            _groq_key = os.getenv("GROQ_API_KEY")
            if _groq_key and Groq:
                _gc = Groq(api_key=_groq_key)
                _prompt = (
                    f"MakeMyTrip uses internal area tag codes for hotel searches.\n"
                    f"The format is a short uppercase code derived from the area name.\n"
                    f"Examples:\n"
                    f"  Andheri East → ARANDH\n"
                    f"  Paharganj → PHRGNJ\n"
                    f"  Connaught Place → CNTPLC\n"
                    f"  Bandra West → ARBNDW\n"
                    f"  Koramangala → KRMGLA\n"
                    f"  Whitefield → WHTFLD\n"
                    f"  Vashi → ARVASH\n"
                    f"  Karol Bagh → KRLBGH\n"
                    f"  MG Road → ARMGRD\n"
                    f"  Hitech City → HTCHCT\n"
                    f"\nWhat is the MakeMyTrip area tag code for: \"{area_display_name}\" in {canonical_city_name}?\n"
                    f"Reply with ONLY the code (e.g. ARANDH). Nothing else."
                )
                _resp = _gc.chat.completions.create(
                    model="llama-3.3-70b-versatile",
                    messages=[{"role": "user", "content": _prompt}],
                    max_tokens=20, temperature=0,
                )
                _area_code = _resp.choices[0].message.content.strip().upper()
                # Sanitize: only alphanumeric, 3-10 chars
                _area_code = re.sub(r'[^A-Z0-9]', '', _area_code)
                if 3 <= len(_area_code) <= 10:
                    mm_area_tag = f"{area_display_name}|{_area_code}"
                    print(f"  ✅ MMT area tag resolved: {area_display_name} → {mm_area_tag}")
        except Exception as _area_err:
            print(f"  ⚠️ MMT area tag resolution failed: {_area_err}")

    if is_area:
        search_text = matched_key.title()   # just "Andheri East", not "Andheri East, Mumbai"
    else:
        search_text = canonical_city_name

    return {
        "mmt_city_code":   area_city_code if is_area else city_code,
        "locus_type":      "city",
        "type":            "area" if is_area else "city",
        "country":         country,
        "mmt_search_text": search_text,
        "is_area":         is_area,
        "region_code":     city_code,
        "area_tag_id":     "",
        "mm_area_tag":     mm_area_tag,
    }


def get_mmt_params_from_llm(meeting_location: str, dest_iata: str) -> dict:
    """Wrapper kept for API compatibility — now uses hardcoded table, no LLM."""
    cache_key = f"{meeting_location}_{dest_iata}".lower().strip()
    if cache_key in _mmt_city_cache:
        cached = _mmt_city_cache[cache_key]
        # Invalidate old cache entries that lack mm_area_tag field
        if cached.get("area_tag_id") or "mm_area_tag" not in cached:
            del _mmt_city_cache[cache_key]
        else:
            return cached
    result = _resolve_mmt_locus(meeting_location, dest_iata)
    _mmt_city_cache[cache_key] = result
    _save_mmt_cache(_mmt_city_cache)
    return result


def get_expert_mmt_url(meeting_location, dest_iata, check_in_dt, check_out_dt, guests, rooms):
    cache_key = f"expert_{meeting_location.lower().strip()}_{guests}_{rooms}"
    if cache_key in _mmt_city_cache:
        if isinstance(_mmt_city_cache[cache_key], dict) and "url" in _mmt_city_cache[cache_key]:
            return _mmt_city_cache[cache_key]["url"]
    api_key = os.environ.get("GROQ_API_KEY")
    if not (api_key and Groq):
        params = get_mmt_params_from_llm(meeting_location, dest_iata)
        return get_live_mmt_url(check_in_dt, check_out_dt, guests, meeting_location,
                                params.get("mmt_city_code", dest_iata), rooms=rooms)
    try:
        client = Groq(api_key=api_key)
        mmt_ci = check_in_dt.strftime("%m%d%Y")
        mmt_co = check_out_dt.strftime("%m%d%Y")
        prompt = f"""You are an expert on MakeMyTrip (MMT) internal hotel search URL parameters.
Generate a working MMT hotel listing URL.

CITY CODES: International→"CT"+abbrev (CTLONDO,CTDUBAI,CTSINGAP,CTNEWYORK,CTPARIS,CTTOKYO)
            Indian→"RG"+abbrev (RGMUM,RGDEL,RGBLR,RGHYD,RGMAA,RGCCU,RGPNQ,RGAMD,RGGOI,RGJPR)

RULES:
1. Derive city from meeting_location, NOT dest_iata
2. roomStayQualifier: "{{adults_per_room}}e0e" per room joined by "_"
3. rsc: "{{rooms}}e{{total_guests}}e0e"
4. checkin/checkout: MMDDYYYY
5. ALWAYS include _uCurrency=INR, reference=hotel

INPUT: meeting_location="{meeting_location}", dest_iata="{dest_iata}", check_in="{mmt_ci}", check_out="{mmt_co}", guests={guests}, rooms={rooms}

OUTPUT — ONLY valid JSON:
{{"mmt_city_code":"RGNCR","country":"IN","locus_type":"region","type":"area","room_stay_qualifier":"1e0e","rsc":"1e1e0e","url":"https://www.makemytrip.com/hotels/hotel-listing/?..."}}"""
        response = client.chat.completions.create(
            model="llama-3.1-8b-instant",
            messages=[{"role": "user", "content": prompt}],
            temperature=0, response_format={"type": "json_object"}
        )
        data = json.loads(response.choices[0].message.content.strip())
        if data and "url" in data:
            _mmt_city_cache[cache_key] = data
            _save_mmt_cache(_mmt_city_cache)
            return data["url"]
    except Exception as e:
        print(f"Expert Groq API error: {e}")
    params = get_mmt_params_from_llm(meeting_location, dest_iata)
    return get_live_mmt_url(check_in_dt, check_out_dt, guests, meeting_location,
                            params.get("mmt_city_code", dest_iata), rooms=rooms)


def get_iata_from_city(city_query: str) -> str:
    if not city_query or len(city_query) < 2:
        return "BOM"
    city_query = city_query.strip().lower()
    _aliases = {
        "bombay": "mumbai", "madras": "chennai", "calcutta": "kolkata",
        "cochin": "kochi", "baroda": "vadodara", "banaras": "varanasi",
        "trivandrum": "thiruvananthapuram", "vizag": "visakhapatnam",
        "mangalore": "mangaluru", "bengaluru": "bangalore",
    }
    city_query = _aliases.get(city_query, city_query)
    if len(city_query) == 3 and city_query.upper() in airport_db:
        return city_query.upper()
    matches = [(code, data) for code, data in airport_db.items()
               if data.get('city', '').lower() == city_query]
    if matches:
        for code, data in matches:
            if data.get('country') == 'IN':
                return code
        return matches[0][0]
    partial_city = [(code, data) for code, data in airport_db.items()
                    if city_query in data.get('city', '').lower()]
    if partial_city:
        for code, data in partial_city:
            if data.get('country') == 'IN':
                return code
        return partial_city[0][0]
    partial_name = [(code, data) for code, data in airport_db.items()
                    if city_query in data.get('name', '').lower()]
    if partial_name:
        for code, data in partial_name:
            if data.get('country') == 'IN':
                return code
        return partial_name[0][0]
    return city_query.upper()


_LOCUS_TO_CITY_NAME = {
    # Indian cities
    "RGMUM": "Mumbai", "RGNCR": "Delhi", "RGBLR": "Bangalore",
    "RGHYD": "Hyderabad", "RGMAA": "Chennai", "RGCCU": "Kolkata",
    "RGPNQ": "Pune", "RGAMD": "Ahmedabad", "RGGOI": "Goa",
    "RGJPR": "Jaipur", "RGBDQ": "Vadodara", "RGSTV": "Surat",
    "RGCOK": "Kochi", "RGLKO": "Lucknow", "RGIXC": "Chandigarh",
    "RGNAG": "Nagpur", "RGCJB": "Coimbatore", "RGVTZ": "Visakhapatnam",
    "RGBBI": "Bhubaneswar",
    # International cities
    "CTLONDO": "London", "CTDUBAI": "Dubai", "CTSINGAP": "Singapore",
    "CTNEWYORK": "New York", "CTPARIS": "Paris", "CTFRANKF": "Frankfurt",
    "CTAMSTER": "Amsterdam", "CTZURICH": "Zurich", "CTTOKYO": "Tokyo",
    "CTOSAKA": "Osaka", "CTHONGKO": "Hong Kong", "CTBANGKO": "Bangkok",
    "CTKUALAL": "Kuala Lumpur", "CTSYDNEY": "Sydney", "CTMELB": "Melbourne",
    "CTTORONT": "Toronto", "CTCHICAG": "Chicago", "CTSANFRA": "San Francisco",
    "CTLOSANG": "Los Angeles", "CTDOHA": "Doha", "CTABUDHA": "Abu Dhabi",
    "CTMUSCAT": "Muscat",
}

def _locus_to_city_name(locus_id: str) -> str:
    """Return the human-readable city name for an MMT locus ID."""
    return _LOCUS_TO_CITY_NAME.get(locus_id, locus_id)


def make_mmt_link(
    origin: str,
    destination: str,
    travel_date,
    travel_count: int,
    arrival_cutoff_dt=None,
    airline_code: str = "",
    meeting_location: str = None,
    is_round_trip: bool = False,
    return_date=None,
    rooms: int = 1,
) -> str:
    """
    Bulletproof MMT URL builder.
    - meeting_location → hotel listing URL
    - otherwise        → flight search URL (round-trip when is_round_trip=True)
    NEVER raises; always returns a valid MMT URL as fallback.
    """
    # ── Hotel URL ──────────────────────────────────────────────────────────
    if meeting_location:
        try:
            _td       = travel_date if hasattr(travel_date, "strftime") else datetime.now().date()
            check_out = _td + timedelta(days=1)
            mmt_ci    = _td.strftime("%m%d%Y")
            mmt_co    = check_out.strftime("%m%d%Y")
            mmt_params    = _resolve_mmt_locus(meeting_location, str(destination or ""))
            city_locus_id = mmt_params["mmt_city_code"]
            country       = mmt_params["country"]
            search_text   = mmt_params["mmt_search_text"]
            guests  = max(1, int(travel_count or 1))
            _rooms  = max(1, int(rooms or 1))
            _base, _ex = divmod(guests, _rooms)
            qualifier  = "_".join([f"{_base + (1 if i < _ex else 0)}e0e" for i in range(_rooms)])
            rsc        = f"{_rooms}e{guests}e0e"
            mmt_type = mmt_params.get("type", "city")
            params = {
                "checkin":           mmt_ci,
                "checkout":          mmt_co,
                "locusId":           city_locus_id,
                "locusType":         "city",
                "city":              city_locus_id,
                "country":           country,
                "searchText":        search_text,
                "roomStayQualifier": qualifier,
                "_uCurrency":        "INR",
                "reference":         "hotel",
                "type":              mmt_type,
                "rsc":               rsc,
            }
            if mmt_params.get("is_area"):
                if mmt_params.get("mm_area_tag"):
                    params["mmAreaTag"] = mmt_params["mm_area_tag"]
                params["semanticResults"] = "true"
            return f"https://www.makemytrip.com/hotels/hotel-listing/?{urllib.parse.urlencode(params)}"
        except Exception as e:
            print(f"MMT hotel link error: {e}")
            return "https://www.makemytrip.com/hotels/"

    # ── Flight URL ─────────────────────────────────────────────────────────
    try:
        _org  = str(origin      or "BOM").strip().upper()[:3]
        _dst  = str(destination or "DEL").strip().upper()[:3]
        _tc   = max(1, int(travel_count or 1))

        # Detect international route — MMT dates shift for intl
        _is_mmt_intl = is_international(_org, _dst)

        try:
            _td = travel_date if hasattr(travel_date, "strftime") else datetime.now().date()
            # International outbound: depart the day BEFORE the meeting
            # so the traveller arrives on or before meeting day
            if _is_mmt_intl:
                _td_mmt = _td - timedelta(days=1)
            else:
                _td_mmt = _td
            formatted_date = _td_mmt.strftime("%d/%m/%Y")
        except Exception:
            formatted_date = datetime.now().strftime("%d/%m/%Y")

        if is_round_trip and return_date is not None:
            try:
                _rd = return_date if hasattr(return_date, "strftime") else datetime.now().date()
                # International return: fly the day AFTER the meeting ends
                if _is_mmt_intl:
                    _rd_mmt = _rd + timedelta(days=1)
                else:
                    _rd_mmt = _rd
                ret_formatted = _rd_mmt.strftime("%d/%m/%Y")
            except Exception:
                ret_formatted = formatted_date
            itinerary = f"{_org}-{_dst}-{formatted_date}_{_dst}-{_org}-{ret_formatted}"
            trip_type = "R"
        else:
            itinerary = f"{_org}-{_dst}-{formatted_date}"
            trip_type = "O"

        params = [
            f"itinerary={itinerary}",
            f"tripType={trip_type}",
            f"paxType=A-{_tc}_C-0_I-0",
            f"intl={'true' if _is_mmt_intl else 'false'}",
            "cabinClass=E",
        ]

        if arrival_cutoff_dt and hasattr(arrival_cutoff_dt, "hour"):
            try:
                cutoff_hour = int(arrival_cutoff_dt.hour)
                buckets = []
                if cutoff_hour > 0:  buckets.append("00-06")
                if cutoff_hour > 6:  buckets.append("06-12")
                if cutoff_hour > 12: buckets.append("12-18")
                if cutoff_hour > 18: buckets.append("18-24")
                if buckets:
                    params.append(f"arrTime={','.join(buckets)}")
            except Exception:
                pass

        if airline_code and isinstance(airline_code, str) and len(airline_code.strip()) == 2:
            params.append(f"selectedAirlines={airline_code.strip().upper()}")

        return f"https://www.makemytrip.com/flight/search?{'&'.join(params)}"

    except Exception as e:
        print(f"MMT flight link error: {e}")
        return "https://www.makemytrip.com/flight/search"


# ==============================================================================
#  HOTEL ENGINE
# ==============================================================================

_hotel_key_index = 0

def get_serp_hotel_results_with_fallback(params):
    global _hotel_key_index
    all_keys = [
        os.getenv("SERP_API_1"), os.getenv("SERP_API_2"), os.getenv("SERP_API_3"),
        os.getenv("SERP_API_4"), os.getenv("SERP_API_5"), os.getenv("SERP_API_6"),
        os.getenv("SERP_API_7"), os.getenv("SERP_API_8"), os.getenv("SERP_API_9")
    ]
    keys = [k for k in all_keys if k]
    if not keys:
        return {}
    n = len(keys); start = _hotel_key_index % n; _hotel_key_index = (start + 1) % n
    for i in range(n):
        key = keys[(start + i) % n]
        key_num = all_keys.index(key) + 1 if key in all_keys else "?"
        try:
            params["api_key"] = key
            results = GoogleSearch(params).get_dict()
            if "error" in results:
                print(f"⚠️ Key {key_num} Quota Error (Hotels): {results['error']} — trying next…")
                continue
            print(f"  ✅ Key {key_num} used for Hotels")
            return results
        except Exception as e:
            print(f"⚠️ Key {key_num} Connection Error (Hotels): {e} — trying next…")
            continue
    return {}


def _hvarint(value: int) -> bytes:
    out = []
    while value > 0x7F:
        out.append((value & 0x7F) | 0x80); value >>= 7
    out.append(value & 0x7F)
    return bytes(out)

def _hvfield(field_num: int, value: int) -> bytes:
    return _hvarint((field_num << 3) | 0) + _hvarint(value)

def _hlfield(field_num: int, data: bytes) -> bytes:
    return _hvarint((field_num << 3) | 2) + _hvarint(len(data)) + data

def _date_msg(year: int, month: int, day: int) -> bytes:
    return _hvfield(1, year) + _hvfield(2, month) + _hvfield(3, day)

def build_hotels_ts(check_in: date, check_out: date, currency: str = "INR",
                    adults: int = 2, rooms: int = 1) -> str:
    nights = (check_out - check_in).days; rooms = int(rooms); adults = int(adults)
    guest_list  = b'\x0a\x02\x08\x03' * adults
    rooms_count = _hvfield(2, rooms)
    guests_block = guest_list + rooms_count
    ci_msg = _date_msg(check_in.year,  check_in.month,  check_in.day)
    co_msg = _date_msg(check_out.year, check_out.month, check_out.day)
    date_range = _hlfield(1, ci_msg) + _hlfield(2, co_msg) + _hvfield(3, nights)
    inner  = _hlfield(2, date_range) + _hlfield(6, b'\x08\x02')
    field3 = _hlfield(1, b'\x1a\x00') + _hlfield(2, inner)
    curr_block = _hlfield(1, _hlfield(7, currency.encode())) + b'\x1a\x00'
    proto = (_hvfield(1, 1) + _hlfield(2, guests_block) + _hlfield(3, field3) + _hlfield(5, curr_block))
    return base64.urlsafe_b64encode(proto).decode().rstrip("=")

def build_hotels_ap(adults: int, rooms: int = 1) -> str:
    rooms = max(1, int(rooms)); adults = max(1, int(adults))
    base, extra = divmod(adults, rooms); ap_msg = b''
    for i in range(rooms):
        guests_in_room = base + (1 if i < extra else 0)
        ap_msg += _hlfield(1, _hvfield(2, guests_in_room))
    return base64.urlsafe_b64encode(ap_msg).decode().rstrip("=")

def build_hotels_url(destination, check_in, check_out, adults, rooms=1, currency="INR") -> str:
    ts = build_hotels_ts(check_in, check_out, currency, adults=adults, rooms=rooms)
    ap = build_hotels_ap(adults, rooms=rooms)
    params = {"q": destination, "ts": ts, "ap": ap, "hl": "en", "gl": "in",
              "curr": currency, "_cb": str(int(time.time()))}
    return f"https://www.google.com/travel/hotels?{urllib.parse.urlencode(params)}"

_STAR_SUFFIX_RE = re.compile(r'\s*\d+\s*[-–]?\s*[Ss]tars?\b', re.VERBOSE)

def _strip_star_suffix(name: str) -> str:
    cleaned = _STAR_SUFFIX_RE.sub("", name).strip()
    return re.sub(r"\s{2,}", " ", cleaned)

def _clean_hotel_name(name: str) -> str:
    return _strip_star_suffix(name.strip())

def build_hotel_direct_url(hotel_name: str, check_in: date, check_out: date,
                           adults: int, currency: str = "INR") -> str:
    ts = build_hotels_ts(check_in, check_out, currency, adults=adults)
    ap = build_hotels_ap(adults)
    params = {"q": hotel_name, "ts": ts, "ap": ap, "hl": "en", "gl": "in",
              "curr": currency, "_cb": str(int(time.time()))}
    return f"https://www.google.com/travel/hotels?{urllib.parse.urlencode(params)}"

def clean_hotel_name(raw_name: str, destination: str = "", meeting_loc: str = "") -> str:
    if not raw_name: return ""
    raw = raw_name.strip()
    vague_pattern = r'(?i)\b(?:at|near|by|with|close to|opposite)\b'
    has_vague = bool(re.search(vague_pattern, raw))
    if has_vague:
        text = re.sub(vague_pattern + r'.*$', '', raw)
        text = re.sub(r'\(.*?\)', '', text); text = re.sub(r'-.*$', '', text)
        clean_text = text.strip()
        if destination and destination.lower() not in clean_text.lower():
            return f"{clean_text}, {destination}"
        return clean_text
    if destination and destination.lower() not in raw.lower():
        return f"{raw}, {destination}"
    return raw

def get_live_hotel_url(hotel_obj, ci_dt, co_dt, passenger_count,
                       meeting_loc="", city_name="", budget_min=None,
                       budget_max=None, rooms=1) -> str:
    raw_name   = hotel_obj.get("name", "").strip()
    clean_name = clean_hotel_name(raw_name, destination=city_name, meeting_loc=meeting_loc)
    rooms      = max(1, int(rooms))
    url = build_hotels_url(destination=clean_name, check_in=ci_dt, check_out=co_dt,
                           adults=int(passenger_count), rooms=rooms)
    print(f"  🔗 [Book URL] '{raw_name}' → q='{clean_name}' adults={passenger_count} rooms={rooms}")
    return url

def get_airport_city(iata_code: str) -> str:
    airport_city_map = {
        "DEL": "New Delhi",  "BOM": "Mumbai",    "BLR": "Bengaluru",
        "MAA": "Chennai",    "HYD": "Hyderabad",  "CCU": "Kolkata",
        "GOI": "Goa",        "PNQ": "Pune",       "AMD": "Ahmedabad",
        "JAI": "Jaipur",     "COK": "Kochi",      "LKO": "Lucknow",
        "TRV": "Trivandrum", "IXE": "Mangaluru",  "VNS": "Varanasi",
        "BDQ": "Vadodara",   "STV": "Surat",
    }
    return airport_city_map.get(iata_code.strip().upper(), iata_code.upper())

def get_live_mmt_url(check_in, check_out, guests: int, meeting_location: str,
                     dest_iata: str = "", rooms: int = 1) -> str:
    mmt_ci = check_in.strftime("%m%d%Y"); mmt_co = check_out.strftime("%m%d%Y")
    rooms = max(1, int(rooms)); guests = max(1, int(guests))
    base_per_room, extra = divmod(guests, rooms)
    qualifiers = [f"{base_per_room + (1 if i < extra else 0)}e0e" for i in range(rooms)]
    room_stay_qualifier = "_".join(qualifiers)
    rsc = f"{rooms}e{guests}e0e"
    mmt_params    = _resolve_mmt_locus(meeting_location, dest_iata)
    city_locus_id = mmt_params["mmt_city_code"]
    country       = mmt_params["country"]
    search_text   = mmt_params["mmt_search_text"]
    mmt_type      = mmt_params.get("type", "city")
    params = {
        "checkin": mmt_ci, "checkout": mmt_co, "locusId": city_locus_id,
        "locusType": "city", "city": city_locus_id,
        "country": country, "searchText": search_text,
        "roomStayQualifier": room_stay_qualifier, "_uCurrency": "INR",
        "reference": "hotel", "type": mmt_type, "rsc": rsc,
    }
    # Area-specific params: mmAreaTag + semanticResults for neighbourhood filtering
    if mmt_params.get("is_area"):
        if mmt_params.get("mm_area_tag"):
            params["mmAreaTag"] = mmt_params["mm_area_tag"]
        params["semanticResults"] = "true"
    return f"https://www.makemytrip.com/hotels/hotel-listing/?{urllib.parse.urlencode(params)}"


def _parse_hotels_from_properties(properties: list) -> list:
    hotels = []
    for prop in properties:
        try:
            name      = prop.get("name", "Unknown Hotel")
            raw_dist  = prop.get("distance", "")
            dist_match = re.search(r"(\d+\.?\d*)", raw_dist)
            if dist_match:
                distance_km = float(dist_match.group(1)); distance_explicit = True
            else:
                distance_km = 1.0; distance_explicit = False
            rating  = prop.get("overall_rating", 0.0)
            reviews = prop.get("reviews", 0)
            rate_per_night = prop.get("rate_per_night", {}) or {}
            total_rate     = prop.get("total_rate",     {}) or {}
            price_val = (rate_per_night.get("extracted_before_taxes_fees") or
                         total_rate.get("extracted_before_taxes_fees") or
                         rate_per_night.get("extracted_lowest") or
                         total_rate.get("extracted_lowest"))
            price_val = float(price_val) if price_val else 999999.0
            if price_val == float('inf'):
                continue
            if rate_per_night.get("extracted_lowest"):
                price_fmt_str = rate_per_night.get("lowest", f"₹{int(price_val):,}")
            elif total_rate.get("extracted_lowest"):
                price_fmt_str = total_rate.get("lowest", f"₹{int(price_val):,}")
            else:
                price_fmt_str = "Check Rates"
            images = prop.get("images", []); thumb = None
            if isinstance(images, list) and len(images) > 0:
                first_img = images[0]
                if isinstance(first_img, dict):
                    thumb = first_img.get("thumbnail") or first_img.get("original_image")
                elif isinstance(first_img, str):
                    thumb = first_img
                if not (isinstance(thumb, str) and (thumb.startswith("http") or thumb.startswith("data:image"))):
                    thumb = None
            hotels.append({
                "name": name, "price_raw": price_val, "price_fmt": price_fmt_str,
                "rating": rating, "reviews": reviews, "thumbnail": thumb,
                "distance_val": distance_km, "distance_fmt": f"{distance_km} km away" if distance_explicit else "Near meeting location",
                "distance_explicit": distance_explicit,
                "property_token": prop.get("property_token", ""), "api_link": prop.get("link", ""),
            })
        except Exception as e:
            print(f"  ⚠️ Error parsing hotel '{prop.get('name', '?')}': {e}")
            continue
    return hotels


def _filter_and_sort(hotels, max_dist, min_b, max_b, min_r=3.5, max_r=5.0, traveler_type="Employee"):
    filtered = []
    office_hotels = {"ginger mumbai andheri east", "ginger mumbai airport",
                     "holiday inn express kolkata new town", "holiday inn express kolkata newtown"}
    for h in hotels:
        rating   = h.get('rating', 0.0)
        is_office = h["name"].lower().strip() in office_hotels
        if not is_office and (rating < min_r or rating > max_r):
            continue
        dist = h['distance_val']; explicit = h.get('distance_explicit', False)
        if not is_office and explicit and dist > max_dist:
            continue
        if h['price_raw'] <= 0:
            continue
        if not (min_b <= h['price_raw'] <= max_b):
            continue
        filtered.append(h)
    if traveler_type == "Management":
        filtered.sort(key=lambda x: (-x.get('rating', 0.0), -x['price_raw']))
    else:
        filtered.sort(key=lambda x: (x['price_raw'], -x.get('rating', 0.0)))
    return filtered


def _serpapi_hotel_search(query, check_in_date, check_out_date, passenger_count,
                          min_price=None, max_price=None, location="", _actual_adults=None) -> list:
    adults_val = str(int(_actual_adults)) if _actual_adults and int(_actual_adults) > 1 else "1"
    params = {
        "engine": "google_hotels", "q": query, "check_in_date": check_in_date,
        "check_out_date": check_out_date, "adults": adults_val, "currency": "INR",
        "gl": "in", "hl": "en", "sort_by": "8", "api_key": "DUMMY_KEY",
    }
    if location:   params["location"]  = location
    if min_price:  params["min_price"] = str(min_price)
    if max_price:  params["max_price"] = str(max_price)
    try:
        print(f"  🔍 SerpAPI query='{query}' | pax={passenger_count} | ₹{min_price}–{max_price}")
        results = get_serp_hotel_results_with_fallback(params)
        if "error" in results:
            return []
        props = results.get("properties", [])
        print(f"  → {len(props)} raw properties")
        return props
    except Exception as e:
        print(f"  ❌ SerpAPI error: {e}")
        return []

def _enrich_prices_with_actual_adults(hotels_list, location_anchor, check_in_date, check_out_date, adults):
    if not hotels_list or int(adults) <= 1:
        return
    props = _serpapi_hotel_search(location_anchor, check_in_date, check_out_date,
                                  passenger_count=adults, location=location_anchor, _actual_adults=adults)
    if not props:
        return
    price_map = {}
    for prop in props:
        name  = prop.get("name", "").lower().strip()
        rate  = prop.get("rate_per_night", {}) or {}
        total = prop.get("total_rate",     {}) or {}
        raw   = rate.get("extracted_lowest") or total.get("extracted_lowest")
        fmt   = rate.get("lowest")           or total.get("lowest")
        if raw and name:
            price_map[name] = {"price_raw": float(raw), "price_fmt": fmt or f"₹{int(raw):,}"}
    for h in hotels_list:
        key = h["name"].lower().strip()
        if key in price_map:
            h["price_raw"] = price_map[key]["price_raw"]
            h["price_fmt"] = price_map[key]["price_fmt"]


@st.cache_data
def search_hotels(location, check_in_date, check_out_date,
                  budget_min=0, budget_max=float('inf'),
                  traveler_type="Employee", passenger_count=1,
                  meeting_location="", dest_airport_iata=None, _cache_bust=1) -> dict:
    city_name  = get_airport_city(dest_airport_iata) if dest_airport_iata else ""
    anchor_loc = location
    if city_name and city_name.lower() not in location.lower():
        anchor_loc = f"{location}, {city_name}"
    min_rating, max_rating = 3.5, 5.0
    props = _serpapi_hotel_search(anchor_loc, check_in_date, check_out_date, passenger_count,
                                  min_price=budget_min, max_price=budget_max, location=anchor_loc)
    if props:
        all_hotels = _parse_hotels_from_properties(props)
        results    = _filter_and_sort(all_hotels, 4.0, budget_min, budget_max, min_rating, max_rating, traveler_type)
        if not results:
            results = _filter_and_sort(all_hotels, 5.0, budget_min, budget_max, min_rating, max_rating, traveler_type)
        if results:
            _enrich_prices_with_actual_adults(results, anchor_loc, check_in_date, check_out_date, passenger_count)
            return {"hotels": results[:30], "search_area": "meeting", "fallback_city": None}
    return {"hotels": [], "search_area": "none", "fallback_city": None}


def search_hotels_budget_fallback(meeting_location, check_in_date, check_out_date,
                                  total_count, min_budget, dest_airport_iata=None) -> list:
    max_budget = min_budget + 4000; MIN_RATING = 3.0
    city_name  = get_airport_city(dest_airport_iata) if dest_airport_iata else ""
    anchor     = f"{meeting_location}, {city_name}" if city_name and city_name.lower() not in meeting_location.lower() else meeting_location
    props  = _serpapi_hotel_search(anchor, check_in_date, check_out_date, total_count,
                                   min_price=min_budget, max_price=max_budget, location=anchor)
    if not props:
        return []
    parsed  = _parse_hotels_from_properties(props)
    results = []
    office_hotels = {"ginger mumbai andheri east", "ginger mumbai airport",
                     "holiday inn express kolkata new town", "holiday inn express kolkata newtown"}
    for h in parsed:
        price = h.get("price_raw", 0); rating = h.get("rating", 0.0)
        is_office = h["name"].lower().strip() in office_hotels
        if is_office: rating = 5.0
        if min_budget <= price <= max_budget and rating >= MIN_RATING and price > 0:
            results.append(h)
    results.sort(key=lambda x: x["price_raw"])
    return results[:10]


def search_hotels_coordinated(meeting_location, check_in_date, check_out_date,
                               m_count, e_count, dest_airport_iata=None, _cache_bust=1) -> dict:
    MGMT_MIN, MGMT_MAX     = 3500, 19000
    EMP_MIN,  EMP_MAX      = 2500, 3500
    MIN_RATING, MAX_RATING = 3.5, 5.0
    MEETING_RADIUS         = 5.0
    city_name = get_airport_city(dest_airport_iata) if dest_airport_iata else ""
    if city_name and city_name.lower() not in meeting_location.lower():
        location_anchor = f"{meeting_location}, {city_name}"
    else:
        location_anchor = meeting_location
    all_props  = _serpapi_hotel_search(location_anchor, check_in_date, check_out_date,
                                       passenger_count=1, min_price=EMP_MIN, max_price=MGMT_MAX,
                                       location=location_anchor)
    all_parsed = _parse_hotels_from_properties(all_props)
    near_hotels = _filter_and_sort(all_parsed, MEETING_RADIUS, EMP_MIN, MGMT_MAX,
                                   MIN_RATING, MAX_RATING, traveler_type="Combined")
    shared_results = []; mgmt_results = []
    for h in near_hotels:
        price = h.get("price_raw", 0)
        if EMP_MIN <= price <= EMP_MAX:
            shared_results.append(h)
        elif MGMT_MIN < price <= MGMT_MAX:
            mgmt_results.append(h)
    EXPAND_RADIUS = 4.0; shared_from_expand = False
    if not shared_results:
        expanded = _filter_and_sort(all_parsed, EXPAND_RADIUS, EMP_MIN, MGMT_MAX,
                                    MIN_RATING, MAX_RATING, traveler_type="Combined")
        for h in expanded:
            price = h.get("price_raw", 0)
            if EMP_MIN <= price <= EMP_MAX:
                shared_results.append(h)
            elif MGMT_MIN < price <= MGMT_MAX:
                if h not in mgmt_results:
                    mgmt_results.append(h)
        if shared_results:
            shared_from_expand = True
    shared_names = {h["name"].lower().strip() for h in shared_results}
    mgmt_results = [h for h in mgmt_results if h["name"].lower().strip() not in shared_names]
    if shared_results:
        _enrich_prices_with_actual_adults(shared_results, location_anchor, check_in_date, check_out_date, m_count + e_count)
    if mgmt_results:
        _enrich_prices_with_actual_adults(mgmt_results, location_anchor, check_in_date, check_out_date, m_count)
    return {
        "mgmt": {"hotels": mgmt_results[:30], "search_area": "meeting" if mgmt_results else "none"},
        "emp":  {"hotels": [],                "search_area": "none"},
        "shared_hotels":      shared_results[:30],
        "shared_area":        "expand" if shared_from_expand else ("meeting" if shared_results else "none"),
        "shared_from_expand": shared_from_expand,
        "fallback_city":      None,
    }


# ==============================================================================
#  EMAIL AGENT
# ==============================================================================

RECIPIENT   = "krishna.kotecha@si2tech.com"
SENDER_NAME = "Travel Desk, SI2Tech"


class EmailAgent:

    def __init__(self):
        api_key = os.getenv("GROQ_API_KEY")
        if not api_key:
            raise ValueError("GROQ_API_KEY not found in environment variables")
        self.client = Groq(api_key=api_key)
        self.embedded_images = []

    @staticmethod
    def _fmt_price(val) -> str:
        try:    return f"&#8377;{int(float(val)):,}"
        except: return str(val)

    @staticmethod
    def _fmt_price_plain(val) -> str:
        try:    return f"Rs.{int(float(val)):,}"
        except: return str(val)

    def _process_image_for_email(self, url: str, raw_bytes: bytes, use_cid: bool) -> str:
        if not url or not str(url).startswith("http"):
            return ""
        if use_cid:
            cid = f"img_{uuid.uuid4().hex[:8]}@si2tech.com"
            self.embedded_images.append((cid, url, raw_bytes))
            return f"cid:{cid}"
        return url

    @staticmethod
    def _flight_url(flight: dict, context: dict) -> str:
        td = context.get("travel_date")
        ds = td.strftime("%Y-%m-%d") if isinstance(td, datetime) else str(td)
        # Use arr_time_raw if available (guaranteed clean), else clean arr_time
        arr_t = flight.get("arr_time_raw") or _clean_time_str(flight.get("arr_time", ""))
        return make_google_flights_link(
            origin=context["origin_iata"], destination=context["destination_iata"],
            travel_date=ds, airline_name=flight.get("airline", ""),
            travel_count=context.get("travel_count", 1),
            dep_time_str=flight.get("dep_time", ""),
            arr_time_str=arr_t,
            price_per_person=int(flight.get("price", 0) / max(context.get("travel_count", 1), 1)),
        )

    @staticmethod
    def _hotel_url(h: dict, context: dict, pax=None, bmin=None, bmax=None) -> str:
        ci, co = context.get("check_in"), context.get("check_out")
        if not ci or not co:
            return "#"
        return get_live_hotel_url(hotel_obj=h, ci_dt=ci, co_dt=co,
                                  passenger_count=pax if pax is not None else context.get("travel_count", 1),
                                  meeting_loc=context.get("meeting_location", ""),
                                  city_name=context.get("dest_city", ""))

    def _generate_prose(self, context: dict) -> dict:
        prompt = f"""
You are a professional corporate travel assistant.
Generate FOUR plain-text blocks for a travel approval email. Return ONLY valid JSON.

1. "greeting"  — e.g. "Dear Krishna,"
2. "intro"     — 2-3 sentences. State this is a travel approval request from
                 {context['origin']} to {context['destination']} on {context['date_str']}.
                 Mention flight and hotel options are listed below.
3. "outro"     — 2-3 sentences. Ask to review and approve. Note fares may change.
4. "signoff"   — e.g. "Warm regards,\\nTravel Desk, SI2Tech"

Rules: Plain text only. No markdown, no HTML. Return ONLY the JSON object.
{{"greeting": "...", "intro": "...", "outro": "...", "signoff": "..."}}
"""
        try:
            res = self.client.chat.completions.create(
                messages=[
                    {"role": "system", "content": "Corporate travel assistant. Return only valid JSON."},
                    {"role": "user",   "content": prompt},
                ],
                model="llama-3.3-70b-versatile", temperature=0.2,
            )
            match = re.search(r'\{.*\}', res.choices[0].message.content, re.DOTALL)
            if match:
                return json.loads(match.group())
        except Exception:
            pass
        return {
            "greeting": "Dear Krishna,",
            "intro": (f"I hope this message finds you well. I am writing to request approval for an upcoming "
                      f"business trip from {context['origin']} to {context['destination']} on {context['date_str']}. "
                      f"Please find the available flight and accommodation options below for your review."),
            "outro": ("Kindly review and approve at your earliest convenience as booking links are live "
                      "and fares are subject to change."),
            "signoff": "Warm regards,\nTravel Desk, SI2Tech",
        }

    _FONT   = "font-family: Segoe UI, Arial, sans-serif;"
    _BLUE   = "#1d4ed8"; _DKBLUE = "#1e3a8a"; _GREY = "#6b7280"
    _LGREY  = "#f8fafc"; _BORDER = "#e5e7eb"; _TEXT = "#111827"
    _MUTED  = "#374151"; _ORANGE = "#e37400"; _EMP  = "#1a73e8"
    _GREEN  = "#059669"

    def _th(self, label, align="center", width=""):
        w = f'width="{width}"' if width else ""
        return (f'<th {w} align="{align}" style="{self._FONT} padding:8px 12px; font-size:11px; '
                f'color:{self._GREY}; font-weight:600; text-transform:uppercase; letter-spacing:0.4px; '
                f'background:{self._LGREY}; border-bottom:2px solid {self._BORDER};">{label}</th>')

    def _book_btn(self, url, label="Book &rarr;"):
        return (f'<a href="{url}" style="display:inline-block; background:{self._BLUE}; color:#ffffff; '
                f'text-decoration:none; padding:6px 16px; font-size:12px; font-weight:600; {self._FONT}">'
                f'{label}</a>')

    def _flight_rows(self, flights, context, use_cid=False):
        rows = ""
        for i, f in enumerate(flights):
            url       = self._flight_url(f, context)
            raw_price = f.get('custom_price', f.get('price', 0))
            price     = self._fmt_price(raw_price)
            vendor    = f.get("booking_vendor", "MakeMyTrip")
            stops     = "Non-stop" if f.get("stops", 1) == 0 else f"{f.get('stops',1)} stop(s)"
            img_src   = self._process_image_for_email(f.get("thumbnail"), f.get("safe_img_data"), use_cid)
            logo_tag  = f'<img src="{img_src}" width="20" height="20" style="vertical-align:middle;border-radius:3px;margin-right:6px;">' if img_src else ""
            bg = "#ffffff" if i % 2 == 0 else "#f9fafb"
            td = f'style="{self._FONT} padding:10px 12px; border-bottom:1px solid {self._BORDER}; background:{bg}; font-size:13px;"'
            rows += f"""
<tr>
  <td {td} align="left">
    {logo_tag}<strong style="color:{self._TEXT};">{f.get('airline','').upper()}</strong><br>
    <span style="color:{self._GREY}; font-size:11px;">{f.get('flight_no','')} &nbsp;&middot;&nbsp; {stops} &nbsp;&middot;&nbsp; {f.get('duration','')}</span>
  </td>
  <td {td} align="center"><strong style="font-size:14px; color:{self._TEXT};">{f.get('dep_time','')}</strong><br><span style="color:{self._GREY}; font-size:11px;">{context.get('origin_iata','')}</span></td>
  <td {td} align="center" style="{self._FONT} padding:10px 6px; border-bottom:1px solid {self._BORDER}; background:{bg}; color:{self._GREY}; font-size:18px;">&#x2192;</td>
  <td {td} align="center"><strong style="font-size:14px; color:{self._TEXT};">{f.get('arr_time','')}</strong><br><span style="color:{self._GREY}; font-size:11px;">{context.get('destination_iata','')}</span></td>
  <td {td} align="center"><strong style="font-size:14px; color:{self._TEXT};">{price}</strong><br><span style="color:{self._GREY}; font-size:10px;">via {vendor}</span></td>
</tr>"""
        return rows

    def _hotel_rows(self, hotels, context, pax=None, bmin=None, bmax=None, use_cid=False):
        rows = ""
        for i, h in enumerate(hotels):
            url       = self._hotel_url(h, context, pax=pax, bmin=bmin, bmax=bmax)
            stars     = "&#9733;" * int(float(h.get("rating", 0)))
            raw_price = h.get('custom_price', h.get('price_raw', 0))
            price_fmt = self._fmt_price(raw_price) if h.get('custom_price') else h.get('price_fmt', 'N/A')
            vendor    = h.get("booking_vendor", "MakeMyTrip")
            bg        = "#ffffff" if i % 2 == 0 else "#f9fafb"
            td        = f'style="{self._FONT} padding:10px 12px; border-bottom:1px solid {self._BORDER}; background:{bg}; font-size:13px;"'
            img_src   = self._process_image_for_email(h.get("thumbnail"), h.get("safe_img_data"), use_cid)
            thumb_tag = f'<img src="{img_src}" width="80" height="80" style="vertical-align:middle;border-radius:4px;margin-right:10px;">' if img_src else ""
            rows += f"""
<tr>
  <td {td} align="left">
    <table cellpadding="0" cellspacing="0" border="0"><tr>
      <td style="vertical-align:middle;">{thumb_tag}</td>
      <td style="vertical-align:middle;">
        <strong style="color:{self._TEXT};">{h.get('name','').upper()}</strong><br>
        <span style="color:#f59e0b; font-size:12px;">{stars}</span>
        <span style="color:{self._GREY}; font-size:11px;"> {h.get('rating','N/A')} Stars &nbsp;&middot;&nbsp; {h.get('distance_fmt','')}</span>
      </td>
    </tr></table>
  </td>
  <td {td} align="center"><strong style="font-size:14px; color:{self._TEXT};">{price_fmt}</strong><br><span style="color:{self._GREY}; font-size:10px;">via {vendor}</span><br><span style="color:{self._GREY}; font-size:11px;">per night</span></td>
</tr>"""
        return rows

    def _section_table(self, title, thead_html, tbody_html, note="", title_color=None):
        color    = title_color if title_color else self._BLUE
        note_row = (f'<tr><td colspan="10" style="{self._FONT} padding:6px 12px 2px; font-size:11px; '
                    f'color:{self._GREY}; font-style:italic;">{note}</td></tr>') if note else ""
        return f"""
<table width="100%" cellpadding="0" cellspacing="0" border="0" style="margin-bottom:4px;">
  <tr><td style="{self._FONT} padding:20px 0 8px; font-size:11px; font-weight:700; text-transform:uppercase; letter-spacing:0.8px; color:{color};">{title}</td></tr>
</table>
<table width="100%" cellpadding="0" cellspacing="0" border="0" style="border-collapse:collapse; border:1px solid {self._BORDER};">
  <thead><tr>{thead_html}</tr></thead>
  <tbody>{tbody_html}{note_row}</tbody>
</table>"""

    def _group_label(self, label, color):
        return f"""
<table width="100%" cellpadding="0" cellspacing="0" border="0" style="margin:18px 0 6px;">
  <tr><td style="background:{color}; padding:6px 14px;">
    <span style="{self._FONT} font-size:12px; font-weight:700; color:#ffffff;">{label}</span>
  </td></tr>
</table>"""

    def _build_hotel_block(self, context, h_thead, use_cid=False):
        hotels_mgmt   = context.get("hotels_mgmt",   [])
        hotels_emp    = context.get("hotels_emp",     [])
        hotels_legacy = context.get("hotels",         [])
        m_count       = int(context.get("m_count",  0))
        e_count       = int(context.get("e_count",  0))
        dual          = m_count > 0 and e_count > 0 and (hotels_mgmt or hotels_emp)
        if dual:
            mgmt_names   = {h["name"] for h in hotels_mgmt}
            emp_names    = {h["name"] for h in hotels_emp}
            shared_names = mgmt_names & emp_names
            shared    = [h for h in hotels_mgmt if h["name"] in shared_names]
            mgmt_only = [h for h in hotels_mgmt if h["name"] not in shared_names]
            emp_only  = [h for h in hotels_emp  if h["name"] not in shared_names]
            block = ""
            if shared:
                block += self._group_label("&#127968; Shared Hotel &mdash; Both Groups Can Stay Here", self._GREEN)
                block += self._section_table("Management &amp; Employee Accommodation", h_thead,
                                             self._hotel_rows(shared, context, pax=m_count+e_count, bmin=1500, bmax=19000, use_cid=use_cid),
                                             note="Both management and employees can book rooms at this property.", title_color=self._GREEN)
            if mgmt_only:
                block += self._group_label("&#128188; Management Hotels", self._ORANGE)
                block += self._section_table("Management Accommodation", h_thead,
                                             self._hotel_rows(mgmt_only, context, pax=m_count, bmin=3000, bmax=19000, use_cid=use_cid), title_color=self._ORANGE)
            elif not shared:
                block += self._group_label("&#128188; Management Hotels", self._ORANGE)
                block += f'<p style="{self._FONT} font-size:13px; color:{self._GREY}; font-style:italic; margin:6px 0 16px;">No management hotels found.</p>'
            if emp_only:
                block += self._group_label("&#128100; Employee Hotels", self._EMP)
                block += self._section_table("Employee Accommodation", h_thead,
                                             self._hotel_rows(emp_only, context, pax=e_count, bmin=1500, bmax=3000, use_cid=use_cid), title_color=self._EMP)
            elif not shared:
                block += self._group_label("&#128100; Employee Hotels", self._EMP)
                block += f'<p style="{self._FONT} font-size:13px; color:{self._GREY}; font-style:italic; margin:6px 0 16px;">No employee hotels found.</p>'
            return block or f'<p style="{self._FONT} font-size:13px; color:{self._GREY}; font-style:italic; margin:20px 0;">No hotel required.</p>'
        if hotels_legacy:
            return self._section_table("&#127968; Accommodation Options", h_thead,
                                       self._hotel_rows(hotels_legacy, context, use_cid=use_cid))
        return f'<p style="{self._FONT} font-size:13px; color:{self._GREY}; font-style:italic; margin:20px 0;">No hotel required &mdash; same-day return.</p>'

    def _build_html_body(self, context, prose, use_cid=False):
        origin = context.get("origin", ""); destination = context.get("destination", "")
        date_str = context.get("date_str", "")
        main_flights   = context.get("flights", [])
        cheapest       = context.get("cheapest_flight")
        nearest        = context.get("nearest_flights", [])
        is_best_search = context.get("is_best_search", False)
        return_flights = context.get("return_flights", [])
        is_round_trip  = context.get("is_round_trip",  False)

        f_thead = (self._th("Airline", "left", "240") + self._th("Departure", "center", "90") +
                   self._th("", "center", "24") + self._th("Arrival", "center", "90") +
                   self._th("Total Fare", "center", "110"))
        flight_tables = ""
        if is_best_search:
            flight_tables += self._section_table("&#9992; Recommended &mdash; Nearest to Arrival", f_thead,
                                                 self._flight_rows(main_flights, context, use_cid=use_cid),
                                                 "These flights land closest to your meeting start time.")
            if cheapest:
                flight_tables += self._section_table("&#128176; Low-Cost Alternative", f_thead,
                                                     self._flight_rows([cheapest], context, use_cid=use_cid),
                                                     "Lower fare — may require earlier travel or an additional hotel night.")
        else:
            flight_tables += self._section_table("&#9992; Best-Value Flights (Outbound)", f_thead,
                                                 self._flight_rows(main_flights, context, use_cid=use_cid))
            seen = {(f.get("airline"), f.get("flight_no")) for f in main_flights}
            unique_nearest = [f for f in nearest if (f.get("airline"), f.get("flight_no")) not in seen]
            if unique_nearest:
                flight_tables += self._section_table("&#8987; Nearest to Arrival Options", f_thead,
                                                     self._flight_rows(unique_nearest, context, use_cid=use_cid),
                                                     "These options arrive closest to your meeting time.")

        # Return flights section — use swapped origin/destination context
        if is_round_trip and return_flights:
            rt_ctx = {**context,
                      "origin_iata": context.get("destination_iata", ""),
                      "destination_iata": context.get("origin_iata", "")}
            flight_tables += self._section_table("&#128260; Return Flights", f_thead,
                                                 self._flight_rows(return_flights, rt_ctx, use_cid=use_cid),
                                                 "Departing after meeting ends.")

        h_thead     = self._th("Hotel", "left", "340") + self._th("Rate / Night", "center", "140")
        hotel_block = self._build_hotel_block(context, h_thead, use_cid=use_cid)
        intro_html  = prose["intro"].replace("\n", "<br>")
        outro_html  = prose["outro"].replace("\n", "<br>")
        signoff_html = prose["signoff"].replace("\n", "<br>")

        return f"""<!DOCTYPE html>
<html xmlns:v="urn:schemas-microsoft-com:vml" xmlns:o="urn:schemas-microsoft-com:office:office">
<head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1"></head>
<body style="margin:0; padding:0; background:#f1f5f9; {self._FONT}">
<table width="100%" cellpadding="0" cellspacing="0" border="0" style="background:#f1f5f9; padding:24px 0;">
<tr><td align="center">
  <table width="680" cellpadding="0" cellspacing="0" border="0"
         style="background:#ffffff; border:1px solid {self._BORDER}; max-width:680px; width:100%;">
    <tr>
      <td style="background:{self._DKBLUE}; padding:28px 36px;">
        <p style="{self._FONT} font-size:10px; font-weight:600; letter-spacing:1.4px; text-transform:uppercase; color:#93c5fd; margin:0 0 6px;">Travel Approval Request &nbsp;&middot;&nbsp; SI2Tech</p>
        <p style="{self._FONT} font-size:22px; font-weight:700; color:#ffffff; margin:0 0 4px;">{origin} &rarr; {destination}{"&nbsp;&nbsp;&#128260;&nbsp;&nbsp;" + destination + " &rarr; " + origin if is_round_trip else ""}</p>
        <p style="{self._FONT} font-size:13px; color:#bfdbfe; margin:0;">Travel Date: {date_str}</p>
      </td>
    </tr>
    <tr>
      <td style="padding:32px 36px;">
        <p style="{self._FONT} font-size:14px; line-height:1.8; color:{self._MUTED}; margin:0 0 24px;">{prose['greeting']}<br><br>{intro_html}</p>
        <table width="100%" cellpadding="0" cellspacing="0" border="0" style="margin:0 0 24px;"><tr><td style="border-top:1px solid {self._BORDER}; font-size:0; line-height:0;">&nbsp;</td></tr></table>
        <table width="100%" cellpadding="0" cellspacing="0" border="0" style="background:{self._LGREY}; border:1px solid {self._BORDER}; margin-bottom:24px;">
          <tr><td style="{self._FONT} padding:14px 20px; font-size:13px; color:{self._MUTED};"><strong>Route:</strong> {origin} &rarr; {destination} &nbsp;&nbsp;&nbsp; <strong>Date:</strong> {date_str}</td></tr>
        </table>
        {flight_tables}
        <table width="100%" cellpadding="0" cellspacing="0" border="0" style="margin:16px 0 0;"><tr><td>&nbsp;</td></tr></table>
        {hotel_block}
        <table width="100%" cellpadding="0" cellspacing="0" border="0" style="margin:28px 0 20px;"><tr><td style="border-top:1px solid {self._BORDER}; font-size:0; line-height:0;">&nbsp;</td></tr></table>
        <p style="{self._FONT} font-size:14px; line-height:1.8; color:{self._MUTED}; margin:0 0 16px;">{outro_html}</p>
        <p style="{self._FONT} font-size:14px; font-weight:600; color:{self._MUTED}; margin:0;">{signoff_html}</p>
      </td>
    </tr>
    <tr>
      <td style="background:{self._LGREY}; border-top:1px solid {self._BORDER}; padding:14px 36px;">
        <p style="{self._FONT} font-size:11px; color:#9ca3af; text-align:center; margin:0;">Generated by the SI2Tech Travel Management System.</p>
      </td>
    </tr>
  </table>
</td></tr>
</table>
</body>
</html>"""

    def _build_plain_body(self, context, prose):
        SEP = "=" * 60; LINE = "-" * 40
        def f_block(flights, title, note=""):
            out = [title, LINE]
            for f in flights:
                raw_price = f.get('custom_price', f.get('price', 0))
                price     = self._fmt_price_plain(raw_price)
                vendor    = f.get("booking_vendor", "MakeMyTrip")
                out += [f"  {f.get('airline','').upper()}  |  {f.get('flight_no','')}",
                        f"   Departure : {f.get('dep_time','')}", f"   Arrival   : {f.get('arr_time','')}",
                        f"   Duration  : {f.get('duration','')}", f"   Fare      : {price} (via {vendor})",
                        f"   Book Now  : {self._flight_url(f, context)}", ""]
            if note: out += [f"  >> {note}", ""]
            return "\n".join(out)

        def h_block(hotels, title, pax=None, bmin=None, bmax=None):
            out = [title, LINE]
            for h in hotels:
                raw_price = h.get('custom_price', h.get('price_raw', 0))
                price_fmt = self._fmt_price_plain(raw_price) if h.get('custom_price') else h.get('price_fmt', 'N/A')
                vendor    = h.get("booking_vendor", "MakeMyTrip")
                out += [f"  {h.get('name','').upper()}", f"   Rating : {h.get('rating','N/A')} Stars",
                        f"   Rate   : {price_fmt} per night (via {vendor})",
                        f"   Book   : {self._hotel_url(h, context, pax=pax, bmin=bmin, bmax=bmax)}", ""]
            return "\n".join(out)

        main_flights  = context.get("flights", []); cheapest = context.get("cheapest_flight")
        nearest       = context.get("nearest_flights", []); is_best_search = context.get("is_best_search", False)
        return_flights = context.get("return_flights", []); is_round_trip  = context.get("is_round_trip", False)

        if is_best_search:
            f_sec = f_block(main_flights, "OUTBOUND FLIGHTS -- NEAREST TO ARRIVAL (RECOMMENDED)")
            if cheapest: f_sec += "\n" + f_block([cheapest], "LOW-COST ALTERNATIVE")
        else:
            f_sec = f_block(main_flights, "BEST-VALUE OUTBOUND FLIGHTS")
            seen  = {(f.get("airline"), f.get("flight_no")) for f in main_flights}
            uniq  = [f for f in nearest if (f.get("airline"), f.get("flight_no")) not in seen]
            if uniq: f_sec += "\n" + f_block(uniq, "NEAREST TO ARRIVAL OPTIONS")

        if is_round_trip and return_flights:
            f_sec += "\n" + f_block(return_flights, "RETURN FLIGHTS (after meeting ends)")

        hotels_mgmt = context.get("hotels_mgmt", []); hotels_emp = context.get("hotels_emp", [])
        hotels_legacy = context.get("hotels", [])
        m_count = int(context.get("m_count", 0)); e_count = int(context.get("e_count", 0))
        dual = m_count > 0 and e_count > 0 and (hotels_mgmt or hotels_emp)
        if dual:
            mn = {h["name"] for h in hotels_mgmt}; en = {h["name"] for h in hotels_emp}; sn = mn & en
            shared = [h for h in hotels_mgmt if h["name"] in sn]
            mo     = [h for h in hotels_mgmt if h["name"] not in sn]
            eo     = [h for h in hotels_emp  if h["name"] not in sn]
            h_sec = ""
            if shared: h_sec += h_block(shared, "SHARED ACCOMMODATION", pax=m_count+e_count, bmin=1500, bmax=19000)
            if mo:     h_sec += h_block(mo, "MANAGEMENT HOTELS", pax=m_count, bmin=3000, bmax=19000)
            elif not shared: h_sec += f"MANAGEMENT HOTELS\n{LINE}\n  No management hotels found.\n\n"
            if eo:     h_sec += h_block(eo, "EMPLOYEE HOTELS", pax=e_count, bmin=1500, bmax=3000)
            elif not shared: h_sec += f"EMPLOYEE HOTELS\n{LINE}\n  No employee hotels found.\n\n"
            h_sec = h_sec or "No hotel required (same-day return).\n"
        elif hotels_legacy:
            h_sec = h_block(hotels_legacy, "ACCOMMODATION OPTIONS")
        else:
            h_sec = "No hotel required (same-day return).\n"

        return (f"{prose['greeting']}\n\n{prose['intro']}\n\n{SEP}\n"
                f"  ROUTE   : {context.get('origin','')} to {context.get('destination','')}\n"
                f"  DATE    : {context.get('date_str','')}\n{SEP}\n\n"
                f"{f_sec}\n{h_sec}\n{SEP}\n\n{prose['outro']}\n\n{prose['signoff']}\n\n"
                f"{SEP}\nGenerated by SI2Tech Travel Management System.\n{SEP}")

    def _build_eml(self, subject, html_body, plain_body, recipient=RECIPIENT):
        msg = MIMEMultipart("related")
        msg["Subject"] = subject; msg["To"] = recipient
        msg["From"] = SENDER_NAME; msg["X-Unsent"] = "1"
        alt_part = MIMEMultipart("alternative"); msg.attach(alt_part)
        alt_part.attach(MIMEText(plain_body, "plain", "utf-8"))
        alt_part.attach(MIMEText(html_body,  "html",  "utf-8"))
        for cid, url, img_bytes in self.embedded_images:
            if img_bytes:
                try:
                    img_part = MIMEImage(img_bytes)
                    img_part.add_header("Content-ID", f"<{cid}>")
                    img_part.add_header("Content-Disposition", "inline")
                    msg.attach(img_part)
                except Exception as e:
                    print(f"Failed to attach image {cid}: {e}")
        return msg.as_bytes()

    def _build_preview(self, context, prose, eml_download_html=""):
        raw_html = self._build_html_body(context, prose, use_cid=False)
        return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap" rel="stylesheet">
<style>
  *{{box-sizing:border-box;margin:0;padding:0}}
  body{{font-family:'Inter',sans-serif;background:#f1f5f9;padding:20px 12px 48px}}
  .wrap{{max-width:700px;margin:0 auto}}
  .top-bar{{background:#fff;border:1px solid #e5e7eb;padding:14px 20px;display:flex;justify-content:space-between;align-items:center;margin-bottom:16px;border-radius:8px}}
  .top-bar span{{font-size:12px;color:#6b7280}}
  .dl-btn{{display:inline-flex;align-items:center;gap:6px;background:#1d4ed8;color:#fff;text-decoration:none;padding:9px 20px;border-radius:6px;font-size:13px;font-weight:600;font-family:'Inter',sans-serif}}
  .preview-container{{border-radius:8px;overflow:hidden;box-shadow:0 4px 6px -1px rgba(0,0,0,0.1)}}
</style>
</head>
<body>
<div class="wrap">
  <div class="top-bar">
    <span>&#128196; Preview — click button to open in Outlook with full formatting</span>
    {eml_download_html}
  </div>
  <div class="preview-container">{raw_html}</div>
</div>
</body>
</html>"""

    def prepare_email(self, context: dict) -> dict:
        prose = self._generate_prose(context)
        self.embedded_images = []
        html_body_for_eml = self._build_html_body(context, prose, use_cid=True)
        plain_body        = self._build_plain_body(context, prose)
        subject = f"Travel Approval: {context.get('origin','')} to {context.get('destination','')} on {context.get('date_str','')}"
        eml_bytes = self._build_eml(subject, html_body_for_eml, plain_body)
        mailto_link = (f"mailto:{urllib.parse.quote(RECIPIENT)}?subject={urllib.parse.quote(subject)}"
                       f"&body={urllib.parse.quote(plain_body)}")
        preview_html = self._build_preview(context, prose, "")
        return {
            "html_preview": preview_html, "eml_bytes": eml_bytes,
            "subject": subject, "mailto_link": mailto_link,
            "html_body": html_body_for_eml, "embedded_images": self.embedded_images,
        }


# ==============================================================================
#  EXCEL EXPORT — Fixed: separate outbound vs return flight columns
# ==============================================================================

def _collect_export_rows(
    flight_res, hotel_res,
    project_number, travel_reason,
    meeting_date, meeting_time, meeting_location,
    return_flight_res=None,
) -> tuple:
    """
    Extracts structured row data from scraped results.
    Returns (flight_rows, hotel_rows) — each is a list of dicts.
    Employee names are split into a list so _append_to_master_xlsx can expand them.
    """
    import re as _re_mod

    def _safe_val(v):
        if v is None: return ""
        if isinstance(v, (datetime, date)): return str(v)
        return v

    def _parse_date_row(value):
        """Parse a date from any common string format. Never raises."""
        if value is None: return None
        if isinstance(value, datetime): return value.date()
        if isinstance(value, date):     return value
        if not isinstance(value, str):  return None
        value = value.strip()
        if not value: return None
        # Extended format list covers scraper output variations
        _FMTS = (
            "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y",
            "%d %b %Y", "%d %B %Y", "%b %d, %Y", "%B %d, %Y",
            "%d-%b-%Y", "%d-%B-%Y", "%Y/%m/%d",
            "%d %b %y", "%d/%m/%y", "%Y-%m-%dT%H:%M:%S",
            "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M",
        )
        for fmt in _FMTS:
            try: return datetime.strptime(value, fmt).date()
            except ValueError: pass
        try: return datetime.fromisoformat(value).date()
        except: pass
        # LLM fallback for exotic formats (e.g. "11th March 2026", "Mar-11-26", "11.03.2026")
        try:
            import re as _re_d
            # Regex pattern: day-month_name-year
            m = _re_d.search(r"(\d{1,2})[\s/.\-]([A-Za-z]{3,9})[\s/.\-](\d{2,4})", value)
            if m:
                d, mon, y = m.group(1), m.group(2), m.group(3)
                y = f"20{y}" if len(y) == 2 else y
                try:
                    return datetime.strptime(f"{d} {mon[:3].capitalize()} {y}", "%d %b %Y").date()
                except Exception:
                    pass
            # Regex: year-month-day numeric
            m2 = _re_d.search(r"(\d{4})[/.\-](\d{1,2})[/.\-](\d{1,2})", value)
            if m2:
                return date(int(m2.group(1)), int(m2.group(2)), int(m2.group(3)))
            # Regex: day.month.year numeric (Indian style dd.mm.yyyy)
            m3 = _re_d.search(r"(\d{1,2})\.(\d{1,2})\.(\d{4})", value)
            if m3:
                return date(int(m3.group(3)), int(m3.group(2)), int(m3.group(1)))
        except Exception:
            pass
        # True LLM fallback: ask Groq to extract the date
        try:
            _groq_key = os.getenv("GROQ_API_KEY")
            if _groq_key:
                _gc = Groq(api_key=_groq_key)
                _resp = _gc.chat.completions.create(
                    model="meta-llama/llama-4-scout-17b-16e-instruct",
                    messages=[{
                        "role": "user",
                        "content": (
                            f"Extract the date from this text and return ONLY in YYYY-MM-DD format. "
                            f"Nothing else — no explanation. Text: \"{value}\""
                        )
                    }],
                    max_tokens=20, temperature=0,
                )
                raw_llm = _resp.choices[0].message.content.strip()
                # Validate it's a date string
                import re as _re_v
                if _re_v.match(r"\d{4}-\d{2}-\d{2}", raw_llm):
                    return datetime.strptime(raw_llm[:10], "%Y-%m-%d").date()
        except Exception:
            pass
        return None

    def _parse_amount_row(value):
        if value is None: return None
        if isinstance(value, (int, float)): return float(value)
        if isinstance(value, str):
            cleaned = value.strip().replace(",","").replace("\u20b9","").replace("Rs.","").replace("INR","").strip()
            try: return float(cleaned)
            except: return None
        return None

    def _split_names(raw):
        if not raw or not raw.strip(): return []
        return [n.strip() for n in _re_mod.split(r'[,&]|\band\b', raw) if n.strip()]

    def _primary(val):
        """Return only the first item from a comma-joined multi-leg string."""
        if not val: return ""
        return str(val).split(",")[0].strip()

    flight_rows = []
    hotel_rows  = []
    raw_traveler = ""

    # ── Outbound leg ──────────────────────────────────────────
    if flight_res and "error" not in flight_res:
        out_travel  = (
            _parse_date_row(flight_res.get("departure_datetime"))
            or _parse_date_row(flight_res.get("travel_date"))
            or meeting_date
        )
        # Use only primary airline/flight_no — never the "AIR INDIA, IndiGo" combined form
        out_airline  = _primary(flight_res.get("airline_primary") or flight_res.get("airline", ""))
        out_fno      = _primary(flight_res.get("flight_no_base") or flight_res.get("flight_number", ""))
        out_origin   = flight_res.get("origin", "")
        out_dest     = flight_res.get("destination", "")
        out_vendor   = flight_res.get("ota_source", "Unknown")
        out_amount   = _parse_amount_row(flight_res.get("total_amount"))
        out_bk_date  = _safe_val(_parse_date_row(flight_res.get("date_of_booking")))
        raw_traveler = flight_res.get("traveler_name", "") or ""
    else:
        out_airline = out_fno = out_origin = out_dest = out_vendor = ""
        out_amount = out_bk_date = out_travel = None

    # ── Return leg ────────────────────────────────────────────
    ret_airline = ret_fno = ret_origin = ret_dest = ret_vendor = ""
    ret_amount = ret_bk_date = ret_travel = None
    if return_flight_res and "error" not in return_flight_res:
        ret_travel  = (
            _parse_date_row(return_flight_res.get("departure_datetime"))
            or _parse_date_row(return_flight_res.get("travel_date"))
        )
        ret_airline  = _primary(
            return_flight_res.get("return_airline")
            or return_flight_res.get("airline_primary")
            or return_flight_res.get("airline", "")
        )
        ret_fno      = _primary(
            return_flight_res.get("return_flight_number")
            or return_flight_res.get("flight_no_base")
            or return_flight_res.get("flight_number", "")
        )
        ret_origin   = return_flight_res.get("origin") or out_dest
        ret_dest     = return_flight_res.get("destination") or out_origin
        ret_vendor   = return_flight_res.get("ota_source", "Unknown")
        ret_amount   = _parse_amount_row(return_flight_res.get("total_amount"))
        ret_bk_date  = _safe_val(_parse_date_row(return_flight_res.get("date_of_booking")))
        if not raw_traveler:
            raw_traveler = return_flight_res.get("traveler_name", "") or ""

    emp_names = _split_names(raw_traveler) or [""]

    if flight_res or return_flight_res:
        flight_rows.append({
            "employees":        emp_names,
            "timestamp":        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "project_no":       project_number or "",
            "reason":           travel_reason or "",
            "out_airline":      out_airline,
            "out_fno":          out_fno,
            "out_origin":       out_origin,
            "out_dest":         out_dest,
            "out_travel":       _safe_val(out_travel),
            "out_vendor":       out_vendor,
            "out_amount":       out_amount,
            "ret_airline":      ret_airline,
            "ret_fno":          ret_fno,
            "ret_origin":       ret_origin,
            "ret_dest":         ret_dest,
            "ret_travel":       _safe_val(ret_travel),
            "ret_vendor":       ret_vendor,
            "ret_amount":       ret_amount,
            "bk_date":          out_bk_date or ret_bk_date or "",
            "meeting_date":     _safe_val(meeting_date),
            "meeting_time":     _safe_val(meeting_time),
            "meeting_location": meeting_location or "",
        })

    # ── Hotel ─────────────────────────────────────────────────
    if hotel_res and "error" not in hotel_res:
        hotel_traveler_raw  = hotel_res.get("traveler_name", "") or raw_traveler
        hotel_emp_names     = _split_names(hotel_traveler_raw) or [""]
        hotel_rows.append({
            "employees":        hotel_emp_names,
            "timestamp":        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "project_no":       project_number or "",
            "reason":           travel_reason or "",
            "hotel_name":       hotel_res.get("hotel_name", ""),
            "checkin":          _safe_val(_parse_date_row(hotel_res.get("checkin_date"))),
            "checkout":         _safe_val(_parse_date_row(hotel_res.get("checkout_date"))),
            "vendor":           hotel_res.get("ota_source", "Unknown"),
            "amount":           _parse_amount_row(hotel_res.get("total_amount")),
            "bk_date":          _safe_val(_parse_date_row(hotel_res.get("date_of_booking"))),
            "meeting_date":     _safe_val(meeting_date),
            "meeting_time":     _safe_val(meeting_time),
            "meeting_location": meeting_location or "",
        })

    return flight_rows, hotel_rows


# ── Master Excel headers (fixed schema) ──────────────────────────────────────
_FLIGHT_SHEET_HEADERS = [
    "Booking Timestamp", "Employee Name", "Project No", "Reason",
    "Travel Airline", "Travel Flight No", "Travel Origin", "Travel Destination",
    "Travel Date", "Booking Vendor", "Amount (INR)",
    "Return Airline", "Return Flight No", "Return Origin", "Return Destination",
    "Return Travel Date", "Return Booking Vendor", "Return Amount (INR)",
    "Booking Date", "Meeting Date", "Meeting Time", "Meeting Location",
]

_HOTEL_SHEET_HEADERS = [
    "Booking Timestamp", "Employee Name", "Project No", "Reason",
    "Hotel Name", "Check-in Date", "Checkout Date",
    "Booking Vendor", "Amount (INR)", "Booking Date",
    "Meeting Date", "Meeting Time", "Meeting Location",
]


def _append_to_master_xlsx(flight_rows: list, hotel_rows: list, filepath: str):
    """
    Appends flight and hotel rows to the master Excel file on disk.
    Creates the file with headers if it doesn't exist yet.
    Never requires a download — the file is always on disk and auto-updated.
    """
    import io as _io
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl required — run: pip install openpyxl")

    HEADER_FILL = PatternFill("solid", fgColor="1D4ED8")
    HEADER_FONT = Font(color="FFFFFF", bold=True, name="Calibri", size=10)
    DATA_FONT   = Font(name="Calibri", size=10)
    ALT_FILL    = PatternFill("solid", fgColor="EFF6FF")
    BORDER_SIDE = Side(style="thin", color="D1D5DB")
    CELL_BORDER = Border(left=BORDER_SIDE, right=BORDER_SIDE,
                         top=BORDER_SIDE,  bottom=BORDER_SIDE)
    CENTER      = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT        = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    def _ensure_sheet(wb, title, headers):
        if title in wb.sheetnames:
            ws = wb[title]
        else:
            ws = wb.create_sheet(title)
            ws.freeze_panes = "A2"
            for col_idx, h in enumerate(headers, 1):
                cell           = ws.cell(row=1, column=col_idx, value=h)
                cell.fill      = HEADER_FILL
                cell.font      = HEADER_FONT
                cell.alignment = CENTER
                cell.border    = CELL_BORDER
        return ws

    def _style_data(ws, row_idx, n_cols):
        fill = ALT_FILL if row_idx % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        for col_idx in range(1, n_cols + 1):
            cell           = ws.cell(row=row_idx, column=col_idx)
            cell.fill      = fill
            cell.font      = DATA_FONT
            cell.alignment = LEFT
            cell.border    = CELL_BORDER

    def _autofit(ws, headers):
        for col_idx, h in enumerate(headers, 1):
            col_letter = get_column_letter(col_idx)
            max_len    = max(len(str(h)), 12)
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
                for cell in row:
                    val = str(cell.value) if cell.value is not None else ""
                    max_len = max(max_len, len(val))
            ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

    # Load or create workbook
    import os as _os
    if _os.path.exists(filepath):
        try:
            wb = openpyxl.load_workbook(filepath)
        except Exception:
            wb = openpyxl.Workbook()
    else:
        wb = openpyxl.Workbook()
        # Remove default "Sheet" created by openpyxl
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    # ── Flight Bookings sheet ─────────────────────────────────
    ws_f = _ensure_sheet(wb, "Flight Bookings", _FLIGHT_SHEET_HEADERS)
    n_f  = len(_FLIGHT_SHEET_HEADERS)

    for r in flight_rows:
        # Each employee in the list gets its own row
        emps = r.get("employees", [""])
        for emp in emps:
            row_data = [
                r.get("timestamp", ""),
                emp,
                r.get("project_no", ""),   r.get("reason", ""),
                r.get("out_airline", ""),   r.get("out_fno", ""),
                r.get("out_origin", ""),    r.get("out_dest", ""),
                r.get("out_travel", ""),    r.get("out_vendor", ""),
                r.get("out_amount"),
                r.get("ret_airline", ""),   r.get("ret_fno", ""),
                r.get("ret_origin", ""),    r.get("ret_dest", ""),
                r.get("ret_travel", ""),    r.get("ret_vendor", ""),
                r.get("ret_amount"),
                r.get("bk_date", ""),
                r.get("meeting_date", ""),  r.get("meeting_time", ""),
                r.get("meeting_location", ""),
            ]
            row_idx = ws_f.max_row + 1
            ws_f.append(row_data)
            _style_data(ws_f, row_idx, n_f)

    _autofit(ws_f, _FLIGHT_SHEET_HEADERS)

    # ── Hotel Bookings sheet ──────────────────────────────────
    ws_h = _ensure_sheet(wb, "Hotel Bookings", _HOTEL_SHEET_HEADERS)
    n_h  = len(_HOTEL_SHEET_HEADERS)

    for r in hotel_rows:
        emps = r.get("employees", [""])
        for emp in emps:
            row_data = [
                r.get("timestamp", ""),
                emp,
                r.get("project_no", ""),    r.get("reason", ""),
                r.get("hotel_name", ""),
                r.get("checkin", ""),        r.get("checkout", ""),
                r.get("vendor", ""),         r.get("amount"),
                r.get("bk_date", ""),
                r.get("meeting_date", ""),   r.get("meeting_time", ""),
                r.get("meeting_location", ""),
            ]
            row_idx = ws_h.max_row + 1
            ws_h.append(row_data)
            _style_data(ws_h, row_idx, n_h)

    _autofit(ws_h, _HOTEL_SHEET_HEADERS)

    wb.save(filepath)
    total_f = ws_f.max_row - 1  # subtract header
    total_h = ws_h.max_row - 1
    return total_f, total_h


@st.cache_data(show_spinner=False)
def get_safe_image(img_url):
    if not img_url:
        return None
    if isinstance(img_url, bytes):
        return img_url
    if isinstance(img_url, str) and img_url.startswith("data:image"):
        try:
            _, encoded = img_url.split(",", 1)
            return base64.b64decode(encoded)
        except Exception:
            return None
    if not isinstance(img_url, str):
        return None
    if img_url.startswith("http://"):
        img_url = img_url.replace("http://", "https://", 1)
    elif img_url.startswith("//"):
        img_url = "https:" + img_url
    try:
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/122.0.0.0 Safari/537.36",
            "Accept": "image/webp,image/apng,image/*,*/*;q=0.8",
            "Accept-Language": "en-US,en;q=0.9", "Referer": "https://www.google.com/",
        }
        response = requests.get(img_url, headers=headers, timeout=6, allow_redirects=True)
        content_type = response.headers.get("Content-Type", "").lower()
        if response.status_code == 200 and "text/html" not in content_type:
            return response.content
        return None
    except Exception as e:
        print(f"Image fetch failed for {img_url}: {e}")
        return None


def _fetch_image_bytes_aggressive(url: str):
    if not url or not isinstance(url, str):
        return None
    if url.startswith("data:image"):
        try:
            _, encoded = url.split(",", 1); return base64.b64decode(encoded)
        except Exception:
            return None
    if url.startswith("//"): url = "https:" + url
    elif url.startswith("http://"): url = url.replace("http://", "https://", 1)
    user_agents = [
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/122.0.0.0 Safari/537.36",
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:124.0) Gecko/20100101 Firefox/124.0",
    ]
    for ua in user_agents:
        try:
            resp = requests.get(url, headers={"User-Agent": ua, "Accept": "image/webp,image/apng,image/*,*/*;q=0.8",
                                              "Referer": "https://www.google.com/"}, timeout=10, allow_redirects=True)
            ct = resp.headers.get("Content-Type", "").lower()
            if resp.status_code == 200 and "text/html" not in ct and len(resp.content) > 500:
                return resp.content
        except Exception:
            pass
    return None


@st.cache_data(show_spinner=False)
def preload_images_in_parallel(hotel_list):
    urls = [h.get("thumbnail") for h in hotel_list if h.get("thumbnail")]
    with concurrent.futures.ThreadPoolExecutor(max_workers=10) as executor:
        executor.map(get_safe_image, set(urls))


# ── Graph API ─────────────────────────────────────────────────────────────────

def _bytes_to_attachment(cid: str, img_bytes: bytes):
    try:
        img = Image.open(io.BytesIO(img_bytes))
        if img.mode in ("RGBA", "LA"):
            bg = Image.new("RGB", img.size, (255, 255, 255)); bg.paste(img, mask=img.split()[-1]); img = bg
        elif img.mode == "P":
            img = img.convert("RGBA"); bg = Image.new("RGB", img.size, (255, 255, 255)); bg.paste(img, mask=img.split()[-1]); img = bg
        elif img.mode != "RGB":
            img = img.convert("RGB")
        max_dim = 600
        if img.width > max_dim or img.height > max_dim:
            img.thumbnail((max_dim, max_dim), Image.LANCZOS)
        buf = io.BytesIO(); img.save(buf, format="JPEG", quality=72, optimize=True)
        final_bytes = buf.getvalue()
        return {"@odata.type": "#microsoft.graph.fileAttachment", "name": f"{cid}.jpg",
                "contentType": "image/jpeg", "contentBytes": base64.b64encode(final_bytes).decode("utf-8"),
                "isInline": True, "contentId": cid}
    except Exception as e:
        print(f"   _bytes_to_attachment failed for {cid}: {e}")
        return None


def send_via_graph_api(subject, html_content, to_email, embedded_images=None):
    TENANT_ID     = os.getenv("AZURE_TENANT_ID"); CLIENT_ID = os.getenv("AZURE_CLIENT_ID")
    CLIENT_SECRET = os.getenv("AZURE_CLIENT_SECRET"); SENDER_EMAIL = os.getenv("SENDER_EMAIL")
    authority    = f"https://login.microsoftonline.com/{TENANT_ID}"
    app          = msal.ConfidentialClientApplication(CLIENT_ID, authority=authority, client_credential=CLIENT_SECRET)
    token_result = app.acquire_token_for_client(scopes=["https://graph.microsoft.com/.default"])
    if "access_token" not in token_result:
        return False, "Failed to authenticate with Microsoft Entra."
    access_token = token_result["access_token"]
    graph_attachments = []
    if embedded_images:
        for i, item in enumerate(embedded_images):
            cid = item[0]; img_bytes = None
            for idx in range(1, len(item)):
                if isinstance(item[idx], (bytes, bytearray)) and len(item[idx]) > 100:
                    img_bytes = item[idx]; break
            if img_bytes is None:
                url_to_fetch = None
                for idx in range(1, len(item)):
                    if isinstance(item[idx], str) and item[idx].startswith("http"):
                        url_to_fetch = item[idx]; break
                if url_to_fetch:
                    img_bytes = _fetch_image_bytes_aggressive(url_to_fetch)
            if img_bytes is None:
                continue
            attachment = _bytes_to_attachment(cid, img_bytes)
            if attachment:
                graph_attachments.append(attachment)
    message_payload = {
        "subject": subject, "body": {"contentType": "HTML", "content": html_content},
        "toRecipients": [{"emailAddress": {"address": to_email}}],
    }
    if graph_attachments:
        message_payload["attachments"] = graph_attachments
    email_msg  = {"message": message_payload, "saveToSentItems": "true"}
    payload_mb = len(json.dumps(email_msg).encode("utf-8")) / (1024 * 1024)
    if payload_mb > 4.0:
        return False, f"Email too large ({payload_mb:.2f} MB)."
    endpoint    = f"https://graph.microsoft.com/v1.0/users/{SENDER_EMAIL}/sendMail"
    api_headers = {"Authorization": f"Bearer {access_token}", "Content-Type": "application/json"}
    response    = requests.post(endpoint, headers=api_headers, json=email_msg)
    if response.status_code == 202:
        return True, f"Email sent! ({len(graph_attachments)} image(s) embedded)"
    else:
        return False, f"Graph API Error {response.status_code}: {response.text[:300]}"


# ── Session State ─────────────────────────────────────────────────────────────

for _k, _v in {
    "show_welcome":           True,
    "display_flights":        [],
    "hotel_results_mgmt":     [],
    "hotel_results_emp":      [],
    "hotel_results_shared":   [],
    "hotel_results":          [],
    "cheapest_flight":        None,
    "nearest_flights":        [],
    "is_best_search":         False,
    "email_package":          None,
    "last_meeting_location":  "",
    "view":                   "results",
    "selected_flight_keys":   set(),
    "selected_hotel_keys":    set(),
    "selected_flights":       [],
    "selected_hotels":        [],
    "last_m_count":           0,
    "last_e_count":           0,
    "project_number":         "",
    "travel_reason":          "",
    "show_tracking_overlay":  False,
    "return_flights":              [],
    "is_round_trip":               False,
    "meeting_end_date_val":        None,
    "meeting_end_time_val":        None,
    "selected_return_flight_keys": set(),
    "selected_return_flights":     [],
    # Persisted MMT URLs — survive navigation away and back to the results page
    "mmt_flight_url":              "https://www.makemytrip.com/flight/search",
    "mmt_hotel_url":               "https://www.makemytrip.com/hotels/",
    "mmt_flight_label":            "✈️ Book on MMT",
    "mmt_hotel_label":             "🏨 Hotels on MMT",
}.items():
    if _k not in st.session_state:
        st.session_state[_k] = _v


# ── Key helpers ───────────────────────────────────────────────────────────────

def _flight_key(f): return f"{f.get('airline','')}__{f.get('flight_no','')}__{f.get('dep_time','')}"
def _hotel_key(h):  return f"{h.get('name','')}__{h.get('price_fmt','')}"

def _rebuild_selected_lists():
    all_flights = st.session_state.get("display_flights", [])
    all_hotels  = (st.session_state.get("hotel_results_shared", []) +
                   st.session_state.get("hotel_results_mgmt",   []) +
                   st.session_state.get("hotel_results_emp",    []))
    all_returns = st.session_state.get("return_flights", [])

    sel_outbound = [f for f in all_flights if _flight_key(f) in st.session_state["selected_flight_keys"]]

    sel_return = []
    for f in all_returns:
        if _flight_key(f) in st.session_state["selected_return_flight_keys"]:
            tagged = f.copy()
            tagged["_is_return"] = True
            sel_return.append(tagged)

    st.session_state["selected_flights"]        = sel_outbound + sel_return
    st.session_state["selected_hotels"]         = [h for h in all_hotels  if _hotel_key(h)  in st.session_state["selected_hotel_keys"]]
    st.session_state["selected_return_flights"] = sel_return

MGMT_MIN, MGMT_MAX = 2700, 19000
EMP_MIN,  EMP_MAX  = 1500,  3500

# ── Page config — MUST be first Streamlit call ────────────────────────────────
st.set_page_config(
    page_title="Si2 Travel Planner",
    page_icon="✈️",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── Global CSS ────────────────────────────────────────────────────────────────
st.markdown("""
<style>
/* ── Main content: never clip, full width, no horizontal scroll ── */
.block-container {
    max-width: 100% !important;
    width: 100% !important;
    padding-top: 4rem !important;
    padding-bottom: 2rem !important;
    padding-left: 1rem !important;
    padding-right: 1rem !important;
    box-sizing: border-box !important;
    overflow-x: hidden !important;
}

/* Streamlit's outermost app wrapper — must not clip */
[data-testid="stAppViewContainer"] > section[data-testid="stMain"] {
    overflow-x: hidden !important;
    width: 100% !important;
}
[data-testid="stMain"] > div {
    overflow-x: hidden !important;
    width: 100% !important;
}

/* ── Columns: constrain to their grid cell, never clip content ── */
[data-testid="column"] {
    min-width: 0 !important;
    overflow: visible !important;
}

/* ── Sidebar: fixed width, scrollable internally ── */
[data-testid="stSidebar"] > div:first-child {
    width: 280px !important;
    min-width: 280px !important;
    max-width: 280px !important;
    padding: 1rem 0.75rem 2rem !important;
    overflow-y: auto !important;
    overflow-x: hidden !important;
}

/* ── Sidebar inputs ── */
[data-testid="stSidebar"] .stNumberInput input {
    font-size: 15px !important; height: 40px !important; text-align: center !important;
}
[data-testid="stSidebar"] .stNumberInput button {
    height: 38px !important; font-size: 18px !important;
}
[data-testid="stSidebar"] .stTextInput input,
[data-testid="stSidebar"] .stDateInput input,
[data-testid="stSidebar"] .stTimeInput input {
    height: 40px !important; font-size: 14px !important;
}

/* ── Sidebar buttons ── */
[data-testid="stSidebar"] .stButton > button {
    width: 100% !important; font-size: 14px !important;
    padding: 8px 12px !important; border-radius: 8px !important;
    white-space: nowrap !important; overflow: hidden !important;
    text-overflow: ellipsis !important;
}
[data-testid="stSidebar"] .stLinkButton a {
    display: block !important; width: 100% !important;
    text-align: center !important; white-space: nowrap !important;
    font-size: 14px !important; padding: 8px 12px !important;
    border-radius: 8px !important; box-sizing: border-box !important;
    overflow: hidden !important; text-overflow: ellipsis !important;
}

/* ── All buttons: no wrap ── */
.stButton > button, .stLinkButton a {
    white-space: nowrap !important;
}

/* ── Images inside cards: never exceed column width ── */
[data-testid="stImage"] img {
    max-width: 100% !important;
    height: auto !important;
}
</style>
""", unsafe_allow_html=True)

# ── Sidebar ───────────────────────────────────────────────────────────────────

with st.sidebar:
    st.header("✈️ Travel Details")
    st.markdown("---")

    if "origin_city" not in st.session_state: st.session_state["origin_city"] = ""
    if "dest_city"   not in st.session_state: st.session_state["dest_city"]   = ""

    def _swap_cities():
        st.session_state["origin_city"], st.session_state["dest_city"] = (
            st.session_state["dest_city"], st.session_state["origin_city"])

    col_orig, col_swap, col_dest = st.columns([10, 2, 10])
    with col_orig: origin_input = st.text_input("Origin",      key="origin_city", placeholder="e.g. Vadodara")
    with col_swap:
        st.button("🔄", on_click=_swap_cities, help="Swap Origin and Destination", use_container_width=True)
    with col_dest: dest_input = st.text_input("Destination", key="dest_city", placeholder="e.g. London")

    origin      = get_flight_iata_from_city(origin_input)
    destination = get_flight_iata_from_city(dest_input)

    if "last_search_params" not in st.session_state:
        st.session_state["last_search_params"] = {}

    st.markdown("### 👥 Passenger Details")
    exec_count  = st.number_input("Number of Executives",  min_value=0, max_value=9, value=int(st.session_state.get("exec_count",  0)), step=1)
    m_count     = st.number_input("Number of Management",  min_value=0, max_value=9, value=int(st.session_state.get("last_m_count", 0)), step=1)
    e_count     = st.number_input("Number of Employee",    min_value=0, max_value=9, value=int(st.session_state.get("last_e_count", 0)), step=1)
    rooms_count = st.number_input("🛏️ Number of Rooms",   min_value=1, max_value=9, value=int(st.session_state.get("rooms_count",  1)), step=1)

    st.markdown("### 📅 Schedule & Purpose")

    trip_type_choice = st.radio(
        "Trip Type", ["One-way", "Round Trip"],
        horizontal=True, key="trip_type_radio",
    )
    is_round_trip = (trip_type_choice == "Round Trip")

    meeting_date = st.date_input("Meeting Date",
                                 value=st.session_state.get("meeting_date_val", None))

    _time_slots = [""] + [
        datetime.strptime(f"{h:02d}:{m:02d}", "%H:%M").strftime("%I:%M %p")
        for h in range(0, 24) for m in (0, 30)
    ]
    _saved_time = st.session_state.get("meeting_time_val")
    _saved_time_str = _saved_time.strftime("%I:%M %p") if _saved_time else ""
    _time_idx = _time_slots.index(_saved_time_str) if _saved_time_str in _time_slots else 0
    _meeting_time_sel = st.selectbox("Meeting Time", _time_slots, index=_time_idx, key="meeting_time_select")
    meeting_time = datetime.strptime(_meeting_time_sel, "%I:%M %p").time() if _meeting_time_sel else None

    if is_round_trip:
        meeting_end_date = st.date_input(
            "Meeting End Date",
            value=st.session_state.get("meeting_end_date_val", None),
            key="meeting_end_date_input",
        )
        _saved_end_time = st.session_state.get("meeting_end_time_val")
        _saved_end_time_str = _saved_end_time.strftime("%I:%M %p") if _saved_end_time else ""
        _end_time_idx = _time_slots.index(_saved_end_time_str) if _saved_end_time_str in _time_slots else 0
        _meeting_end_time_sel = st.selectbox("Meeting End Time", _time_slots, index=_end_time_idx, key="meeting_end_time_select")
        meeting_end_time = datetime.strptime(_meeting_end_time_sel, "%I:%M %p").time() if _meeting_end_time_sel else None
        st.session_state["meeting_end_date_val"] = meeting_end_date
        st.session_state["meeting_end_time_val"] = meeting_end_time
        if meeting_end_date and meeting_end_time:
            meeting_end_dt = datetime.combine(meeting_end_date, meeting_end_time)
        elif meeting_end_date:
            meeting_end_dt = datetime.combine(meeting_end_date, datetime.strptime("18:00", "%H:%M").time())
        else:
            meeting_end_dt = datetime.combine(meeting_date, datetime.strptime("18:00", "%H:%M").time()) if meeting_date else None
    else:
        meeting_end_date = meeting_date
        meeting_end_time = meeting_time
        meeting_end_dt   = None

    if "meeting_location_input" not in st.session_state:
        st.session_state["meeting_location_input"] = ""
    meeting_location = st.text_input("Meeting Location", key="meeting_location_input", placeholder="e.g. Canary Wharf, London")
    project_number   = st.session_state.get("project_number", "")
    travel_reason    = st.session_state.get("travel_reason", "")

    st.session_state["meeting_date_val"]  = meeting_date
    st.session_state["meeting_time_val"]  = meeting_time

    _has_results = bool(st.session_state.get("display_flights") or
                        st.session_state.get("hotel_results_mgmt") or
                        st.session_state.get("hotel_results_emp")  or
                        st.session_state.get("hotel_results_shared"))

    current_params = {
        "origin": origin_input, "dest": dest_input,
        "exec_count": exec_count, "m_count": m_count, "e_count": e_count,
        "rooms_count": rooms_count, "meeting_date": meeting_date, "meeting_time": meeting_time,
        "meeting_end_date": meeting_end_date if is_round_trip else None,
        "meeting_end_time": meeting_end_time if is_round_trip else None,
        "is_round_trip": is_round_trip,
        "location": meeting_location, "project_number": project_number, "reason": travel_reason
    }
    # Always sync params — no need for an explicit Update button (search reruns from scratch)
    st.session_state["rooms_count"] = int(rooms_count)
    st.session_state["exec_count"]  = int(exec_count)
    t_count = int(m_count) + int(e_count) + int(exec_count)

    _rooms_invalid = t_count > 0 and int(rooms_count) > t_count
    if _rooms_invalid:
        st.error(f"❌ Rooms ({int(rooms_count)}) cannot exceed total passengers ({t_count}).")

    st.markdown("---")
    st.caption("Search for Best Flight (Nearest to Arrival)")
    if "top_pick" not in st.session_state:
        st.session_state["top_pick"] = False
    best_per_airline = st.toggle("Top Pick Per Airline", key="top_pick")

    search_triggered = st.button("Find Best Flights", type="primary", width='stretch')
    if not st.session_state.get("meeting_location_input", "").strip():
        st.warning("📍 Please enter a Meeting Location to enable search.")

    if search_triggered:
        st.session_state["last_search_params"] = current_params
        today = datetime.now().date()
        if not st.session_state.get("meeting_location_input", "").strip():
            st.error("❌ Meeting Location is required.")
            st.stop()
        elif not meeting_date:
            st.error("❌ Please select a Meeting Date.")
            st.stop()
        elif not meeting_time:
            st.error("❌ Please select a Meeting Time.")
            st.stop()
        elif meeting_date < today:
            st.error("❌ Meeting date cannot be in the past!")
            st.stop()
        elif is_round_trip and meeting_end_date and meeting_end_date < meeting_date:
            st.error("❌ Meeting end date cannot be before meeting start date!")
            st.stop()
        else:
            st.success("✅ Meeting date is valid!")

    st.divider()

    n_sf = len(st.session_state["selected_flight_keys"])
    n_sh = len(st.session_state["selected_hotel_keys"])
    if n_sf > 0 or n_sh > 0:
        st.caption(f"✅ **{n_sf} flight{'s' if n_sf != 1 else ''}** and **{n_sh} hotel{'s' if n_sh != 1 else ''}** selected")

    if st.button("✅ Verify Before Approval", type="primary", width='stretch'):
        st.session_state["show_tracking_overlay"] = False
        st.session_state["show_welcome"] = False
        st.session_state["view"] = "verify"
        st.rerun()

    if st.session_state.get("draft_email_triggered", False):
        st.session_state["draft_email_triggered"] = False
        _rebuild_selected_lists()
        sel_f    = st.session_state["selected_flights"]
        sel_h    = st.session_state["selected_hotels"]
        has_any  = bool(st.session_state.get("display_flights") or
                        st.session_state.get("hotel_results_mgmt") or
                        st.session_state.get("hotel_results_emp") or
                        st.session_state.get("hotel_results_shared"))
        if not has_any:
            st.sidebar.warning("Search for flights before generating the email!")
        else:
            has_sel_flights   = bool(sel_f)
            raw_email_flights = sel_f if has_sel_flights else st.session_state.get("display_flights", [])
            shared_for_email  = st.session_state.get("hotel_results_shared", [])
            has_hotels        = bool(sel_h)
            # Only include hotels that were explicitly selected — if none selected, send no hotels
            if has_hotels:
                raw_email_h_mgmt = [h for h in (shared_for_email + st.session_state.get("hotel_results_mgmt", [])) if _hotel_key(h) in st.session_state["selected_hotel_keys"]]
                raw_email_h_emp  = [h for h in (shared_for_email + st.session_state.get("hotel_results_emp",  [])) if _hotel_key(h) in st.session_state["selected_hotel_keys"]]
            else:
                raw_email_h_mgmt = []
                raw_email_h_emp  = []

            def _enrich_with_image(items):
                enriched = []
                for item in items:
                    copy  = item.copy(); thumb = copy.get("thumbnail")
                    data  = get_safe_image(thumb)
                    if not data and thumb:
                        data = _fetch_image_bytes_aggressive(thumb)
                    copy["safe_img_data"] = data; enriched.append(copy)
                return enriched

            email_flights  = _enrich_with_image(raw_email_flights)
            email_h_mgmt   = _enrich_with_image(raw_email_h_mgmt)
            email_h_emp    = _enrich_with_image(raw_email_h_emp)
            _sel_ret_keys  = st.session_state.get("selected_return_flight_keys", set())
            _all_ret       = st.session_state.get("return_flights", [])
            _raw_ret       = [f for f in _all_ret if _flight_key(f) in _sel_ret_keys] if _sel_ret_keys else _all_ret
            email_returns  = _enrich_with_image(_raw_ret)

            try:
                _ci = meeting_date; _co = meeting_date + timedelta(days=1)
                _last_m = int(st.session_state.get("last_m_count", 0))
                _last_e = int(st.session_state.get("last_e_count", 0))
                _last_t = _last_m + _last_e
                bmin, bmax = (3000, 19000) if _last_m > 0 else (1500, 3000)

                email_context = {
                    "origin": origin_input, "origin_iata": origin,
                    "destination": dest_input, "destination_iata": destination,
                    "date_str": meeting_date.strftime("%A, %d %B %Y"),
                    "travel_date": datetime.combine(meeting_date, meeting_time),
                    "check_in": _ci, "check_out": _co,
                    "budget_min": bmin, "budget_max": bmax,
                    "travel_count": _last_t, "m_count": _last_m, "e_count": _last_e,
                    "meeting_location": meeting_location,
                    "dest_city": get_airport_city(destination),
                    "flights": email_flights,
                    "is_best_search": False if has_sel_flights else st.session_state.get("is_best_search", False),
                    "cheapest_flight": None if has_sel_flights else st.session_state.get("cheapest_flight"),
                    "nearest_flights": [] if has_sel_flights else st.session_state.get("nearest_flights", []),
                    "hotels_mgmt": email_h_mgmt, "hotels_emp": email_h_emp,
                    "hotels": email_h_mgmt + email_h_emp,
                    "is_round_trip":  st.session_state.get("is_round_trip", False),
                    "return_flights": email_returns,
                }

                with st.spinner("🤖 Drafting approval email…"):
                    agent   = EmailAgent()
                    package = agent.prepare_email(email_context)
                    st.session_state["email_package"] = package
                    st.session_state["view"] = "email"

            except Exception as e:
                st.sidebar.error(f"Error: {e}")

    if st.button("📄 Upload and Track", type="secondary", width='stretch'):
        st.session_state["view"] = "tracking"
        st.session_state["show_tracking_overlay"] = True
        st.session_state["show_welcome"] = False
        st.rerun()

    if st.button("💾 Add to Database", type="primary", width='stretch'):
        st.session_state["show_welcome"] = False
        flight_res        = st.session_state.get("scraped_flight")
        hotel_res         = st.session_state.get("scraped_hotel")
        return_flight_res = st.session_state.get("scraped_return_flight")
        if not flight_res and not hotel_res and not return_flight_res:
            st.warning("⚠️ No scraped data. Upload and scrape receipts first.")
        else:
            def _primary_val(v):
                """Return only the first value from a comma-joined multi-leg string."""
                if not v: return ""
                return str(v).split(",")[0].strip()

            def _clean_for_db(rec, out_rec=None):
                """
                Strips fields we no longer store (departure_time, arrival_time, status)
                and ensures airline / flight_number contain only the primary leg value.
                For return flights, ensures origin/dest are correctly swapped from outbound.
                """
                if not rec: return rec
                c = dict(rec)
                # Remove time and status fields — not needed in DB
                for drop in ("departure_time", "arrival_time", "dep_time", "arr_time",
                             "arr_time_raw", "status", "dep_dt", "arr_dt"):
                    c.pop(drop, None)
                # Use only primary airline / flight_number (never "AIR INDIA, IndiGo" style)
                c["airline"]       = _primary_val(c.get("airline_primary") or c.get("airline", ""))
                c["flight_number"] = _primary_val(c.get("flight_no_base") or c.get("flight_number", ""))
                # Backfill origin/dest for return flights from outbound if missing
                if out_rec:
                    if not c.get("origin"):
                        c["origin"] = _primary_val(out_rec.get("destination", ""))
                    if not c.get("destination"):
                        c["destination"] = _primary_val(out_rec.get("origin", ""))
                return c

            _inserted_booking_ids = []
            try:
                from addtodatabase_v2 import add_flight_booking, add_hotel_booking, build_ui_inputs, fetch_flight_transactions, fetch_hotel_transactions
                traveler = ""
                if flight_res and "error" not in flight_res:
                    traveler = flight_res.get("traveler_name", "")
                elif return_flight_res and "error" not in return_flight_res:
                    traveler = return_flight_res.get("traveler_name", "")
                elif hotel_res and "error" not in hotel_res:
                    traveler = hotel_res.get("traveler_name", "")
                ui = build_ui_inputs(emp_name=traveler, project_no=project_number, reason=travel_reason,
                                     meeting_date=meeting_date, meeting_time=meeting_time,
                                     meeting_location=meeting_location)
                if flight_res and "error" not in flight_res:
                    try:
                        clean_out = _clean_for_db(flight_res)
                        r = add_flight_booking(clean_out, ui)
                        _inserted_booking_ids.append(r['booking_id'])
                        st.success(f"✅ Outbound flight saved! DB ID: {r['booking_id']}")
                    except Exception as e:
                        st.error(f"Outbound flight DB error: {e}")
                if return_flight_res and "error" not in return_flight_res:
                    try:
                        clean_ret = _clean_for_db(return_flight_res, out_rec=flight_res)
                        r = add_flight_booking(clean_ret, ui)
                        _inserted_booking_ids.append(r['booking_id'])
                        st.success(f"✅ Return flight saved! DB ID: {r['booking_id']}")
                    except Exception as e:
                        st.error(f"Return flight DB error: {e}")
                if hotel_res and "error" not in hotel_res:
                    try:
                        # Ensure check_in_date / check_out_date are always set (DB NOT NULL)
                        # Try: scraped fields first, then session state meeting date as fallback
                        _hotel_db_rec = dict(hotel_res)
                        def _to_date_str(v):
                            if v is None: return None
                            if hasattr(v, "strftime"): return v.strftime("%Y-%m-%d")
                            return str(v)[:10] if v else None
                        _ci = (_to_date_str(hotel_res.get("checkin_date"))
                               or _to_date_str(hotel_res.get("check_in_date"))
                               or _to_date_str(hotel_res.get("check_in"))
                               or _to_date_str(st.session_state.get("meeting_date")))
                        _co = (_to_date_str(hotel_res.get("checkout_date"))
                               or _to_date_str(hotel_res.get("check_out_date"))
                               or _to_date_str(hotel_res.get("check_out")))
                        if _ci and not _co:
                            # Default checkout = check_in + 1 day
                            from datetime import datetime as _dt2, timedelta as _td2
                            try:
                                _co = (_dt2.strptime(_ci, "%Y-%m-%d") + _td2(days=1)).strftime("%Y-%m-%d")
                            except Exception:
                                pass
                        if _ci: _hotel_db_rec["check_in_date"]  = _ci
                        if _co: _hotel_db_rec["check_out_date"] = _co
                        # Also normalise key names that add_hotel_booking may expect
                        if _ci: _hotel_db_rec["checkin_date"]   = _ci
                        if _co: _hotel_db_rec["checkout_date"]  = _co
                        r = add_hotel_booking(_hotel_db_rec, ui)
                        _inserted_booking_ids.append(r['booking_id'])
                        st.success(f"✅ Hotel saved! DB ID: {r['booking_id']}")
                    except Exception as e:
                        st.error(f"Hotel DB error: {e}")
                # Track inserted IDs for revert
                if _inserted_booking_ids:
                    st.session_state["_last_db_booking_ids"] = _inserted_booking_ids
            except Exception as outer_e:
                st.error(f"Error saving to database: {outer_e}")

    _MASTER_XLSX = os.path.join(os.path.dirname(os.path.abspath(__file__)), "SI2Tech_Travel_Master.xlsx")

    if st.button("📊 Save to Master Excel", type="secondary", width='stretch'):
        st.session_state["show_welcome"] = False
        flight_res        = st.session_state.get("scraped_flight")
        hotel_res         = st.session_state.get("scraped_hotel")
        return_flight_res = st.session_state.get("scraped_return_flight")
        if not flight_res and not hotel_res and not return_flight_res:
            st.warning("⚠️ No scraped data. Upload and scrape receipts first.")
        else:
            try:
                new_f_rows, new_h_rows = _collect_export_rows(
                    flight_res=flight_res if (flight_res and "error" not in flight_res) else None,
                    hotel_res=hotel_res   if (hotel_res  and "error" not in hotel_res)  else None,
                    project_number=project_number, travel_reason=travel_reason,
                    meeting_date=meeting_date, meeting_time=meeting_time,
                    meeting_location=meeting_location,
                    return_flight_res=return_flight_res if (return_flight_res and "error" not in return_flight_res) else None,
                )
                # Track row counts before insert for revert
                _xlsx_rows_before = {"flights": 0, "hotels": 0}
                if os.path.exists(_MASTER_XLSX):
                    try:
                        import openpyxl as _opx
                        _wb_pre = _opx.load_workbook(_MASTER_XLSX, read_only=True)
                        if "Flight Bookings" in _wb_pre.sheetnames:
                            _xlsx_rows_before["flights"] = _wb_pre["Flight Bookings"].max_row
                        if "Hotel Bookings" in _wb_pre.sheetnames:
                            _xlsx_rows_before["hotels"] = _wb_pre["Hotel Bookings"].max_row
                        _wb_pre.close()
                    except Exception:
                        pass
                total_f, total_h = _append_to_master_xlsx(new_f_rows, new_h_rows, _MASTER_XLSX)
                st.session_state["_last_xlsx_rows_before"] = _xlsx_rows_before
                st.success(f"✅ Saved! Master Excel now has **{total_f}** flight row(s) and **{total_h}** hotel row(s).")
            except Exception as e:
                st.error(f"Export error: {e}")

    # ── Revert Last Insertion ─────────────────────────────────────────────
    _has_db_revert   = bool(st.session_state.get("_last_db_booking_ids"))
    _has_xlsx_revert = bool(st.session_state.get("_last_xlsx_rows_before"))
    if _has_db_revert or _has_xlsx_revert:
        if st.button("↩ Revert Last Insertion", type="secondary", width='stretch'):
            _reverted_any = False
            # Revert DB
            if _has_db_revert:
                try:
                    from addtodatabase_v2 import revert_booking
                    for _bid in st.session_state["_last_db_booking_ids"]:
                        revert_booking(_bid)
                    st.success(f"✅ DB reverted: booking ID(s) {st.session_state['_last_db_booking_ids']} deleted.")
                    st.session_state.pop("_last_db_booking_ids", None)
                    st.session_state.pop("_cached_flight_txns", None)
                    st.session_state.pop("_cached_hotel_txns", None)
                    _reverted_any = True
                except Exception as _rev_err:
                    st.error(f"DB revert failed: {_rev_err}")
            # Revert Excel
            if _has_xlsx_revert and os.path.exists(_MASTER_XLSX):
                try:
                    import openpyxl as _opx
                    _wb = _opx.load_workbook(_MASTER_XLSX)
                    _before = st.session_state["_last_xlsx_rows_before"]
                    for _sheet_name, _key in [("Flight Bookings", "flights"), ("Hotel Bookings", "hotels")]:
                        if _sheet_name in _wb.sheetnames:
                            _ws = _wb[_sheet_name]
                            _old_max = _before.get(_key, 0)
                            if _old_max > 0 and _ws.max_row > _old_max:
                                _ws.delete_rows(_old_max + 1, _ws.max_row - _old_max)
                    _wb.save(_MASTER_XLSX)
                    _wb.close()
                    st.success("✅ Excel reverted: last inserted rows removed.")
                    st.session_state.pop("_last_xlsx_rows_before", None)
                    _reverted_any = True
                except Exception as _xlsx_rev_err:
                    st.error(f"Excel revert failed: {_xlsx_rev_err}")
            if _reverted_any:
                st.rerun()

    if os.path.exists(_MASTER_XLSX):
        _fsize_kb = os.path.getsize(_MASTER_XLSX) / 1024
        st.caption(f"📁 SI2Tech_Travel_Master.xlsx  ({_fsize_kb:.1f} KB)")

    if meeting_date is None:
        meeting_date = datetime.now().date()
    if meeting_time is None:
        meeting_time = datetime.strptime("09:00", "%H:%M").time()
    meet_dt           = datetime.combine(meeting_date, meeting_time)
    date_str          = meeting_date.strftime("%Y-%m-%d")
    target_arrival_dt = meet_dt - timedelta(hours=3)


# ── Helpers ───────────────────────────────────────────────────────────────────

def _render_search_header(origin, destination, origin_input, dest_input, meeting_date,
                          target_arrival_dt, is_round_trip=False, meeting_end_date=None):
    target_time_str = target_arrival_dt.strftime("%I:%M %p") if target_arrival_dt else "N/A"
    route_icon = "&#x21C4;" if is_round_trip else "&#x2708;"
    if is_round_trip and meeting_end_date and meeting_end_date != meeting_date:
        date_label = f"{meeting_date.strftime('%b %d')} &rarr; {meeting_end_date.strftime('%b %d')}"
    else:
        date_label = meeting_date.strftime('%b %d, %Y')
    cutoff_badge = (
        f"<div style='margin-top:6px;display:inline-block;background:rgba(15,157,88,0.12);"
        f"border:1px solid #0f9d58;color:#0f9d58;padding:3px 10px;"
        f"border-radius:4px;font-size:0.8em;font-weight:600;'>&#x2B06; Must land by {target_time_str}</div>"
        if target_arrival_dt else ""
    )
    st.markdown(f"""
<div style="
    display:flex;
    flex-wrap:wrap;
    align-items:center;
    justify-content:space-between;
    gap:10px;
    background:#161b27;
    border:1px solid #2d3748;
    border-radius:10px;
    padding:14px 18px;
    margin-bottom:8px;
    box-sizing:border-box;
    width:100%;
    overflow:hidden;
">
  <div style="flex:1 1 120px;min-width:0;">
    <div style="font-size:0.62em;color:#6b7280;font-weight:700;text-transform:uppercase;
                letter-spacing:0.08em;margin-bottom:3px;">Departure ({origin})</div>
    <div style="font-size:1.55em;font-weight:800;color:#f3f4f6;line-height:1.15;
                overflow:hidden;text-overflow:ellipsis;white-space:nowrap;">{origin_input}</div>
  </div>
  <div style="flex:0 0 auto;text-align:center;color:#9ca3af;font-size:1.3em;padding:0 4px;">
    {route_icon}
    <div style="font-size:0.42em;color:#6b7280;margin-top:3px;white-space:nowrap;">{date_label}</div>
  </div>
  <div style="flex:1 1 120px;min-width:0;text-align:right;">
    <div style="font-size:0.62em;color:#6b7280;font-weight:700;text-transform:uppercase;
                letter-spacing:0.08em;margin-bottom:3px;">Target Arrival ({destination})</div>
    <div style="font-size:1.55em;font-weight:800;color:#f3f4f6;line-height:1.15;
                overflow:hidden;text-overflow:ellipsis;white-space:nowrap;">{dest_input}</div>
    {cutoff_badge}
  </div>
</div>
""", unsafe_allow_html=True)


def _hotel_mode(m_count, e_count, exec_count=0):
    total = m_count + e_count + exec_count
    if total == 0: return "none"
    if m_count > 0 and e_count > 0: return "dual"
    if total > 0: return "single"
    return "none"

def _handle_back_from_email():
    # Navigate back to the verify/approval page, not all the way back to results
    st.session_state["view"] = "verify"

def _handle_back_from_verify():
    st.session_state["view"] = "results"

@st.fragment
def _select_button_flight(f):
    fk = _flight_key(f)
    def _toggle():
        if fk in st.session_state["selected_flight_keys"]:
            st.session_state["selected_flight_keys"].discard(fk)
        else:
            st.session_state["selected_flight_keys"].add(fk)
    st.button("✅" if fk in st.session_state["selected_flight_keys"] else "Select",
              key=f"btn_flight_{fk}", on_click=_toggle, type="secondary", width='stretch')


def _render_return_flight_card(f_item, origin, destination, disp_t_count, index):
    """Renders a single return-flight card with Select + Book buttons."""
    si = int(f_item.get("stops", 0))
    fk = _flight_key(f_item)
    with st.container():
        f_c1, f_c2, f_c3 = st.columns([1, 2, 1])
        with f_c1:
            f_img = f_item.get("thumbnail")
            st.image(f_img if isinstance(f_img, str) and f_img.startswith("http")
                     else "https://cdn-icons-png.flaticon.com/512/727/727142.png",
                     width=80 if isinstance(f_img, str) and f_img.startswith("http") else 40)
            st.markdown(f"**{f_item['airline']}**")
            st.caption(f_item["flight_no"])
            if si == 0:
                st.markdown("<span style='background:#0f9d58;color:white;padding:2px 8px;"
                            "border-radius:10px;font-size:0.75em;'>Non-stop</span>",
                            unsafe_allow_html=True)
            else:
                st.markdown(f"<span style='background:#e37400;color:white;padding:2px 8px;"
                            f"border-radius:10px;font-size:0.75em;'>{si} Stop(s)</span>",
                            unsafe_allow_html=True)
        with f_c2:
            st.markdown(f"**{f_item['dep_time']}** ⎯ **{f_item['arr_time']}**")
            st.caption(f"⏱ {f_item['duration']}")
            if f_item.get("same_day_return"):
                st.caption("✅ Same-day return — no hotel needed")
            elif f_item.get("needs_hotel"):
                ci  = f_item.get("hotel_check_in")
                co  = f_item.get("hotel_check_out")
                ci_s = ci.strftime("%b %d") if ci else "?"
                co_s = co.strftime("%b %d") if co else "?"
                st.caption(f"🏨 Overnight stay at destination ({ci_s} → {co_s})")
        with f_c3:
            pl = f"₹{int(f_item['price'])}"
            if disp_t_count > 1:
                st.markdown(f"<div style='text-align:center;font-weight:bold;font-size:1.5em;'>{pl}</div>"
                            f"<div style='text-align:center;color:grey;font-size:.8em;'>Total for {disp_t_count}</div>",
                            unsafe_allow_html=True)
            else:
                st.markdown(f"<div style='text-align:center;font-weight:bold;font-size:1.5em;'>{pl}</div>",
                            unsafe_allow_html=True)

            @st.fragment
            def _return_select_btn(key=fk):
                is_sel = key in st.session_state["selected_return_flight_keys"]
                def _toggle():
                    if key in st.session_state["selected_return_flight_keys"]:
                        st.session_state["selected_return_flight_keys"].discard(key)
                    else:
                        st.session_state["selected_return_flight_keys"].add(key)
                btn_label = "✅ Selected" if is_sel else "Select"
                st.button(btn_label, key=f"ret_sel_{key}_{index}",
                          on_click=_toggle, type="primary" if is_sel else "secondary",
                          use_container_width=True)
            _return_select_btn()
            # Use <a> tag — st.link_button URL-encodes & in Chrome, breaking multi-param URLs
            st.markdown(
                f'<a href="{f_item["link"]}" target="_blank" rel="noopener noreferrer" '
                f'style="display:block;width:100%;text-align:center;background:#1d4ed8;color:#fff;'
                f'text-decoration:none;padding:10px 16px;border-radius:8px;font-weight:600;font-size:14px;">'
                f'Book Return</a>',
                unsafe_allow_html=True,
            )
    st.divider()


def _render_verify_view():
    _rebuild_selected_lists()
    flights = st.session_state.get("selected_flights", [])
    hotels  = st.session_state.get("selected_hotels",  [])
    if not flights and not hotels:
        st.warning("⚠️ No items selected. Please go back and select flights/hotels.")
        if st.button("← Back to Results"):
            _handle_back_from_verify(); st.rerun()
        return
    c1, c2 = st.columns([1, 4])
    with c1:
        if st.button("← Back", width='stretch'):
            _handle_back_from_verify(); st.rerun()
    with c2:
        st.title("✅ Verify & Customize Approval Data")
    st.markdown("---")
    with st.form("main_verify_form"):
        if flights:
            st.subheader("🛫 Selected Flights")
            for idx, f in enumerate(flights):
                is_return = f.get("_is_return", False)
                label_tag = " 🔄 Return" if is_return else " ✈️ Outbound"
                with st.container(border=True):
                    st.caption(label_tag)
                    cols = st.columns([2, 2, 2])
                    with cols[0]:
                        st.markdown(f"**{f.get('airline')}**")
                        st.caption(f"Flight {f.get('flight_no')} | {f.get('dep_time')} → {_clean_time_str(f.get('arr_time', ''))}")
                    with cols[1]:
                        f_price = st.number_input("Final Price (₹)", value=float(f.get('custom_price', f.get('price', 0))), key=f"vf_p_{idx}")
                    with cols[2]:
                        st.selectbox("Booking Vendor", [
                            "MakeMyTrip", "Cleartrip", "Yatra", "EaseMyTrip", "Goibibo",
                            "Booking.com", "Expedia", "Trip.com",
                            "Air India (Direct)", "IndiGo (Direct)", "Vistara (Direct)",
                            "SpiceJet (Direct)", "Akasa Air (Direct)", "Airline Website (Direct)",
                            "Corporate Travel Desk", "Offline Agent", "Other"
                        ], key=f"vf_v_{idx}")
                    st.text_input("Custom vendor name", key=f"vf_vc_{idx}", placeholder="Type vendor name (used when 'Other' is selected)")
        st.markdown("<br>", unsafe_allow_html=True)
        if hotels:
            st.subheader("🏨 Selected Hotels")
            for idx, h in enumerate(hotels):
                with st.container(border=True):
                    cols = st.columns([2, 2, 2])
                    with cols[0]:
                        st.markdown(f"**{h.get('name')}**")
                        st.caption(f"⭐ {h.get('rating', 'N/A')} | {h.get('distance_fmt', '')}")
                    with cols[1]:
                        h_price = st.number_input("Price per night (₹)", value=float(h.get('custom_price', h.get('price_raw', 0))), key=f"vh_p_{idx}")
                    with cols[2]:
                        st.selectbox("Booking Vendor", [
                            "MakeMyTrip", "Cleartrip", "Yatra", "EaseMyTrip", "Goibibo",
                            "Booking.com", "Agoda", "Expedia", "Trip.com",
                            "Hotels.com", "OYO", "Treebo", "FabHotels",
                            "Direct Hotel Website", "Other"
                        ], key=f"vh_v_{idx}")
                    st.text_input("Custom vendor name", key=f"vh_vc_{idx}", placeholder="Type vendor name (used when 'Other' is selected)")
        st.markdown("<br>", unsafe_allow_html=True)
        if st.form_submit_button("✨ Done & Generate Approval Draft", type="primary", use_container_width=True):
            for idx, f in enumerate(flights):
                f['custom_price']   = st.session_state.get(f"vf_p_{idx}", f.get('price', 0))
                vendor_sel          = st.session_state.get(f"vf_v_{idx}", "MakeMyTrip")
                vendor_custom       = st.session_state.get(f"vf_vc_{idx}", "").strip()
                f['booking_vendor'] = vendor_custom if vendor_sel == "Other" and vendor_custom else vendor_sel
            for idx, h in enumerate(hotels):
                h['custom_price']   = st.session_state.get(f"vh_p_{idx}", h.get('price_raw', 0))
                vendor_sel          = st.session_state.get(f"vh_v_{idx}", "MakeMyTrip")
                vendor_custom       = st.session_state.get(f"vh_vc_{idx}", "").strip()
                h['booking_vendor'] = vendor_custom if vendor_sel == "Other" and vendor_custom else vendor_sel
            st.session_state["draft_email_triggered"] = True
            st.session_state["view"] = "results"
            st.rerun()


def _render_hotel_card(h, ci_dt, co_dt, adults, meeting_loc="", city_name="",
                       label_color="#1a73e8", shared_badge=False, group="", index=0, rooms=1,
                       city_code="", property_token=""):
    with st.container(border=True):
        if shared_badge:
            st.markdown(f"<div style='background:{label_color}18;border-left:3px solid {label_color};"
                        f"padding:3px 8px;margin-bottom:8px;font-size:0.75em;color:{label_color};"
                        f"font-weight:600;'>✅ Available for both groups</div>", unsafe_allow_html=True)
        col_img, col_info = st.columns([1.2, 2.5])
        with col_img:
            img = get_safe_image(h.get("thumbnail"))
            st.image(img if img else "https://cdn-icons-png.flaticon.com/512/2983/2983803.png", width='stretch')
        with col_info:
            dist_val = h.get("distance_val", 0.0)
            dist_str = f"📍 {h.get('distance_fmt', '')}" if dist_val and dist_val > 0.0 else "📍 Near meeting location"
            st.markdown(f"<p style='margin:0 0 2px;font-weight:700;font-size:1em;'>{h.get('name','')}</p>", unsafe_allow_html=True)
            st.markdown(f"<p style='margin:0 0 2px;font-size:0.85em;color:#9ca3af;'>⭐ {h.get('rating','N/A')} · {dist_str}</p>", unsafe_allow_html=True)
            try:
                nights = (co_dt - ci_dt).days
                nights_label = f" · {nights} night{'s' if nights != 1 else ''}"
            except Exception:
                nights_label = ""
            st.markdown(f"<p style='margin:0 0 6px;font-size:0.85em;color:#0f9d58;'>💰 <strong>{h.get('price_fmt', 'N/A')}</strong>{nights_label}</p>", unsafe_allow_html=True)

        act_c1, act_c2 = st.columns([1, 1])
        with act_c1:
            @st.fragment
            def _render_hotel_select_btn():
                hk = _hotel_key(h)
                is_sel = hk in st.session_state["selected_hotel_keys"]
                def _toggle_sel():
                    if is_sel: st.session_state["selected_hotel_keys"].discard(hk)
                    else: st.session_state["selected_hotel_keys"].add(hk)
                st.button("✅ Added" if is_sel else "➕ Add",
                          key=f"sel_{group}_{index}_{hk}", on_click=_toggle_sel, type="primary", width='stretch')
            _render_hotel_select_btn()
        with act_c2:
            @st.fragment
            def _render_hotel_book_btn():
                hk       = _hotel_key(h)
                book_url = get_live_hotel_url(h, ci_dt, co_dt, adults,
                                             meeting_loc=meeting_loc, city_name=city_name, rooms=rooms)
                if st.button("🚀 Book", key=f"book_{group}_{index}_{hk}", type="secondary", width='stretch'):
                    webbrowser.open_new_tab(book_url)
                    st.toast("🌍 Opening Google Hotels...", icon="⚡")
            _render_hotel_book_btn()


def _render_tracking_overlay():
    st.title("📄 Upload and Track Receipts")
    st.markdown("---")

    if st.button("← Back to Results", width='stretch'):
        st.session_state["show_tracking_overlay"] = False
        st.session_state["view"] = "results"
        st.rerun()

    tab_upload, tab_flights, tab_hotels = st.tabs(
        ["Upload Receipts", "Flight Transactions", "Hotel Transactions"]
    )

    # ── Tab 2: Flight Transactions ──────────────────────────────────────────
    with tab_flights:
        if st.button("🔄 Refresh", key="refresh_flights"):
            st.session_state.pop("_cached_flight_txns", None)
        if "_cached_flight_txns" not in st.session_state:
            try:
                st.session_state["_cached_flight_txns"] = fetch_flight_transactions()
            except Exception as _ft_err:
                st.error(f"Could not load flight transactions: {_ft_err}")
                st.session_state["_cached_flight_txns"] = []
        _f_txns = st.session_state["_cached_flight_txns"]
        if _f_txns:
            import pandas as _pd_f
            # Pivot: collapse passenger rows into one row per segment
            _f_pivoted = {}  # key = (booking_id, pnr, flight_number, origin, destination, travel_date)
            for row in _f_txns:
                _fk = (row.get("booking_id"), row.get("pnr"), row.get("flight_number"),
                        row.get("origin"), row.get("destination"), str(row.get("travel_date")))
                if _fk not in _f_pivoted:
                    _f_pivoted[_fk] = {
                        "Project No": row.get("project_no"),
                        "Booking Date": row.get("booking_date"),
                        "Airline": row.get("airline"),
                        "Flight No": row.get("flight_number"),
                        "Origin": row.get("origin"),
                        "Destination": row.get("destination"),
                        "Travel Date": row.get("travel_date"),
                        "Amount": row.get("segment_amount"),
                        "PNR": row.get("pnr"),
                        "_passengers": [],
                    }
                pax = row.get("passenger")
                if pax and pax not in _f_pivoted[_fk]["_passengers"]:
                    _f_pivoted[_fk]["_passengers"].append(pax)
            # Build passenger columns
            _f_rows = list(_f_pivoted.values())
            _f_max_pax = max((len(r["_passengers"]) for r in _f_rows), default=0)
            for r in _f_rows:
                for i in range(_f_max_pax):
                    r[f"Passenger {i+1}"] = r["_passengers"][i] if i < len(r["_passengers"]) else ""
                del r["_passengers"]
            _f_display = _pd_f.DataFrame(_f_rows)
            # Metrics — amount counted once per segment (already de-duped by pivot)
            _total_segments = len(_f_display)
            _f_spend = _f_display["Amount"].dropna().sum() if "Amount" in _f_display.columns else 0
            m1, m2 = st.columns(2)
            m1.metric("Total Segments", _total_segments)
            m2.metric("Total Spend", f"₹{_f_spend:,.0f}" if _f_spend else "N/A")
            st.dataframe(_f_display, use_container_width=True, hide_index=True)
        else:
            st.info("No flight transactions found.")

    # ── Tab 3: Hotel Transactions ───────────────────────────────────────────
    with tab_hotels:
        if st.button("🔄 Refresh", key="refresh_hotels"):
            st.session_state.pop("_cached_hotel_txns", None)
        if "_cached_hotel_txns" not in st.session_state:
            try:
                st.session_state["_cached_hotel_txns"] = fetch_hotel_transactions()
            except Exception as _ht_err:
                st.error(f"Could not load hotel transactions: {_ht_err}")
                st.session_state["_cached_hotel_txns"] = []
        _h_txns = st.session_state["_cached_hotel_txns"]
        if _h_txns:
            import pandas as _pd_h
            # Pivot: collapse passenger rows into one row per hotel stay
            _h_pivoted = {}  # key = (booking_id, hotel_name)
            for row in _h_txns:
                _hk = (row.get("booking_id"), row.get("hotel_name"))
                if _hk not in _h_pivoted:
                    _h_pivoted[_hk] = {
                        "Project No": row.get("project_no"),
                        "Reason": row.get("reason"),
                        "Booking Date": row.get("booking_date"),
                        "Hotel Name": row.get("hotel_name"),
                        "Amount": row.get("hotel_amount"),
                        "_passengers": [],
                    }
                pax = row.get("passenger")
                if pax and pax not in _h_pivoted[_hk]["_passengers"]:
                    _h_pivoted[_hk]["_passengers"].append(pax)
            # Build passenger columns
            _h_rows = list(_h_pivoted.values())
            _h_max_pax = max((len(r["_passengers"]) for r in _h_rows), default=0)
            for r in _h_rows:
                for i in range(_h_max_pax):
                    r[f"Passenger {i+1}"] = r["_passengers"][i] if i < len(r["_passengers"]) else ""
                del r["_passengers"]
            _h_display = _pd_h.DataFrame(_h_rows)
            # Metrics — amount counted once per stay (already de-duped by pivot)
            _h_bookings = len(_h_display)
            _h_spend = _h_display["Amount"].dropna().sum() if "Amount" in _h_display.columns else 0
            h1, h2 = st.columns(2)
            h1.metric("Total Hotel Bookings", _h_bookings)
            h2.metric("Total Spend", f"₹{_h_spend:,.0f}" if _h_spend else "N/A")
            st.dataframe(_h_display, use_container_width=True, hide_index=True)
        else:
            st.info("No hotel transactions found.")

    # ── Tab 1: Upload Receipts (existing logic) ────────────────────────────
    with tab_upload:
        col_prj, col_rsn = st.columns(2)
        with col_prj:
            _trk_proj = st.text_input("Project Number", value=st.session_state.get("project_number", ""), key="trk_project_number")
        with col_rsn:
            _trk_reason = st.text_input("Reason", value=st.session_state.get("travel_reason", ""), key="trk_travel_reason")
        st.session_state["project_number"] = _trk_proj
        st.session_state["travel_reason"] = _trk_reason

        st.markdown("""
        > **ℹ️ How it works:** Upload your flight receipt PDF — it contains both the outbound **and** return legs.
        > The system will automatically extract both legs and swap origin/destination for the return flight.
        > Upload a separate hotel PDF if applicable.
        """)

        col_flight, col_hotel = st.columns(2)
        with col_flight:
            st.subheader("✈️ Flight Receipt (Outbound + Return)")
            flight_pdf = st.file_uploader("Upload Flight PDF", type=["pdf"], key="flight_pdf_uploader")
        with col_hotel:
            st.subheader("🏨 Hotel Receipt")
            hotel_pdf = st.file_uploader("Upload Hotel PDF", type=["pdf"], key="hotel_pdf_uploader")

        st.markdown("---")
        if st.button("🚀 Scrape Receipts", type="primary", use_container_width=True):
            if not pdfScrapper:
                st.error("pdfScrapper module not found."); return
            with st.spinner("Extracting data..."):
                flight_res = hotel_res = return_flight_res = None

                # Clear stale results from previous scrape runs
                if not flight_pdf:
                    st.session_state["scraped_flight"]        = None
                    st.session_state["scraped_return_flight"] = None
                if not hotel_pdf:
                    st.session_state["scraped_hotel"] = None

                if flight_pdf:
                    flight_path = f"temp_flight_{uuid.uuid4().hex[:8]}.pdf"
                    with open(flight_path, "wb") as f: f.write(flight_pdf.getbuffer())
                    try:
                        # extract_flight may return a list (multi-leg) or a single dict
                        raw = pdfScrapper.extract_flight(flight_path)
                        if isinstance(raw, list) and len(raw) >= 2:
                            # Multi-leg: first item = outbound, second = return
                            flight_res = raw[0]
                            return_flight_res = raw[1]
                        elif isinstance(raw, list) and len(raw) == 1:
                            flight_res = raw[0]
                            return_flight_res = None
                        else:
                            # Single dict — treat as outbound only
                            flight_res = raw
                            return_flight_res = None

                        # Auto-build return flight record from outbound if not separately extracted.
                        # Priority: use return_* fields scraped from PDF; only fall back to
                        # outbound fields when the return-specific ones are absent.
                        if flight_res and "error" not in flight_res and return_flight_res is None:
                            _ret_airline = (
                                flight_res.get("return_airline")         # explicit return airline from scraper
                                or flight_res.get("return_carrier")
                                or ""                                    # do NOT fall back to outbound airline
                            )
                            _ret_fno = (
                                flight_res.get("return_flight_number")   # explicit return flight number
                                or flight_res.get("return_flight_no")
                                or ""                                    # do NOT fall back to outbound flight_no
                            )
                            return_flight_res = {
                                "_is_return":          True,
                                "airline":             _ret_airline,
                                "flight_number":       _ret_fno,
                                # Swap origin ↔ destination
                                "origin":              flight_res.get("destination", ""),
                                "destination":         flight_res.get("origin", ""),
                                "departure_datetime":  flight_res.get("return_departure_datetime") or "",
                                "arrival_datetime":    flight_res.get("return_arrival_datetime",  "") or "",
                                "traveler_name":       flight_res.get("traveler_name", ""),
                                "pnr":                 flight_res.get("return_pnr") or flight_res.get("pnr", ""),
                                "total_amount":        flight_res.get("return_total_amount") or flight_res.get("total_amount", ""),
                                "currency":            flight_res.get("currency", "INR"),
                                "ota_source":          flight_res.get("ota_source", ""),
                                "date_of_booking":     flight_res.get("date_of_booking", ""),
                            }

                        # ── Infer return departure date if not explicitly scraped ──────────
                        if return_flight_res and not return_flight_res.get("departure_datetime"):
                            try:
                                _groq_key = os.getenv("GROQ_API_KEY")
                                _booking_date = flight_res.get("date_of_booking", "")
                                # Collect candidate date fields from the outbound ticket
                                _candidate_vals = [
                                    flight_res.get("departure_datetime"),
                                    flight_res.get("arrival_datetime"),
                                    flight_res.get("travel_date"),
                                    flight_res.get("return_departure_datetime"),
                                    flight_res.get("return_travel_date"),
                                ]
                                _candidate_dates = list(dict.fromkeys(
                                    v for v in _candidate_vals
                                    if v and str(v).strip() and str(v).strip() != str(_booking_date).strip()
                                ))
                                if _groq_key and _candidate_dates:
                                    _gc = Groq(api_key=_groq_key)
                                    _prompt = (
                                        "You are given date values from a flight ticket. "
                                        f"Booking date: {_booking_date}. "
                                        f"All dates found: {_candidate_dates}. "
                                        "TASK: Identify the return/onward travel date (exclude booking date). "
                                        "Rules: "
                                        "- If only ONE travel date exists, that IS the return date (same-day return). "
                                        "- If multiple travel dates exist, the LAST one is the return date (others are stops). "
                                        "Respond with ONLY a date in YYYY-MM-DD format. Nothing else."
                                    )
                                    _resp = _gc.chat.completions.create(
                                        model="meta-llama/llama-4-scout-17b-16e-instruct",
                                        messages=[{"role": "user", "content": _prompt}],
                                        max_tokens=20, temperature=0,
                                    )
                                    _raw_ret_dt = _resp.choices[0].message.content.strip()
                                    import re as _re_ret
                                    if _re_ret.match(r"\d{4}-\d{2}-\d{2}", _raw_ret_dt):
                                        return_flight_res["departure_datetime"] = _raw_ret_dt[:10]
                                        print(f"  ✅ LLM inferred return date: {_raw_ret_dt[:10]}")
                                elif _candidate_dates:
                                    # Fallback without LLM: use last date in candidates
                                    return_flight_res["departure_datetime"] = str(_candidate_dates[-1])[:10]
                            except Exception as _ret_date_err:
                                print(f"  ⚠️ Return date inference error: {_ret_date_err}")
                        # ──────────────────────────────────────────────────────────────────

                        st.session_state["scraped_flight"] = flight_res
                        st.session_state["scraped_return_flight"] = return_flight_res

                    except Exception as e:
                        flight_res = {"error": str(e)}
                        st.session_state["scraped_flight"] = flight_res
                    finally:
                        if os.path.exists(flight_path): os.remove(flight_path)

                if hotel_pdf:
                    hotel_path = f"temp_hotel_{uuid.uuid4().hex[:8]}.pdf"
                    with open(hotel_path, "wb") as f: f.write(hotel_pdf.getbuffer())
                    try:
                        hotel_res = pdfScrapper.extract_hotel(hotel_path)
                        st.session_state["scraped_hotel"] = hotel_res
                    except Exception as e:
                        hotel_res = {"error": str(e)}
                    finally:
                        if os.path.exists(hotel_path): os.remove(hotel_path)

            any_result = flight_res or hotel_res
            if any_result:
                res_cols = st.columns(3)
                with res_cols[0]:
                    if flight_res and "error" not in flight_res:
                        st.success("✅ Outbound Flight Extracted!")
                        with st.container(border=True):
                            st.markdown(f"### ✈️ {flight_res.get('airline', 'Flight Details')}")
                            st.divider()
                            fc1, fc2 = st.columns(2)
                            with fc1:
                                st.markdown(f"**From:** {flight_res.get('origin', 'N/A')}")
                                st.markdown(f"**Flight No:** {flight_res.get('flight_number', 'N/A')}")
                                st.markdown(f"**Departure:** {flight_res.get('departure_datetime', 'N/A')}")
                            with fc2:
                                st.markdown(f"**To:** {flight_res.get('destination', 'N/A')}")
                                st.markdown(f"**PNR:** {flight_res.get('pnr', 'N/A')}")
                                st.markdown(f"**Arrival:** {flight_res.get('arrival_datetime', 'N/A')}")
                            st.divider()
                            st.markdown(f"**Traveler:** {flight_res.get('traveler_name', 'N/A')}")
                            st.markdown(f"**Total:** {flight_res.get('currency', '')} {flight_res.get('total_amount', 'N/A')}")
                    elif flight_res and "error" in flight_res:
                        st.error(f"Outbound failed: {flight_res['error']}")

                with res_cols[1]:
                    return_flight_res_display = st.session_state.get("scraped_return_flight")
                    if return_flight_res_display and "error" not in return_flight_res_display:
                        label = "🔄 Return Flight (Auto-derived)" if not isinstance(raw if 'raw' in dir() else None, list) else "🔄 Return Flight Extracted!"
                        st.success(label)
                        with st.container(border=True):
                            st.markdown(f"### 🔄 {return_flight_res_display.get('airline', 'Return Flight')}")
                            st.divider()
                            rc1, rc2 = st.columns(2)
                            with rc1:
                                st.markdown(f"**From:** {return_flight_res_display.get('origin', 'N/A')}")
                                st.markdown(f"**Flight No:** {return_flight_res_display.get('flight_number', 'N/A')}")
                                st.markdown(f"**Departure:** {return_flight_res_display.get('departure_datetime', 'N/A')}")
                            with rc2:
                                st.markdown(f"**To:** {return_flight_res_display.get('destination', 'N/A')}")
                                st.markdown(f"**PNR:** {return_flight_res_display.get('pnr', 'N/A')}")
                                st.markdown(f"**Arrival:** {return_flight_res_display.get('arrival_datetime', 'N/A')}")
                            st.divider()
                            st.markdown(f"**Traveler:** {return_flight_res_display.get('traveler_name', 'N/A')}")
                            st.markdown(f"**Total:** {return_flight_res_display.get('currency', '')} {return_flight_res_display.get('total_amount', 'N/A')}")
                    elif not flight_res or "error" in (flight_res or {}):
                        st.info("Upload a flight PDF to auto-extract return leg.")

                with res_cols[2]:
                    if hotel_res and "error" not in hotel_res:
                        st.success("✅ Hotel Data Extracted!")
                        with st.container(border=True):
                            st.markdown(f"### 🏨 {hotel_res.get('hotel_name', 'Hotel Details')}")
                            st.divider()
                            hc1, hc2 = st.columns(2)
                            with hc1:
                                st.markdown(f"**Check-in:** {hotel_res.get('checkin_date', 'N/A')}")
                                st.markdown(f"**Booking ID:** {hotel_res.get('booking_id', 'N/A')}")
                            with hc2:
                                st.markdown(f"**Check-out:** {hotel_res.get('checkout_date', 'N/A')}")
                                st.markdown(f"**Traveler:** {hotel_res.get('traveler_name', 'N/A')}")
                            st.divider()
                            st.markdown(f"**Total:** {hotel_res.get('currency', '')} {hotel_res.get('total_amount', 'N/A')}")
                    elif hotel_res and "error" in hotel_res:
                        st.error(f"Hotel failed: {hotel_res['error']}")
            else:
                st.warning("Please upload at least one PDF.")


def _render_dual_hotels(hotel_results_mgmt, hotel_results_emp, shared_hotels,
                        ci_dt, co_dt, m_count, e_count, meeting_location, destination,
                        ci_str="", co_str="", city_code="", check_in_date=None,
                        shared_from_expand=False, rooms=1):
    city_name   = get_airport_city(destination)
    shared_list = shared_hotels
    if shared_list:
        header_note = " (nearby area)" if shared_from_expand else ""
        st.markdown("<div style='background:linear-gradient(90deg,#1a73e8,#e37400);color:white;"
                    "padding:6px 12px;border-radius:6px;font-weight:bold;margin-bottom:6px;'>"
                    f"🏨 Shared Hotels — Both Groups Can Stay Here{header_note}</div>",
                    unsafe_allow_html=True)
        for idx, h_item in enumerate(shared_list):
            _render_hotel_card(h_item, ci_dt, co_dt, int(m_count) + int(e_count),
                               meeting_location, city_name, shared_badge=True, label_color="#0f9d58",
                               group="shared", index=idx, rooms=rooms, city_code=city_code,
                               property_token=h_item.get("property_token", ""))
        return

    st.warning("⚠️ No shared hotel found. Enter a minimum budget to find options.")
    with st.form("budget_fallback_form"):
        min_budget = st.number_input("Minimum budget per night (₹)", min_value=500, max_value=50000, value=2000, step=500)
        max_display = min_budget + 4000
        st.caption(f"Search range: ₹{int(min_budget):,} – ₹{int(max_display):,} / night")
        submitted = st.form_submit_button("🔍 Find Cheapest Hotel", type="primary", width='stretch')
    if submitted:
        total_guests = int(m_count) + int(e_count)
        with st.spinner(f"Searching (₹{int(min_budget):,}–₹{int(max_display):,})…"):
            fallback_hotels = search_hotels_budget_fallback(
                meeting_location, ci_str, co_str, total_guests, int(min_budget), dest_airport_iata=destination)
        if not fallback_hotels:
            st.error("No hotels found. Try a different budget range.")
        else:
            cheapest = fallback_hotels[0]
            st.success(f"✅ Found {len(fallback_hotels)} hotel(s).")
            _render_hotel_card(cheapest, ci_dt, co_dt, total_guests, meeting_location, city_name,
                               group="budget_fallback", index=0, rooms=rooms, city_code=city_code)
            st.session_state["hotel_results_shared"] = [cheapest]


# ── SEARCH + RESULTS ──────────────────────────────────────────────────────────

if search_triggered:
    st.session_state["show_welcome"] = False
    for _cbk in list(st.session_state.keys()):
        if _cbk.startswith("cb_flight_") or _cbk.startswith("cb_hotel_"):
            del st.session_state[_cbk]
    st.session_state.update({
        "selected_flight_keys": set(), "selected_hotel_keys": set(),
        "selected_flights": [], "selected_hotels": [],
        "email_package": None, "view": "results",
        "return_flights": [], "is_round_trip": is_round_trip,
        "selected_return_flight_keys": set(), "selected_return_flights": [],
    })
    for k in ["hotel_results_mgmt","hotel_results_emp","hotel_results_shared","display_flights"]:
        st.session_state[k] = []

    hotel_results_mgmt = []; hotel_results_emp = []; hotel_results_shared = []; display_flights = []
    st.session_state["last_meeting_location"] = meeting_location

    meet_dt            = datetime.combine(meeting_date, meeting_time)
    target_arrival_dt  = None; logistics = None
    h_mode             = _hotel_mode(int(m_count), int(e_count), int(exec_count))

    with st.spinner("🤖 Calculating optimal travel window..."):
        target_arrival_dt, logistics = calculate_dynamic_cutoff(destination, meeting_location, meet_dt)

    _render_search_header(origin, destination, origin_input, dest_input, meeting_date,
                          target_arrival_dt, is_round_trip=is_round_trip,
                          meeting_end_date=meeting_end_date if is_round_trip else None)

    _is_intl = is_international(origin, destination)

    if _is_intl or not best_per_airline:
        left_page, right_page = st.columns([1.5, 1], gap="large")
    else:
        left_page = st.columns([1])[0]; right_page = None

    with left_page:
        st.subheader("🛫 Outbound Flights")
        _mmt_rt_date = meeting_end_date if is_round_trip else None
        _mmt_rc = int(st.session_state.get("rooms_count", 1))
        _mmt_intl_note = " (Intl: depart D-1, return D+1)" if is_international(origin, destination) and is_round_trip else (
            " (Intl: depart D-1)" if is_international(origin, destination) else (
            " (Round Trip)" if is_round_trip else ""))
        filtered_flights = []        # initialised here so hotel block can always reference it
        _skip_hotel_same_day_rt = False  # initialised here so hotel block can always reference it
        filtered_flights = []        # scope guard: defined before logistics block
        _skip_hotel_same_day_rt = False
        _mmt_flight_url = make_mmt_link(origin, destination, meeting_date, t_count,
                          arrival_cutoff_dt=target_arrival_dt,
                          is_round_trip=is_round_trip, return_date=_mmt_rt_date,
                          rooms=_mmt_rc)
        # Use markdown <a> instead of st.link_button — Chrome has a URL-encoding bug
        # with st.link_button that strips query params with & causing MMT to open homepage
        st.markdown(
            f'<a href="{_mmt_flight_url}" target="_blank" rel="noopener noreferrer" '
            f'style="display:block;width:100%;text-align:center;background:#1d4ed8;color:#fff;'
            f'text-decoration:none;padding:10px 16px;border-radius:8px;font-weight:600;font-size:14px;">'
            f'✈️ Book on MMT{_mmt_intl_note}</a>',
            unsafe_allow_html=True,
        )
        if logistics and target_arrival_dt:
            d_mins   = logistics.get("drive_mins", 0)
            kms      = logistics.get("distance_km", 0)
            drive_str = f"{d_mins//60}h {d_mins%60}m" if d_mins >= 60 else f"{d_mins}m"
            cutoff_str = target_arrival_dt.strftime("%I:%M %p")
            st.info(
                f"ℹ️ **Must land by {cutoff_str}** — "
                f"**{drive_str}** drive ({kms} km) + **1h** transit buffer "
                f"before {meet_dt.strftime('%I:%M %p')} meeting."
            )
            date_str    = meeting_date.strftime("%Y-%m-%d")
            prev_date_str = (meeting_date - timedelta(days=1)).strftime("%Y-%m-%d")
            raw_flights = search_flights_raw(origin, destination, date_str, meeting_date_obj=meeting_date)
            # For domestic: also fetch previous day's flights (needed for prev-night fallback)
            if not _is_intl:
                prev_raw = _search_flights_single_date(origin, destination, prev_date_str)
                raw_flights_with_prev = prev_raw + raw_flights
            else:
                raw_flights_with_prev = raw_flights
            st.session_state["raw_flights"] = raw_flights_with_prev

            filtered_flights = []
            if raw_flights_with_prev:
                if "email_package" in st.session_state: del st.session_state["email_package"]

                if best_per_airline and not _is_intl:
                    try:
                        eod        = datetime.combine(meeting_date, datetime.max.time())
                        cheap_list = process_flight_results(raw_flights, eod, origin, destination, meet_dt,
                                                           travel_count=t_count, dedupe=False)
                        if cheap_list:
                            cheap_list.sort(key=lambda x: (x["stops"], x["price"]))
                            st.session_state.cheapest_flight = cheap_list[0]; st.session_state.is_best_search = True
                    except Exception as e:
                        print(f"Cheapest calc error: {e}"); st.session_state.cheapest_flight = None; st.session_state.is_best_search = False
                else:
                    st.session_state.cheapest_flight = None; st.session_state.is_best_search = False

                try:
                    nearest_list = process_flight_results(raw_flights, target_arrival_dt, origin, destination,
                                                          meet_dt, travel_count=t_count, dedupe=False)
                    if nearest_list: nearest_list.sort(key=lambda x: x["price"])
                    st.session_state.nearest_flights = nearest_list[:3]
                except Exception as e:
                    print(f"Nearest calc error: {e}"); st.session_state.nearest_flights = []

                _use_dedupe    = (not best_per_airline) if not _is_intl else False
                filtered_flights = process_flight_results(
                    raw_flights_with_prev, target_arrival_dt, origin, destination,
                    meet_dt, t_count, dedupe=_use_dedupe, allow_prev_night=True,
                )
                display_flights  = []
                _has_prev_night  = filtered_flights and filtered_flights[0].get("category") == "prev_night"

                if filtered_flights:
                    if _has_prev_night:
                        # Adjust hotel check-in to previous night for the accommodation column
                        st.session_state["_prev_night_hotel_override"] = True
                        st.session_state["_prev_night_ci"] = meeting_date - timedelta(days=1)
                        _pn_cutoff = target_arrival_dt.strftime("%I:%M %p") if target_arrival_dt else ""
                        _pn_ci_s   = (meeting_date - timedelta(days=1)).strftime("%b %d")
                        _pn_co_s   = meeting_date.strftime("%b %d")
                        st.markdown(
                            f"<div style='background:#3b1f6a;border-left:3px solid #7c3aed;"
                            f"padding:6px 12px;border-radius:4px;font-size:0.82em;"
                            f"color:#ddd6fe;margin-bottom:4px;'>"
                            f"🌙 No same-day flights before {_pn_cutoff} — showing prev-night arrivals. "
                            f"Hotel: {_pn_ci_s} → {_pn_co_s}</div>",
                            unsafe_allow_html=True
                        )
                    else:
                        st.session_state["_prev_night_hotel_override"] = False
                        st.session_state.pop("_prev_night_ci", None)

                    with st.container(height=700):
                        if best_per_airline and not _is_intl:
                            for f_item in filtered_flights:
                                try:
                                    _arr_base = f_item.get("arr_time_raw") or _clean_time_str(f_item["arr_time"])
                                    _arr_date = datetime.strptime(f_item.get("arrival_date", date_str), "%Y-%m-%d").date()
                                    fa = datetime.combine(_arr_date, datetime.strptime(_arr_base, "%I:%M %p").time())
                                    f_item["time_gap"] = (target_arrival_dt - fa).total_seconds()
                                except: f_item["time_gap"] = float("inf")
                            st.session_state["display_flights"] = sorted(filtered_flights, key=lambda x: x["time_gap"])[:5]
                        else:
                            st.session_state["display_flights"] = filtered_flights

                        display_flights = st.session_state["display_flights"]
                        for f_item in display_flights:
                            hotel_reason = None
                            cat = f_item.get("category", "")
                            if cat == "prev_night":
                                hotel_reason = f"🌙 Prev-night arrival — Hotel check-in {(meeting_date - timedelta(days=1)).strftime('%b %d')}"
                            elif cat == "early": hotel_reason = "1 Day Early — Hotel Required"
                            elif f_item.get("needs_hotel"): hotel_reason = "Long layover (>2h)"
                            else:
                                try:
                                    _arr_clean = f_item.get("arr_time_raw") or _clean_time_str(f_item["arr_time"])
                                    arr_dt = datetime.combine(meeting_date, datetime.strptime(_arr_clean, "%I:%M %p").time())
                                    tg = (meet_dt - arr_dt).total_seconds() / 3600
                                    if tg > 18: hotel_reason = "Previous day arrival"
                                    elif arr_dt.hour >= 23 or arr_dt.hour <= 4: hotel_reason = "Late night landing"
                                    elif tg > 3: hotel_reason = f"Early arrival — {int(tg)}h before meeting"
                                except: pass

                            with st.container():
                                if best_per_airline and not _is_intl:
                                    gap = f_item.get("time_gap", 0)
                                    if gap != float("inf") and gap > 0:
                                        m_g = int(gap / 60)
                                        st.caption(f"⭐ **Best Arrival**: Lands {m_g//60}h {m_g%60}m before target" if m_g >= 60 else f"⭐ **Best Arrival**: Lands {m_g}m before target")
                                _f_img = f_item.get("thumbnail")
                                _f_has_img = isinstance(_f_img, str) and _f_img.startswith("http")
                                stops_int = int(f_item.get("stops", 0)) if str(f_item.get("stops", 0)).isdigit() else 0
                                fc1, fc2, fc3 = st.columns([1.2, 2.5, 1.5])
                                with fc1:
                                    st.image(_f_img if _f_has_img else "https://cdn-icons-png.flaticon.com/512/727/727142.png",
                                             width=70 if _f_has_img else 36)
                                    st.markdown(f"**{f_item['airline']}**")
                                    st.caption(f_item['flight_no'])
                                    if stops_int == 0:
                                        st.markdown("<span style='background:#0f9d58;color:white;padding:1px 7px;border-radius:10px;font-size:0.72em;'>Non-stop</span>", unsafe_allow_html=True)
                                    else:
                                        st.markdown(f"<span style='background:#e37400;color:white;padding:1px 7px;border-radius:10px;font-size:0.72em;'>{stops_int} Stop(s)</span>", unsafe_allow_html=True)
                                with fc2:
                                    st.markdown(f"**{f_item['dep_time']}** → **{f_item['arr_time']}**")
                                    st.caption(f"⏱ {f_item['duration']}")
                                    if hotel_reason: st.caption(hotel_reason)
                                # Category badge
                                _cat = f_item.get("category", "")
                                if is_international(origin, destination):
                                    if _cat == "early":
                                        st.markdown("<span style='background:#e37400;color:white;padding:2px 8px;border-radius:10px;font-size:0.75em;'>🏨 Arrives D-1 — Hotel Required</span>", unsafe_allow_html=True)
                                    elif _cat == "same_day":
                                        st.markdown("<span style='background:#0f9d58;color:white;padding:2px 8px;border-radius:10px;font-size:0.75em;'>✅ Arrives Meeting Day</span>", unsafe_allow_html=True)
                                elif _cat == "prev_night":
                                    st.markdown("<span style='background:#7c3aed;color:white;padding:2px 8px;border-radius:10px;font-size:0.75em;'>🌙 Previous Night — Hotel Required</span>", unsafe_allow_html=True)
                                with fc3:
                                    pl = f"₹{int(f_item['price']):,}" if isinstance(f_item['price'], (int, float)) else f"₹{f_item['price']}"
                                    if t_count > 1:
                                        st.markdown(f"<div style='text-align:right;font-weight:bold;font-size:1.3em;'>{pl}</div>"
                                                    f"<div style='text-align:right;color:grey;font-size:.78em;'>for {t_count} travelers</div>",
                                                    unsafe_allow_html=True)
                                    else:
                                        st.markdown(f"<div style='text-align:right;font-weight:bold;font-size:1.3em;'>{pl}</div>",
                                                    unsafe_allow_html=True)
                                try:
                                    # Use departure_date from the flight dict — already correct
                                    # for prev-night flights (set by process_flight_results)
                                    _f_dep_date   = f_item.get("departure_date") or date_str
                                    _airline_code = get_airline_code(f_item["airline"])
                                    _arr_for_url  = f_item.get("arr_time_raw") or _clean_time_str(f_item.get("arr_time", ""))
                                    gl_url = make_google_flights_link(
                                        origin, destination, _f_dep_date, f_item["airline"], t_count,
                                        dep_time_str=f_item["dep_time"],
                                        arr_time_str=_arr_for_url,
                                        price_per_person=int(f_item["price"] / max(int(t_count), 1)),
                                        max_stops=stops_int, _override_airline_code=_airline_code,
                                        round_trip=False, return_date=None,
                                    )
                                    sb1, sb2 = st.columns(2)
                                    with sb1: _select_button_flight(f_item)
                                    with sb2:
                                        st.markdown(
                                            f'<a href="{gl_url}" target="_blank" rel="noopener noreferrer" '
                                            f'style="display:block;width:100%;text-align:center;background:#1d4ed8;color:#fff;'
                                            f'text-decoration:none;padding:10px 16px;border-radius:8px;font-weight:600;font-size:14px;">'
                                            f'✈ Book on Google</a>',
                                            unsafe_allow_html=True,
                                        )
                                except Exception as e:
                                    st.error(f"Flight Button Error: {e}")
                            st.divider()
                else:
                    st.error(
                        f"❌ No flights found for this route and date. "
                        f"Try adjusting the meeting time or date."
                    )

        # ── Round trip: search + display return flights ────────────────────────
        if is_round_trip and meeting_end_dt:
            st.markdown("---")
            st.subheader("🔄 Return Flights")
            st.caption(f"**{destination}** → **{origin}**  ·  departing after {meeting_end_dt.strftime('%b %d, %I:%M %p')}")
            with st.spinner("🔄 Searching return flights..."):
                return_flights_list = search_return_flights(
                    origin=origin, destination=destination,
                    meeting_end_dt=meeting_end_dt, travel_count=t_count,
                )
            st.session_state["return_flights"] = return_flights_list

            if return_flights_list:
                any_needs_hotel = any(f.get("needs_hotel") for f in return_flights_list)
                all_same_day    = all(f.get("same_day_return") for f in return_flights_list)

                if not all_same_day and any_needs_hotel:
                    earliest_co = min(
                        (f["hotel_check_out"] for f in return_flights_list if f.get("needs_hotel")),
                        default=meeting_end_date + timedelta(days=1)
                    )
                    ci_dt  = meeting_date
                    co_dt  = earliest_co
                    ci_str = ci_dt.strftime("%Y-%m-%d")
                    co_str = co_dt.strftime("%Y-%m-%d")

                with st.container(height=400):
                    for idx, rf in enumerate(return_flights_list[:8]):
                        _render_return_flight_card(rf, origin, destination, t_count, idx)
            else:
                st.info(f"No return flights found departing after {meeting_end_dt.strftime('%I:%M %p')}.")

    # ── Hotel column ──────────────────────────────────────────────────────────
    if right_page and not _rooms_invalid:
        with right_page:
            st.subheader("🏨 Accommodation")
            try:
                mmt_info   = get_mmt_params_from_llm(meeting_location, destination)
                city_code  = mmt_info.get("mmt_city_code", destination)
                _rc        = int(st.session_state.get("rooms_count", 1))

                _skip_hotel_same_day_rt = False
                if is_round_trip and st.session_state.get("return_flights"):
                    rfs = st.session_state["return_flights"]
                    any_needs = any(f.get("needs_hotel") for f in rfs)
                    all_same  = all(f.get("same_day_return") for f in rfs)
                    if all_same or not any_needs:
                        ci_dt = meeting_date; co_dt = meeting_date + timedelta(days=1)
                        if is_round_trip and not is_international(origin, destination) and all_same:
                            _skip_hotel_same_day_rt = True
                    else:
                        earliest_co = min(
                            (f["hotel_check_out"] for f in rfs if f.get("needs_hotel")),
                            default=(meeting_end_date or meeting_date) + timedelta(days=1)
                        )
                        ci_dt = meeting_date; co_dt = earliest_co
                else:
                    ci_dt = meeting_date; co_dt = meeting_date + timedelta(days=1)

                # Prev-night fallback: adjust hotel check-in to the previous evening
                _has_prev_night = (
                    filtered_flights and
                    filtered_flights[0].get("category") == "prev_night"
                )
                if _has_prev_night or st.session_state.get("_prev_night_hotel_override"):
                    _prev_ci = st.session_state.get("_prev_night_ci") or (meeting_date - timedelta(days=1))
                    ci_dt = _prev_ci
                    co_dt = meeting_date  # check-out ON meeting day (prev-night stay)

                if _is_intl and filtered_flights and any(f.get("category") == "early" for f in filtered_flights):
                    ci_dt = meeting_date - timedelta(days=1)

                if _is_intl:
                    co_dt = meeting_end_date + timedelta(days=1)

                ci_str = ci_dt.strftime("%Y-%m-%d")
                co_str = co_dt.strftime("%Y-%m-%d")
                nights = max(1, (co_dt - ci_dt).days)

                if h_mode == "dual":
                    _total_guests = int(m_count) + int(e_count)
                    combined_url  = get_live_mmt_url(ci_dt, co_dt, _total_guests, meeting_location,
                                                     dest_iata=destination, rooms=_rc)
                    st.markdown(
                        f'<a href="{combined_url}" target="_blank" rel="noopener noreferrer" '
                        f'style="display:block;width:100%;text-align:center;background:#0f9d58;color:#fff;'
                        f'text-decoration:none;padding:10px 16px;border-radius:8px;font-weight:600;font-size:13px;margin-bottom:8px;">'
                        f'🔍 Search Hotels on MMT ({_total_guests} guests · {_rc} room(s) · {nights}n)</a>',
                        unsafe_allow_html=True,
                    )
                else:
                    ac = int(m_count) if m_count > 0 else int(e_count)
                    _hotel_url_single = get_live_mmt_url(ci_dt, co_dt, ac, meeting_location,
                                                         dest_iata=destination, rooms=_rc)
                    st.markdown(
                        f'<a href="{_hotel_url_single}" target="_blank" rel="noopener noreferrer" '
                        f'style="display:block;width:100%;text-align:center;background:#0f9d58;color:#fff;'
                        f'text-decoration:none;padding:10px 16px;border-radius:8px;font-weight:600;font-size:13px;margin-bottom:8px;">'
                        f'🔍 Search MMT hotels in {meeting_location} ({nights}n)</a>',
                        unsafe_allow_html=True,
                    )

                if is_round_trip and st.session_state.get("return_flights"):
                    rfs       = st.session_state["return_flights"]
                    all_same  = all(f.get("same_day_return") for f in rfs)
                    any_needs = any(f.get("needs_hotel") for f in rfs)
                    if not (all_same or not any_needs):
                        st.info(f"🏨 Hotel needed: **{ci_str}** → **{co_str}** ({nights} night(s))")
            except Exception as e:
                print(f"MMT Button Error: {e}")

            if not filtered_flights:
                st.caption("Search for flights to see accommodation suggestions.")
            else:
                with st.container(height=700):
                    any_needs_hotel = any(
                        f.get("needs_hotel") or f.get("category") == "early" or
                        (is_round_trip and st.session_state.get("return_flights") and
                         any(rf.get("needs_hotel") for rf in st.session_state["return_flights"]))
                        for f in filtered_flights
                    )

                    _rc  = int(st.session_state.get("rooms_count", 1))
                    _pax = int(m_count) + int(e_count) if h_mode == "dual" else (int(m_count) if int(m_count) > 0 else int(e_count))
                    expert_url = get_expert_mmt_url(meeting_location, destination, ci_dt, co_dt, _pax, _rc)
                    if h_mode == "dual":
                        with st.spinner("Finding hotels for Management & Employees..."):
                            dual = search_hotels_coordinated(meeting_location, ci_str, co_str,
                                                             int(m_count), int(e_count),
                                                             dest_airport_iata=destination, _cache_bust=2)
                        hotel_results_mgmt   = dual["mgmt"]["hotels"]
                        hotel_results_emp    = dual["emp"]["hotels"]
                        hotel_results_shared = dual.get("shared_hotels", [])
                        shared_from_expand   = dual.get("shared_from_expand", False)
                        preload_images_in_parallel(hotel_results_shared + hotel_results_mgmt + hotel_results_emp)
                        _render_dual_hotels(
                            hotel_results_mgmt, hotel_results_emp, hotel_results_shared,
                            ci_dt, co_dt, m_count, e_count, meeting_location, destination,
                            ci_str=ci_str, co_str=co_str,
                            city_code=get_mmt_params_from_llm(meeting_location, destination).get("mmt_city_code", destination),
                            check_in_date=ci_dt, shared_from_expand=shared_from_expand,
                            rooms=int(st.session_state.get("rooms_count", 1)),
                        )
                    elif h_mode == "single":
                        if int(exec_count) > 0: at, ac, nb, xb = "Executive", int(exec_count), 8000, 19000
                        elif int(m_count) > 0:  at, ac, nb, xb = "Management", int(m_count), 2700, 19000
                        else:                   at, ac, nb, xb = "Employee",   int(e_count), 1500,  3500
                        with st.spinner(f"Finding hotels (₹{nb}–{xb})..."):
                            hr  = search_hotels(meeting_location, ci_str, co_str, nb, xb, at,
                                                passenger_count=ac, meeting_location=meeting_location,
                                                dest_airport_iata=destination, _cache_bust=2)
                            hrs = hr["hotels"]
                        preload_images_in_parallel(hrs)
                        for idx, idx_h in enumerate(hrs):
                            _render_hotel_card(idx_h, ci_dt, co_dt, ac, meeting_location,
                                               get_airport_city(destination), group="single", index=idx,
                                               rooms=int(st.session_state.get("rooms_count", 1)),
                                               city_code=destination, property_token=idx_h.get("property_token", ""))
                        hotel_results_mgmt = hrs if at == "Management" else []
                        hotel_results_emp  = hrs if at == "Employee"   else []
                    else:
                        st.error("Please add at least one traveler.")

    st.session_state["display_flights"]      = display_flights
    st.session_state["hotel_results_mgmt"]   = hotel_results_mgmt
    st.session_state["hotel_results_emp"]    = hotel_results_emp
    st.session_state["hotel_results_shared"] = hotel_results_shared
    st.session_state["hotel_results"]        = hotel_results_shared + hotel_results_mgmt + hotel_results_emp
    st.session_state["last_m_count"]         = int(m_count)
    st.session_state["last_e_count"]         = int(e_count)
    # ── Persist MMT URLs so buttons survive navigation ─────────────────────────
    _mmt_rt = st.session_state.get("is_round_trip", False)
    _mmt_ed_safe = st.session_state.get("meeting_end_date_val")
    _mmt_rooms_save = int(st.session_state.get("rooms_count", 1))
    try:
        _saved_mmt_url = make_mmt_link(
            origin, destination, meeting_date, t_count,
            arrival_cutoff_dt=target_arrival_dt,
            is_round_trip=_mmt_rt, return_date=_mmt_ed_safe,
            rooms=_mmt_rooms_save,
        )
    except Exception:
        _saved_mmt_url = "https://www.makemytrip.com/flight/search"

    _hotel_co = (_mmt_ed_safe or meeting_date) + timedelta(days=1) if is_international(origin, destination) else meeting_date + timedelta(days=1)

    _saved_hotel_url = get_live_mmt_url(
        meeting_date, _hotel_co,
        t_count, meeting_location, dest_iata=destination,
        rooms=_mmt_rooms_save,
    ) if meeting_location else "https://www.makemytrip.com/hotels/"
    st.session_state["mmt_flight_url"]   = _saved_mmt_url
    st.session_state["mmt_hotel_url"]    = _saved_hotel_url
    st.session_state["mmt_flight_label"] = "✈️ Book on MMT" + (" (Round Trip)" if _mmt_rt else "")
    st.session_state["mmt_hotel_label"]  = f"🏨 Hotels on MMT — {meeting_location}" if meeting_location else "🏨 Hotels on MMT"


# ── Tracking Overlay ──────────────────────────────────────────────────────────

elif st.session_state.get("view") == "tracking" or st.session_state.get("show_tracking_overlay"):
    _render_tracking_overlay()


# ── Persistent Display ────────────────────────────────────────────────────────

elif not search_triggered and st.session_state.get("display_flights") and st.session_state.get("view") == "results":
    meet_dt   = datetime.combine(meeting_date, meeting_time)
    date_str  = meeting_date.strftime("%Y-%m-%d")
    target_arrival_dt, logistics = calculate_dynamic_cutoff(destination, meeting_location, meet_dt)

    _is_rt    = st.session_state.get("is_round_trip", False)
    _end_date = st.session_state.get("meeting_end_date_val")

    _render_search_header(origin, destination, origin_input, dest_input, meeting_date,
                          target_arrival_dt, is_round_trip=_is_rt, meeting_end_date=_end_date)

    display_flights      = st.session_state["display_flights"]
    hotel_results_mgmt   = st.session_state.get("hotel_results_mgmt",   [])
    hotel_results_emp    = st.session_state.get("hotel_results_emp",    [])
    hotel_results_shared = st.session_state.get("hotel_results_shared", [])
    return_flights       = st.session_state.get("return_flights",        [])

    if st.session_state.pop("_trigger_hotel_refresh", False):
        _loc    = st.session_state.get("last_meeting_location", meeting_location)
        _ci_str = meeting_date.strftime("%Y-%m-%d")
        _co_str = (meeting_date + timedelta(days=1)).strftime("%Y-%m-%d")
        h_mode  = _hotel_mode(int(m_count), int(e_count), int(exec_count))
        hotel_results_mgmt = []; hotel_results_emp = []; hotel_results_shared = []
        if h_mode == "dual":
            with st.spinner("🔄 Refreshing hotels..."):
                dual = search_hotels_coordinated(_loc, _ci_str, _co_str, int(m_count), int(e_count),
                                                 dest_airport_iata=destination, _cache_bust=2)
            hotel_results_mgmt   = dual["mgmt"]["hotels"]
            hotel_results_emp    = dual.get("emp", {}).get("hotels", [])
            hotel_results_shared = dual.get("shared_hotels", [])
        elif h_mode == "single":
            if int(exec_count) > 0: at, ac, nb, xb = "Executive", int(exec_count), 10000, 19000
            elif int(m_count) > 0:  at, ac, nb, xb = "Management", int(m_count), 3500, 19000
            else:                   at, ac, nb, xb = "Employee",   int(e_count), 2500, 4500
            with st.spinner(f"🔄 Refreshing hotels..."):
                hr  = search_hotels(_loc, _ci_str, _co_str, nb, xb, at, passenger_count=ac,
                                    meeting_location=_loc, dest_airport_iata=destination, _cache_bust=2)
                hrs = hr["hotels"]
            if at == "Management": hotel_results_mgmt = hrs
            else:                  hotel_results_emp  = hrs
        raw_flights = st.session_state.get("raw_flights", [])
        if raw_flights and t_count > 0:
            with st.spinner("🔄 Updating flight prices..."):
                target_arrival_dt, _ = calculate_dynamic_cutoff(destination, meeting_location, meet_dt)
                updated = process_flight_results(raw_flights, target_arrival_dt, origin, destination,
                                                 meet_dt, t_count, dedupe=not best_per_airline)
                display_flights = updated
                st.session_state["display_flights"] = display_flights
        st.session_state["hotel_results_mgmt"]   = hotel_results_mgmt
        st.session_state["hotel_results_emp"]    = hotel_results_emp
        st.session_state["hotel_results_shared"] = hotel_results_shared
        st.session_state["last_m_count"] = int(m_count)
        st.session_state["last_e_count"] = int(e_count)
        st.rerun()

    disp_m_count   = int(st.session_state.get("last_m_count", 0))
    disp_e_count   = int(st.session_state.get("last_e_count", 0))
    disp_exec_count = int(st.session_state.get("exec_count", 0))
    disp_t_count   = disp_m_count + disp_e_count + disp_exec_count  # include executives

    if best_per_airline:
        left_page = st.columns([1])[0]; right_page = None
    else:
        left_page, right_page = st.columns([1.5, 1], gap="large")

    with left_page:
        st.subheader("🛫 Outbound Flights")
        st.caption("Check flights to include them in the approval email.")
        # Use persisted MMT URL from session state — survives navigation to other apps and back
        _p_is_rt      = st.session_state.get("is_round_trip", False)
        _p_end_dt     = st.session_state.get("meeting_end_date_val")
        _p_mmt_flight = st.session_state.get("mmt_flight_url") or make_mmt_link(
            origin, destination, meeting_date, disp_t_count,
            arrival_cutoff_dt=target_arrival_dt,
            is_round_trip=_p_is_rt, return_date=_p_end_dt,
            rooms=int(st.session_state.get("rooms_count", 1)),
        )
        _p_mmt_label  = st.session_state.get("mmt_flight_label", "✈️ Book on MMT")
        # Use <a> tag — st.link_button breaks MMT URLs in Chrome (strips & params → opens homepage)
        st.markdown(
            f'<a href="{_p_mmt_flight}" target="_blank" rel="noopener noreferrer" '
            f'style="display:block;width:100%;text-align:center;background:#1d4ed8;color:#fff;'
            f'text-decoration:none;padding:10px 16px;border-radius:8px;font-weight:600;font-size:14px;margin-bottom:8px;">'
            f'{_p_mmt_label}</a>',
            unsafe_allow_html=True,
        )
        if logistics and target_arrival_dt:
            d_mins   = logistics.get("drive_mins", 0)
            kms      = logistics.get("distance_km", 0)
            drive_str = f"{d_mins//60}h {d_mins%60}m" if d_mins >= 60 else f"{d_mins}m"
            cutoff_str = target_arrival_dt.strftime("%I:%M %p")
            st.info(
                f"ℹ️ **Must land by {cutoff_str}** — "
                f"**{drive_str}** drive ({kms} km) + **1h** transit buffer "
                f"before {meet_dt.strftime('%I:%M %p')} meeting."
            )

        if display_flights:
            with st.container(height=700):
                for f_item in display_flights:
                    hotel_reason = None
                    try:
                        # Always use arr_time_raw for time parsing (no newlines)
                        _arr_clean = f_item.get("arr_time_raw") or _clean_time_str(f_item.get("arr_time", ""))
                        ad = datetime.combine(meeting_date, datetime.strptime(_arr_clean, "%I:%M %p").time())
                        tg = (meet_dt - ad).total_seconds() / 3600
                        if tg > 18: hotel_reason = "Previous day arrival"
                        elif ad.hour >= 23 or ad.hour <= 5: hotel_reason = "Late night landing"
                        elif tg > 8: hotel_reason = f"Long wait ({int(tg)}h)"
                    except: pass
                    with st.container():
                        _img = f_item.get("thumbnail")
                        _has_img = isinstance(_img, str) and _img.startswith("http")
                        si = int(f_item.get("stops", 0)) if str(f_item.get("stops", 0)).isdigit() else 0
                        fc_left, fc_mid, fc_right = st.columns([1.2, 2.5, 1.5])
                        with fc_left:
                            st.image(_img if _has_img else "https://cdn-icons-png.flaticon.com/512/727/727142.png",
                                     width=70 if _has_img else 36)
                            st.markdown(f"**{f_item['airline']}**")
                            st.caption(f_item['flight_no'])
                            if si == 0:
                                st.markdown("<span style='background:#0f9d58;color:white;padding:1px 7px;border-radius:10px;font-size:0.72em;'>Non-stop</span>", unsafe_allow_html=True)
                            else:
                                st.markdown(f"<span style='background:#e37400;color:white;padding:1px 7px;border-radius:10px;font-size:0.72em;'>{si} Stop(s)</span>", unsafe_allow_html=True)
                        with fc_mid:
                            st.markdown(f"**{f_item['dep_time']}** → **{f_item['arr_time']}**")
                            st.caption(f"⏱ {f_item['duration']}")
                            if hotel_reason: st.caption(f"🏨 {hotel_reason}")
                            # International arrival category badge
                            if is_international(origin, destination):
                                _cat = f_item.get("category", "")
                                if _cat == "early":
                                    st.markdown(
                                        "<span style='background:#e37400;color:white;"
                                        "padding:2px 8px;border-radius:10px;font-size:0.75em;'>"
                                        "🏨 Arrives D-1 — Hotel Required</span>",
                                        unsafe_allow_html=True)
                                elif _cat == "same_day":
                                    st.markdown(
                                        "<span style='background:#0f9d58;color:white;"
                                        "padding:2px 8px;border-radius:10px;font-size:0.75em;'>"
                                        "✅ Arrives Meeting Day</span>",
                                        unsafe_allow_html=True)
                        with fc_right:
                            pl = f"₹{f_item['price']:,}" if isinstance(f_item['price'], (int, float)) else f"₹{f_item['price']}"
                            if disp_t_count > 1:
                                st.markdown(f"<div style='text-align:right;font-weight:bold;font-size:1.3em;'>{pl}</div>"
                                            f"<div style='text-align:right;color:grey;font-size:.78em;'>for {disp_t_count} travelers</div>",
                                            unsafe_allow_html=True)
                            else:
                                st.markdown(f"<div style='text-align:right;font-weight:bold;font-size:1.3em;'>{pl}</div>",
                                            unsafe_allow_html=True)
                        # Buttons on their own full-width row — never squished
                        _airline_code2 = get_airline_code(f_item["airline"])
                        _arr_for_url = f_item.get("arr_time_raw") or _clean_time_str(f_item.get("arr_time", ""))
                        # Use departure_date from flight dict — correct for prev-night flights
                        _f_dep_date_persist = f_item.get("departure_date") or date_str
                        gl = make_google_flights_link(
                            origin, destination, _f_dep_date_persist, f_item["airline"], disp_t_count,
                            dep_time_str=f_item["dep_time"],
                            arr_time_str=_arr_for_url,
                            price_per_person=int(f_item["price"] / max(int(disp_t_count), 1)),
                            _override_airline_code=_airline_code2,
                            round_trip=False, return_date=None,
                        )
                        btn_c1, btn_c2 = st.columns(2)
                        with btn_c1: _select_button_flight(f_item)
                        with btn_c2:
                            st.markdown(
                                f'<a href="{gl}" target="_blank" rel="noopener noreferrer" '
                                f'style="display:block;width:100%;text-align:center;background:#1d4ed8;color:#fff;'
                                f'text-decoration:none;padding:10px 16px;border-radius:8px;font-weight:600;font-size:14px;">'
                                f'✈ Book on Google</a>',
                                unsafe_allow_html=True,
                            )
                    st.divider()

        # Return flights (persistent display)
        if _is_rt and return_flights:
            st.markdown("---")
            st.subheader("🔄 Return Flights")
            _end_dt_disp = st.session_state.get("meeting_end_date_val")
            _end_tm_disp = st.session_state.get("meeting_end_time_val")
            if _end_dt_disp and _end_tm_disp:
                st.caption(f"**{destination}** → **{origin}**  ·  after {datetime.combine(_end_dt_disp, _end_tm_disp).strftime('%b %d, %I:%M %p')}")
            with st.container(height=400):
                for idx, rf in enumerate(return_flights[:8]):
                    _render_return_flight_card(rf, origin, destination, disp_t_count, idx)
        elif _is_rt:
            st.info("No return flights found.")

    has_any_hotels = hotel_results_mgmt or hotel_results_emp or hotel_results_shared
    if right_page and not best_per_airline and has_any_hotels and not _rooms_invalid:
        with right_page:
            st.subheader("🏨 Accommodation"); st.caption("Check hotels to include in the approval email.")
            h_mode = _hotel_mode(disp_m_count, disp_e_count, int(st.session_state.get("exec_count", 0)))
            _rc    = int(st.session_state.get("rooms_count", 1))

            # MMT hotel button — recompute with current rooms_count so it's always fresh
            # Falls back to saved URL if meeting_location is not set
            _live_hotel_url = None
            if meeting_location:
                try:
                    _hotel_co = (_end_dt_disp or meeting_date) + timedelta(days=1) if is_international(origin, destination) else meeting_date + timedelta(days=1)
                    _live_hotel_url = get_live_mmt_url(
                        meeting_date, _hotel_co,
                        max(1, disp_t_count), meeting_location,
                        dest_iata=destination,
                        rooms=int(st.session_state.get("rooms_count", 1)),
                    )
                except Exception:
                    _live_hotel_url = None
            _p_hotel_url   = _live_hotel_url or st.session_state.get("mmt_hotel_url", "https://www.makemytrip.com/hotels/")
            _p_hotel_label = f"🏨 Hotels on MMT — {meeting_location}" if meeting_location else "🏨 Hotels on MMT"
            st.markdown(
                f'<a href="{_p_hotel_url}" target="_blank" rel="noopener noreferrer" '
                f'style="display:block;width:100%;text-align:center;background:#0f9d58;color:#fff;'
                f'text-decoration:none;padding:10px 16px;border-radius:8px;font-weight:600;font-size:13px;margin-bottom:8px;">'
                f'{_p_hotel_label}</a>',
                unsafe_allow_html=True,
            )

            _skip_hotel_rt_persistent = False
            if _is_rt and return_flights:
                any_needs = any(f.get("needs_hotel") for f in return_flights)
                all_same  = all(f.get("same_day_return") for f in return_flights)
                if all_same or not any_needs:
                    ci_dt = meeting_date; co_dt = meeting_date + timedelta(days=1)
                    if not is_international(origin, destination) and all_same:
                        _skip_hotel_rt_persistent = True
                else:
                    earliest_co = min(
                        (f["hotel_check_out"] for f in return_flights if f.get("needs_hotel")),
                        default=meeting_date + timedelta(days=2)
                    )
                    ci_dt = meeting_date; co_dt = earliest_co
            else:
                ci_dt = meeting_date; co_dt = meeting_date + timedelta(days=1)

            if _skip_hotel_rt_persistent:
                st.success("✅ Same-day return — no hotel required!")
            else:
                nights = (co_dt - ci_dt).days
                if nights > 1:
                    st.info(f"🏨 Hotel: {ci_dt.strftime('%b %d')} → {co_dt.strftime('%b %d')} ({nights} nights)")
                preload_images_in_parallel(hotel_results_shared + hotel_results_mgmt + hotel_results_emp)
            if not _skip_hotel_rt_persistent:
              with st.container(height=700):
                if h_mode == "dual":
                    _ci_str = ci_dt.strftime("%Y-%m-%d"); _co_str = co_dt.strftime("%Y-%m-%d")
                    _render_dual_hotels(
                        hotel_results_mgmt, hotel_results_emp, hotel_results_shared,
                        ci_dt, co_dt, disp_m_count, disp_e_count, meeting_location, destination,
                        ci_str=_ci_str, co_str=_co_str,
                        city_code=get_mmt_params_from_llm(meeting_location, destination).get("mmt_city_code", destination),
                        check_in_date=ci_dt, rooms=_rc,
                    )
                elif h_mode == "single":
                    ac = disp_m_count if disp_m_count > 0 else disp_e_count
                    for idx, h_item in enumerate(hotel_results_mgmt or hotel_results_emp):
                        _render_hotel_card(h_item, ci_dt, co_dt, ac, meeting_location,
                                           get_airport_city(destination), group="single", index=idx,
                                           rooms=_rc, city_code=destination,
                                           property_token=h_item.get("property_token", ""))


# ── Verify View ───────────────────────────────────────────────────────────────

elif st.session_state.get("view") == "verify":
    _render_verify_view()


# ── Welcome Screen ────────────────────────────────────────────────────────────

if (st.session_state.get("show_welcome")
        and not st.session_state.get("display_flights")
        and st.session_state.get("view") == "results"):
    st.markdown("""
        <div style='display:flex;flex-direction:column;justify-content:center;align-items:center;height:75vh;text-align:center;'>
            <div style='font-size:6em;margin-bottom:10px;'>✈️</div>
            <h1 style='color:#FFFFFF;font-size:3.2em;font-weight:700;margin:0 0 15px 0;letter-spacing:-0.5px;'>Welcome to Si2 Travel Planner</h1>
            <p style='color:#D1D5DB;font-size:1.4em;max-width:600px;margin:0 auto;line-height:1.6;'>
                Enter your travel details in the sidebar and click <br>
                <span style='color:#FFFFFF;font-weight:600;'>🔍 Find Best Flights</span> to get started.
            </p>
        </div>""", unsafe_allow_html=True)


# ── Email Preview ─────────────────────────────────────────────────────────────

pkg = st.session_state.get("email_package")
if pkg and st.session_state.get("view") == "email":
    back_col, title_col = st.columns([1, 5])
    with back_col:   st.button("← Back to Results", key="back_to_results_email_btn", width='stretch', on_click=_handle_back_from_email)
    with title_col:  st.markdown("### 📧 Approval Email Draft")
    col_a, col_b = st.columns([3, 1])
    with col_a: st.caption("👇 Preview below. Download .eml → double-click → Outlook → Send.")
    with col_b:
        if st.button("🚀 Send Email via Entra", type="primary", width='stretch'):
            with st.spinner("Authenticating and sending..."):
                success, msg = send_via_graph_api(pkg["subject"], pkg["html_body"],
                                                  "krishna.kotecha@si2tech.com",
                                                  embedded_images=pkg.get("embedded_images", []))
                if success: st.success("✅ " + msg); st.balloons()
                else: st.error("❌ " + msg)
    n_sf = len(st.session_state["selected_flight_keys"]); n_sh = len(st.session_state["selected_hotel_keys"])
    if n_sf > 0 or n_sh > 0:
        st.info(f"📋 Email generated with **{n_sf} flight(s)** and **{n_sh} hotel(s)** selected.")
    else:
        st.info("📋 No items selected — email shows all results.")
    st.divider()
    components.html(pkg["html_preview"], height=1000, scrolling=True)